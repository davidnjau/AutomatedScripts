#!/usr/bin/env python3
"""
ardhisasa_bot.py
================
Telegram bot for Ardhisasa Valuation Officer Assignment.

Conversation flow:
  /assign  (or tap "📋 New Assignment")
    → ask reference numbers
    → [saved valuers exist] pick saved valuer OR search new
      [no saved valuers]    ask valuer name to search
    → choose credential profile
      [valid cached token]  skip login entirely → jump to confirm / valuer list
      [no cached token]     trigger login (OTP sent to device)
                            → ask user to forward OTP here
                            → verify OTP & cache tokens
    → show matching valuers (inline keyboard)   [skipped when saved valuer used]
    → user selects valuer
    → confirm selection
    → run assignments
    → show results  (valuer auto-saved for future use)
"""

import base64
import io
import json
import logging
import asyncio
import os
import re
import signal
import smtplib
import subprocess
import sys
import threading
import time
import uuid
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders as _email_encoders
from dataclasses import dataclass, field
from enum import Enum, auto
from concurrent.futures import ThreadPoolExecutor, as_completed as _futures_as_completed
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Tuple

import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pytesseract
import requests
from dotenv import load_dotenv
from PIL import Image
from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
)
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

from ardhisasa_auth import (
    AUTH_BASE_URL,
    PUBLIC_CREDENTIALS,
    STAFF_CREDENTIALS_ICT,
    STAFF_CREDENTIALS_SUPPORT,
    STAFF_CREDENTIALS_VALUER,
    AuthTokens,
    build_session,
)

# ──────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger("ardhisasa.bot")

load_dotenv()

BOT_TOKEN         = os.environ["TELEGRAM_BOT_TOKEN"]
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
ALLOWED_IDS = set(
    int(x.strip())
    for x in os.getenv("ALLOWED_TELEGRAM_IDS", "").split(",")
    if x.strip()
)

# SMTP config for Auto Fetch email notifications (all optional)
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com").strip()
SMTP_PORT = int(os.getenv("SMTP_PORT", "587").strip())
SMTP_USER = os.getenv("SMTP_USER", "").strip()
SMTP_PASS = os.getenv("SMTP_PASS", "").strip()

BASE_URL = "https://ardhisasa-api.lands.go.ke"

# ──────────────────────────────────────────────────────────
# Persistent storage
# ──────────────────────────────────────────────────────────
DATA_DIR           = os.path.join(os.path.dirname(__file__), "data")
SAVED_VALUERS_FILE      = os.path.join(DATA_DIR, "saved_valuers.json")
SAVED_TOKENS_FILE       = os.path.join(DATA_DIR, "saved_tokens.json")
SAVED_ASSIGNMENTS_FILE  = os.path.join(DATA_DIR, "saved_assignments.json")
SAVED_TASK_BATCHES_FILE = os.path.join(DATA_DIR, "saved_task_batches.json")
SAVED_SCHEDULES_FILE    = os.path.join(DATA_DIR, "saved_schedules.json")
SAVED_DLV_BATCH_FILE          = os.path.join(DATA_DIR, "saved_dlv_batch.json")
SAVED_BULK_EXPORT_SCHED_FILE    = os.path.join(DATA_DIR, "saved_bulk_export_schedule.json")
SAVED_BULK_EXPORT_PARTIAL_FILE  = os.path.join(DATA_DIR, "saved_bulk_export_partial.json")
SAVED_AUTO_FETCH_FILE   = os.path.join(DATA_DIR, "saved_auto_fetch.json")
SAVED_AF_RESULTS_FILE   = os.path.join(DATA_DIR, "saved_af_results.json")

# base64('{"active_role":"DLV"}') — required cparams header for DLV task endpoints
CPARAMS_DLV          = base64.b64encode(b'{"active_role":"DLV"}').decode()
CPARAMS_ASSESSOR     = base64.b64encode(b'{"active_role":"ASSESSOR_OF_STAMP_DUTY"}').decode()
CPARAMS_VALUER_ROLE  = base64.b64encode(b'{"active_role":"VALUER"}').decode()
CPARAMS_SUPPORT      = base64.b64encode(b'{"active_role":"SUPPORT"}').decode()

DAEMON_SCRIPT = os.path.join(os.path.dirname(__file__), "token_refresh_daemon.py")
DAEMON_PID_FILE = os.path.join(DATA_DIR, "daemon.pid")
DAEMON_LOG_FILE = os.path.join(DATA_DIR, "daemon.log")


def _ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)


# ── Valuers ───────────────────────────────────────────────

def load_saved_valuers() -> List[Dict]:
    try:
        with open(SAVED_VALUERS_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def persist_valuer(name: str, uid: str, account_number: str):
    _ensure_data_dir()
    valuers = load_saved_valuers()
    if any(x["uid"] == uid for x in valuers):
        return  # already saved
    valuers.append({"name": name, "uid": uid, "account_number": account_number})
    with open(SAVED_VALUERS_FILE, "w") as f:
        json.dump(valuers, f, indent=2)
    logger.info("Saved valuer %s (uid=%s)", name, uid)


# ── Assignments ───────────────────────────────────────────

def load_saved_assignments() -> Dict:
    """Return dict mapping reference_number → {valuer_name, valuer_uid, assigned_at}."""
    try:
        with open(SAVED_ASSIGNMENTS_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def persist_assignment(ref: str, valuer_name: str, valuer_uid: str):
    _ensure_data_dir()
    assignments = load_saved_assignments()
    assignments[ref] = {
        "valuer_name": valuer_name,
        "valuer_uid":  valuer_uid,
        "assigned_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    with open(SAVED_ASSIGNMENTS_FILE, "w") as f:
        json.dump(assignments, f, indent=2)
    logger.info("Saved assignment %s → %s", ref, valuer_name)


# ── Tokens ────────────────────────────────────────────────

def _jwt_exp(token: str) -> Optional[float]:
    """Decode the JWT payload and return the `exp` claim, or None on failure."""
    try:
        payload = token.split(".")[1]
        payload += "=" * (4 - len(payload) % 4)
        data = json.loads(base64.urlsafe_b64decode(payload))
        return float(data["exp"])
    except Exception:
        return None


def _load_tokens_raw() -> Dict:
    try:
        with open(SAVED_TOKENS_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def persist_tokens(cred_type: str, access_token: str, jwt: str, refresh_token: str = ""):
    _ensure_data_dir()
    tokens = _load_tokens_raw()
    exp = _jwt_exp(jwt) or (time.time() + 3600)
    entry: Dict = {
        "access_token": access_token,
        "jwt":          jwt,
        "expires_at":   exp,
    }
    if refresh_token:
        entry["refresh_token"] = refresh_token
    elif tokens.get(cred_type, {}).get("refresh_token"):
        # Preserve existing refresh_token if a new one wasn't returned
        entry["refresh_token"] = tokens[cred_type]["refresh_token"]
    tokens[cred_type] = entry
    with open(SAVED_TOKENS_FILE, "w") as f:
        json.dump(tokens, f, indent=2)
    logger.info("Cached tokens for cred_type=%s (exp=%s) → %s", cred_type, exp, SAVED_TOKENS_FILE)



def get_valid_tokens(cred_type: str) -> Optional[AuthTokens]:
    """Return cached AuthTokens if still valid (5 min buffer), else None."""
    entry = _load_tokens_raw().get(cred_type)
    if not entry:
        return None
    if entry.get("expires_at", 0) < time.time() + 300:
        logger.info("Cached tokens for %s are expired.", cred_type)
        return None
    return AuthTokens(access_token=entry["access_token"], jwt=entry["jwt"])


# ──────────────────────────────────────────────────────────
# States
# ──────────────────────────────────────────────────────────
class S(Enum):
    INPUT_METHOD       = auto()   # choose text or photo input
    REF_NUMBERS        = auto()   # typing refs manually
    RECV_PHOTOS        = auto()   # receiving photo(s) for OCR
    CONFIRM_REFS       = auto()   # review extracted refs before proceeding
    REASSIGN_CONFIRM   = auto()   # some refs already assigned — ask what to do
    PICK_VALUER_SOURCE = auto()   # choose saved valuer or search new
    VALUER_NAME        = auto()   # enter name when searching new
    CHOOSE_CRED        = auto()
    WAIT_OTP           = auto()
    SELECT_VALUER      = auto()
    CONFIRM            = auto()


# ──────────────────────────────────────────────────────────
# Per-user session data
# ──────────────────────────────────────────────────────────
@dataclass
class Session:
    refs:             List[str] = field(default_factory=list)
    extracted_refs:   List[str] = field(default_factory=list)   # OCR-extracted refs awaiting confirmation
    already_assigned: List[Dict] = field(default_factory=list)  # [{ref, valuer_name, assigned_at}]
    valuer_name:      str = ""
    cred_type:        str = "publicuser"
    session:          Optional[requests.Session] = None
    tokens:           Optional[AuthTokens] = None
    valuers:          List[Dict] = field(default_factory=list)
    selected_idx:     Optional[int] = None
    saved_valuer:     Optional[Dict] = None   # {"name", "uid", "account_number"}


# ──────────────────────────────────────────────────────────
# States — Receive Tasks conversation
# ──────────────────────────────────────────────────────────
class RS(Enum):
    PICK_STAFF_SOURCE = auto()   # choose saved valuer or search new
    STAFF_NAME        = auto()
    SELECT_STAFF      = auto()
    CHOOSE_CRED       = auto()
    WAIT_OTP          = auto()
    TASK_TYPE         = auto()   # choose Stamp Duty vs County Stamp Duty (when staff has both)
    TASK_COUNT        = auto()
    AMOUNT_RANGE      = auto()   # shows Enter / Skip buttons
    AMOUNT_TEXT       = auto()   # text input for min-max after choosing Enter
    SCHEDULE_CHOICE   = auto()
    SCHEDULE_INTERVAL = auto()
    RT_CONFIRM        = auto()


# ──────────────────────────────────────────────────────────
# Per-user session — Receive Tasks
# ──────────────────────────────────────────────────────────
@dataclass
class RTSession:
    staff_name:               str = ""
    staff_results:            List[Dict] = field(default_factory=list)
    staff_data:               Optional[Dict] = None
    saved_valuer:             Optional[Dict] = None   # {"name", "uid", "account_number"} — pre-selected
    cred_type:                str = "publicuser"
    session:                  Optional[object] = None   # requests.Session
    tokens:                   Optional[AuthTokens] = None
    task_count:               int = 0
    amount_min:               Optional[float] = None
    amount_max:               Optional[float] = None
    task_type:                str = ""   # "STAMP_DUTY" or "COUNTY_STAMP_DUTY"
    staff_registry:           str = ""
    staff_county:             str = ""
    matched_tasks:            List[Dict] = field(default_factory=list)
    schedule_interval_minutes: Optional[int] = None


# ──────────────────────────────────────────────────────────
# States — Auto Fetch schedule conversation
# ──────────────────────────────────────────────────────────
class AF(Enum):
    INTERVAL    = auto()   # choose how often to run
    DAYS_BACK   = auto()   # choose days back
    COUNTY      = auto()   # county filter
    REGISTRY    = auto()   # registry filter
    AMOUNT      = auto()   # pick amount range button
    AMOUNT_TEXT = auto()   # custom amount text entry
    SECTIONAL   = auto()   # exclude / only / all sectional
    EMAIL       = auto()   # optional recipient email address


# ──────────────────────────────────────────────────────────
# States — DLV Batch conversation
# ──────────────────────────────────────────────────────────
class DB(Enum):
    INPUT_BATCH   = auto()   # waiting for batch text
    CONFIRM_BATCH = auto()   # waiting for confirm/cancel


@dataclass
class DBSession:
    groups: List[Dict] = field(default_factory=list)
    # groups: [{refs, valuer_name, valuer_uid, valuer_acct, status}]


def _get_db_sess(ctx: ContextTypes.DEFAULT_TYPE) -> DBSession:
    if "db_session" not in ctx.user_data:
        ctx.user_data["db_session"] = DBSession()
    return ctx.user_data["db_session"]


# ──────────────────────────────────────────────────────────
# States — Refresh Auth conversation
# ──────────────────────────────────────────────────────────
class AS(Enum):
    CHOOSE_CRED   = auto()
    FORCE_CONFIRM = auto()
    WAIT_OTP      = auto()


@dataclass
class AuthSession:
    cred_type:    str = ""
    http_session: Optional[requests.Session] = None


def _get_auth_sess(ctx: ContextTypes.DEFAULT_TYPE) -> AuthSession:
    if "auth_session" not in ctx.user_data:
        ctx.user_data["auth_session"] = AuthSession()
    return ctx.user_data["auth_session"]


# ──────────────────────────────────────────────────────────
# States — Fetch Tasks conversation
# ──────────────────────────────────────────────────────────
class FT(Enum):
    CHOOSE_CRED      = auto()
    WAIT_OTP         = auto()
    DAYS_BACK        = auto()   # inline-button preset OR free text
    COUNTY_FILTER    = auto()   # county button picker
    REGISTRY_FILTER  = auto()   # registry button picker
    AMOUNT_FILTER    = auto()   # 4-button amount range picker
    AMOUNT_TEXT      = auto()   # free-text custom min/max amount
    SECTIONAL_FILTER = auto()   # exclude / only / all sectional


@dataclass
class FTSession:
    cred_type:       str = "staff2"   # default to Support Reg (has SUPPORT role)
    http_session:    Optional[requests.Session] = None
    tokens:          Optional[AuthTokens] = None
    days_back:       int = 5
    tasks:           List[Dict] = field(default_factory=list)
    stats:           Dict = field(default_factory=dict)
    county_filter:   str = ""           # "nairobi" or "" (all)
    registry_filter: str = ""           # "central", "nairobi", or "" (all)
    amount_min:       Optional[float] = None
    amount_max:       Optional[float] = None
    sectional_filter: str = "exclude"   # "exclude" | "only" | "all"


def _get_ft_sess(ctx: ContextTypes.DEFAULT_TYPE) -> FTSession:
    if "ft_session" not in ctx.user_data:
        ctx.user_data["ft_session"] = FTSession()
    return ctx.user_data["ft_session"]


# ──────────────────────────────────────────────────────────
# States — Job Distribution conversation
# ──────────────────────────────────────────────────────────
class JD(Enum):
    PICK_CRED = auto()   # select account with valid token
    CONFIRM   = auto()   # confirm → kick off background analysis


# States — Bulk Export conversation
# ──────────────────────────────────────────────────────────
class BE(Enum):
    COUNTY    = auto()   # pick county
    EMAIL     = auto()   # ask for recipient email address
    SCHEDULE  = auto()   # pick repeat interval
    PICK_CRED = auto()   # choose which cached credential to run as
    CONFIRM   = auto()   # confirm → kick off background export


# County → list of registry names as they appear in the API response
_BE_COUNTY_REGISTRIES: Dict[str, List[str]] = {
    "NAIROBI":  ["NAIROBI", "CENTRAL"],
    "KIAMBU":   ["KIAMBU", "LIMURU", "THIKA", "RUIRU", "GITHUNGURI"],
    "MURANGA":  ["MURANGA", "KANDARA", "MARAGUA", "KANGEMA"],
    "MOMBASA":  ["MOMBASA", "COAST"],
}

_BE_COUNTY_LABELS: Dict[str, str] = {
    "NAIROBI": "🌆 Nairobi",
    "KIAMBU":  "🏙 Kiambu",
    "MURANGA": "🏡 Murang'a",
    "MOMBASA": "🌊 Mombasa",
}


_BE_SCHEDULE_OPTIONS: List[tuple] = [
    ("Every Day",      86_400),
    ("Every Week",     604_800),
    ("Bi-Monthly",     1_209_600),   # every 2 weeks
    ("Monthly",        2_592_000),   # 30 days
    ("Every 2 Months", 5_184_000),   # 60 days
    ("Run Once",       0),
]

# Per-chat export status tracker — keyed by chat_id
# Each entry: {phase, started_at, total, pages_done, total_pages,
#              details_done, details_total, errors, completed_at, rows, error_msg}
_BE_STATUS: Dict[int, dict] = {}

# Per-chat job distribution status tracker
_JD_STATUS: Dict[int, dict] = {}


@dataclass
class BESession:
    county:           str = ""
    registries:       List[str] = field(default_factory=list)
    email:            str = ""
    schedule_seconds: int = 0   # 0 = run once
    cred_type:        str = ""


def _get_be_sess(ctx: ContextTypes.DEFAULT_TYPE) -> BESession:
    if "be_session" not in ctx.user_data:
        ctx.user_data["be_session"] = BESession()
    return ctx.user_data["be_session"]


CRED_MAP = {
    "publicuser":   PUBLIC_CREDENTIALS,
    "staff":        STAFF_CREDENTIALS_ICT,
    "staff2":       STAFF_CREDENTIALS_SUPPORT,
    "staff_valuer": STAFF_CREDENTIALS_VALUER,
}

CRED_LABELS = {
    "publicuser":   "👤 Public User",
    "staff":        "🏢 ICT",
    "staff2":       "🏢 Support Reg",
    "staff_valuer": "🏢 Staff Valuer",
}

# ──────────────────────────────────────────────────────────
# Main menu button labels & keyboard
# ──────────────────────────────────────────────────────────
BTN_ASSIGN        = "📋 New Assignment"
BTN_DLV_BATCH     = "📥 DLV Batch"
BTN_DLV_QUEUE     = "🔍 DLV Queue"
BTN_AUTO_FETCH    = "⏰ Auto Fetch"
BTN_ASSIGNMENTS   = "📜 Assignments"
BTN_AUTH          = "🔑 Refresh Auth"
BTN_TOKEN_STATUS  = "🔒 Token Status"
BTN_DAEMON        = "🔄 Token Daemon"
BTN_VALUERS       = "👥 Saved Valuers"
BTN_DELETE        = "🗑 Delete Valuer"
BTN_FETCH_TASKS   = "📊 Fetch Tasks"
BTN_DLV_TASKS     = "📋 DLV Tasks"
BTN_AF_RESULTS    = "🗂 AF Results"
BTN_BULK_EXPORT   = "📤 Export Valuation Report"
BTN_EXPORT_STATUS = "📊 Export Status"
BTN_JOB_DIST      = "🏆 Job Distribution"
BTN_HELP          = "❓ Help"
BTN_RESTART       = "🔁 Restart Bot"
BTN_CANCEL        = "🛑 Cancel"

# Filter that matches any of the persistent menu button texts
_MENU_BUTTON_FILTER = filters.Regex(
    f"^({re.escape(BTN_ASSIGN)}|{re.escape(BTN_DLV_BATCH)}|{re.escape(BTN_DLV_QUEUE)}"
    f"|{re.escape(BTN_AUTO_FETCH)}|{re.escape(BTN_ASSIGNMENTS)}|{re.escape(BTN_AUTH)}"
    f"|{re.escape(BTN_TOKEN_STATUS)}|{re.escape(BTN_DAEMON)}"
    f"|{re.escape(BTN_VALUERS)}|{re.escape(BTN_DELETE)}"
    f"|{re.escape(BTN_FETCH_TASKS)}|{re.escape(BTN_AF_RESULTS)}|{re.escape(BTN_BULK_EXPORT)}"
    f"|{re.escape(BTN_EXPORT_STATUS)}|{re.escape(BTN_JOB_DIST)}"
    f"|{re.escape(BTN_RESTART)}|{re.escape(BTN_HELP)}|{re.escape(BTN_CANCEL)})$"
)
_CANCEL_FILTER = filters.Regex(f"^{re.escape(BTN_CANCEL)}$")


def _main_menu() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        [
            [KeyboardButton(BTN_ASSIGN)],
            [KeyboardButton(BTN_FETCH_TASKS),  KeyboardButton(BTN_AUTO_FETCH)],
            [KeyboardButton(BTN_DLV_BATCH)],
            [KeyboardButton(BTN_AF_RESULTS),   KeyboardButton(BTN_ASSIGNMENTS)],
            [KeyboardButton(BTN_BULK_EXPORT), KeyboardButton(BTN_EXPORT_STATUS)],
            [KeyboardButton(BTN_JOB_DIST)],
            [KeyboardButton(BTN_DAEMON)],
            [KeyboardButton(BTN_AUTH),         KeyboardButton(BTN_TOKEN_STATUS)],
            [KeyboardButton(BTN_RESTART)],
            [KeyboardButton(BTN_VALUERS),      KeyboardButton(BTN_DELETE)],
            [KeyboardButton(BTN_HELP),         KeyboardButton(BTN_CANCEL)],
        ],
        resize_keyboard=True,
        is_persistent=True,
    )


# ──────────────────────────────────────────────────────────
# Auth guard
# ──────────────────────────────────────────────────────────
def allowed(update: Update) -> bool:
    if not ALLOWED_IDS:
        return True
    return update.effective_user.id in ALLOWED_IDS


async def deny(update: Update):
    await update.message.reply_text("⛔ You are not authorised to use this bot.")


# ──────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────
def get_sess(ctx: ContextTypes.DEFAULT_TYPE) -> Session:
    if "session" not in ctx.user_data:
        ctx.user_data["session"] = Session()
    return ctx.user_data["session"]


def parse_refs(raw: str) -> List[str]:
    return [r.strip() for r in re.split(r"[\n,]+", raw) if r.strip()]


def _cred_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(CRED_LABELS["publicuser"],   callback_data="cred:publicuser")],
        [InlineKeyboardButton(CRED_LABELS["staff"],        callback_data="cred:staff")],
        [InlineKeyboardButton(CRED_LABELS["staff2"],       callback_data="cred:staff2")],
        [InlineKeyboardButton(CRED_LABELS["staff_valuer"], callback_data="cred:staff_valuer")],
    ])


def _confirm_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Confirm & Assign", callback_data="confirm:yes"),
        InlineKeyboardButton("❌ Cancel",           callback_data="confirm:no"),
    ]])


async def _do_valuer_search(message, sess: Session) -> Optional[List[Dict]]:
    """Search for valuers; returns list or None (error already sent)."""
    try:
        headers = {
            "Authorization": f"Bearer {sess.tokens.access_token}",
            "JWTAUTH":       f"Bearer {sess.tokens.jwt}",
        }
        params = {
            "account_type": "STAFF",
            "filter_type":  "ACTIVE",
            "page":         1,
            "search":       sess.valuer_name,
        }
        resp = sess.session.get(
            f"{BASE_URL}/acl/api/v1/accounts/list-user-accounts",
            headers=headers, params=params, timeout=30,
        )
        resp.raise_for_status()
        return resp.json().get("results", [])
    except Exception as e:
        await message.reply_text(f"❌ Valuer search failed: `{e}`", parse_mode="Markdown")
        return None


async def _show_valuer_keyboard(message, sess: Session, results: List[Dict]) -> int:
    sess.valuers = results
    rows = []
    for i, v in enumerate(results):
        sd   = v.get("staff_details", {})
        name = " ".join(filter(None, [sd.get("firstname"), sd.get("middlename"), sd.get("lastname")]))
        rows.append([InlineKeyboardButton(name or f"Valuer {i+1}", callback_data=f"valuer:{i}")])
    await message.reply_text(
        f"Found *{len(results)}* valuer(s). Select one:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(rows),
    )
    return S.SELECT_VALUER


# ──────────────────────────────────────────────────────────
# OCR helpers
# ──────────────────────────────────────────────────────────

# Matches patterns like LS/VAL/2024/001 — 3+ slash-separated alphanumeric segments
_REF_RE = re.compile(r'\b[A-Z0-9]{2,}(?:/[A-Z0-9]{2,}){2,}\b')


def _extract_refs_from_text(text: str) -> List[str]:
    return list(dict.fromkeys(_REF_RE.findall(text.upper())))  # dedup, preserve order


async def ocr_extract_refs(photo_bytes: bytes) -> Tuple[List[str], str]:
    """
    Try Tesseract first; fall back to Claude Vision if nothing found.
    Returns (refs, source) where source is "tesseract" or "claude".
    """
    # ── Tesseract ─────────────────────────────────────────
    try:
        img  = Image.open(io.BytesIO(photo_bytes))
        text = pytesseract.image_to_string(img)
        refs = _extract_refs_from_text(text)
        if refs:
            return refs, "tesseract"
    except Exception as e:
        logger.warning("Tesseract failed: %s", e)

    # ── Claude Vision fallback ────────────────────────────
    if not ANTHROPIC_API_KEY:
        return [], "none"
    try:
        client  = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        b64_img = base64.standard_b64encode(photo_bytes).decode()
        resp    = client.messages.create(
            model="claude-opus-4-6",
            max_tokens=512,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {"type": "base64", "media_type": "image/jpeg", "data": b64_img},
                    },
                    {
                        "type": "text",
                        "text": (
                            "Extract every reference number from this document image. "
                            "Reference numbers follow a pattern like LS/VAL/2024/001 — "
                            "alphanumeric segments separated by forward slashes. "
                            "Return ONLY the reference numbers, one per line, nothing else."
                        ),
                    },
                ],
            }],
        )
        text = resp.content[0].text
        refs = _extract_refs_from_text(text)
        return refs, "claude"
    except Exception as e:
        logger.warning("Claude Vision failed: %s", e)
        return [], "none"


# ──────────────────────────────────────────────────────────
# /start  /help
# ──────────────────────────────────────────────────────────
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    await update.message.reply_text(
        "🏛 *Ardhisasa Valuation Bot*\n\n"
        "Use the buttons below to get started:",
        parse_mode="Markdown",
        reply_markup=_main_menu(),
    )


async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await cmd_start(update, ctx)


# ──────────────────────────────────────────────────────────
# /valuers  /delete_valuer — manage saved valuers
# ──────────────────────────────────────────────────────────
async def cmd_assignments(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    assignments = load_saved_assignments()
    if not assignments:
        await update.message.reply_text(
            "📭 No assignments recorded yet.",
            reply_markup=_main_menu(),
        )
        return

    # Sort newest first
    sorted_items = sorted(
        assignments.items(),
        key=lambda kv: kv[1].get("assigned_at", ""),
        reverse=True,
    )

    lines = []
    for ref, info in sorted_items:
        valuer = info.get("valuer_name", "Unknown")
        when   = info.get("assigned_at", "—")
        lines.append(f"• `{ref}`\n  👤 {valuer} | 🕐 {when}")

    header = f"📜 *Assignments ({len(lines)} total)*\n\n"
    chunks = []
    chunk  = header
    for line in lines:
        candidate = (chunk + line + "\n\n").strip()
        if len(candidate) > 4000:
            chunks.append(chunk)
            chunk = line + "\n\n"
        else:
            chunk = candidate + "\n"
    if chunk.strip():
        chunks.append(chunk)

    for i, c in enumerate(chunks):
        await update.message.reply_text(
            c,
            parse_mode="Markdown",
            reply_markup=_main_menu() if i == len(chunks) - 1 else None,
        )


async def cmd_valuers(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    valuers = load_saved_valuers()
    if not valuers:
        await update.message.reply_text(
            "📭 No saved valuers yet.\n"
            "They are saved automatically after a successful assignment.",
            reply_markup=_main_menu(),
        )
        return
    lines = [
        f"{i+1}. *{v['name']}* — ID: `{v['uid']}` | Acct: `{v['account_number']}`"
        for i, v in enumerate(valuers)
    ]
    await update.message.reply_text(
        "📋 *Saved Valuers:*\n\n" + "\n".join(lines),
        parse_mode="Markdown",
        reply_markup=_main_menu(),
    )


async def cmd_delete_valuer(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    valuers = load_saved_valuers()
    if not valuers:
        await update.message.reply_text(
            "📭 No saved valuers to delete.",
            reply_markup=_main_menu(),
        )
        return
    rows = [
        [InlineKeyboardButton(f"🗑 {v['name']}", callback_data=f"del:{i}")]
        for i, v in enumerate(valuers)
    ]
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data="del:cancel")])
    await update.message.reply_text(
        "Select a valuer to delete:",
        reply_markup=InlineKeyboardMarkup(rows),
    )


async def recv_delete_valuer(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data.split(":")[1]
    if data == "cancel":
        await query.edit_message_text("Deletion cancelled.")
        return
    idx = int(data)
    valuers = load_saved_valuers()
    if idx >= len(valuers):
        await query.edit_message_text("⚠️ Valuer not found.")
        return
    removed = valuers.pop(idx)
    _ensure_data_dir()
    with open(SAVED_VALUERS_FILE, "w") as f:
        json.dump(valuers, f, indent=2)
    await query.edit_message_text(
        f"🗑 Removed *{removed['name']}* from saved valuers.", parse_mode="Markdown"
    )


# ──────────────────────────────────────────────────────────
# Step 1 — /assign → ask reference numbers
# ──────────────────────────────────────────────────────────
async def cmd_assign(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    ctx.user_data["session"] = Session()
    await update.message.reply_text(
        "📋 *New Assignment Flow*\n\n"
        "Step 1 — How would you like to provide the reference numbers?",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove(),
    )
    await update.message.reply_text(
        "Choose an input method:",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✏️ Type Reference Numbers", callback_data="input:text")],
            [InlineKeyboardButton("📷 Add Photo(s)",           callback_data="input:photo")],
        ]),
    )
    return S.INPUT_METHOD


# ──────────────────────────────────────────────────────────
# Step 1a — input method chosen
# ──────────────────────────────────────────────────────────
async def recv_input_method(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    method = query.data.split(":")[1]

    if method == "text":
        await query.edit_message_text(
            "✏️ Enter *reference numbers*:\n"
            "_Comma-separated or one per line, e.g._\n"
            "`LS/VAL/2024/001, LS/VAL/2024/002`",
            parse_mode="Markdown",
        )
        return S.REF_NUMBERS
    else:
        await query.edit_message_text(
            "📷 Send a photo of the document.\n"
            "_You can send multiple photos one by one._\n\n"
            "Tap *✅ Done — Review refs* when finished.",
            parse_mode="Markdown",
        )
        return S.RECV_PHOTOS


# ──────────────────────────────────────────────────────────
# Step 1b — receive photo(s) → OCR → accumulate refs
# ──────────────────────────────────────────────────────────
async def recv_photo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sess = get_sess(ctx)
    await update.message.reply_text("🔍 Processing photo…")

    photo_file = await update.message.photo[-1].get_file()   # largest size
    photo_bytes = await photo_file.download_as_bytearray()

    refs, source = await ocr_extract_refs(bytes(photo_bytes))

    if not refs:
        await update.message.reply_text(
            "⚠️ Could not extract any reference numbers from this photo.\n"
            "Try a clearer image, or send another photo.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("✅ Done — Review refs", callback_data="photo:done"),
            ]]) if sess.extracted_refs else None,
        )
        return S.RECV_PHOTOS

    # Merge, avoiding duplicates
    new_refs = [r for r in refs if r not in sess.extracted_refs]
    sess.extracted_refs.extend(new_refs)

    source_label = "Tesseract OCR" if source == "tesseract" else "Claude Vision"
    running = "\n".join(f"  • `{r}`" for r in sess.extracted_refs)
    await update.message.reply_text(
        f"✅ *{len(new_refs)} new ref(s) found* via {source_label}.\n\n"
        f"*Running total ({len(sess.extracted_refs)}):*\n{running}\n\n"
        "_Send another photo or tap Done._",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Done — Review refs", callback_data="photo:done"),
        ]]),
    )
    return S.RECV_PHOTOS


async def recv_photo_done(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess = get_sess(ctx)

    if not sess.extracted_refs:
        await query.edit_message_text(
            "⚠️ No reference numbers extracted yet. Send at least one photo."
        )
        return S.RECV_PHOTOS

    refs_list = "\n".join(f"  • `{r}`" for r in sess.extracted_refs)
    await query.edit_message_text(
        f"📋 *Extracted Reference Numbers ({len(sess.extracted_refs)}):*\n\n"
        f"{refs_list}\n\n"
        "Are these correct?",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Confirm & Proceed", callback_data="refs:confirm")],
            [InlineKeyboardButton("✏️ Edit (retype manually)", callback_data="refs:edit")],
            [InlineKeyboardButton("❌ Cancel",              callback_data="refs:cancel")],
        ]),
    )
    return S.CONFIRM_REFS


# ──────────────────────────────────────────────────────────
# Step 1c — confirm extracted refs
# ──────────────────────────────────────────────────────────
async def recv_confirm_refs(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess   = get_sess(ctx)
    choice = query.data.split(":")[1]

    if choice == "cancel":
        await query.edit_message_text("❌ Assignment cancelled.")
        await query.message.reply_text("Use the menu to start again.", reply_markup=_main_menu())
        return ConversationHandler.END

    if choice == "edit":
        await query.edit_message_text(
            "✏️ Enter the *correct reference numbers*:\n"
            "_Comma-separated or one per line._",
            parse_mode="Markdown",
        )
        return S.REF_NUMBERS

    # confirm — treat extracted refs as the final list
    sess.refs = sess.extracted_refs[:]
    await query.edit_message_text(
        f"✅ *{len(sess.refs)} reference(s) confirmed.*",
        parse_mode="Markdown",
    )
    return await _check_assignments_and_proceed(query.message, sess)


# ──────────────────────────────────────────────────────────
# Step 2 — receive refs (typed) → check existing assignments
# ──────────────────────────────────────────────────────────
async def recv_refs(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sess = get_sess(ctx)
    refs = parse_refs(update.message.text)
    if not refs:
        await update.message.reply_text("⚠️ No valid references found. Try again.")
        return S.REF_NUMBERS

    sess.refs = refs
    return await _check_assignments_and_proceed(update.message, sess)


async def _check_assignments_and_proceed(message, sess: Session) -> int:
    """Check refs against saved assignments, then route accordingly."""
    refs     = sess.refs
    existing = load_saved_assignments()
    already  = [
        {"ref": r, "valuer_name": existing[r]["valuer_name"], "assigned_at": existing[r]["assigned_at"]}
        for r in refs if r in existing
    ]
    new_refs = [r for r in refs if r not in existing]

    if already:
        sess.already_assigned = already

        already_lines = "\n".join(
            f"  • `{a['ref']}` → *{a['valuer_name']}* _(on {a['assigned_at']})_"
            for a in already
        )
        new_lines = ("\n".join(f"  • `{r}`" for r in new_refs)) if new_refs else "_None_"

        await message.reply_text(
            f"⚠️ *Some references are already assigned:*\n{already_lines}\n\n"
            f"*New (unassigned):*\n{new_lines}\n\n"
            "What would you like to do?",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Reassign existing + assign new", callback_data="reassign:all")],
                [InlineKeyboardButton("⏭ Skip existing, assign new only",  callback_data="reassign:skip")],
                [InlineKeyboardButton("❌ Cancel",                          callback_data="reassign:cancel")],
            ]),
        )
        return S.REASSIGN_CONFIRM

    return await _proceed_to_valuer_pick(message, sess, refs)


async def _proceed_to_valuer_pick(message, sess: Session, refs: List[str]) -> int:
    """Show valuer picker after refs are finalised."""
    bullet_list = "\n".join(f"  • `{r}`" for r in refs)
    saved = load_saved_valuers()

    if saved:
        rows = [
            [InlineKeyboardButton(f"👤 {sv['name']}", callback_data=f"src:{i}")]
            for i, sv in enumerate(saved)
        ]
        rows.append([InlineKeyboardButton("🔍 Search new valuer", callback_data="src:new")])
        await message.reply_text(
            f"✅ *{len(refs)} reference(s)* queued:\n{bullet_list}\n\n"
            "Step 2 — Select a saved valuer or search for a new one:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(rows),
        )
        return S.PICK_VALUER_SOURCE
    else:
        await message.reply_text(
            f"✅ *{len(refs)} reference(s)* queued:\n{bullet_list}\n\n"
            "Step 2 — Enter the *valuer name* to search:\n"
            "_Partial names work, e.g._ `JOHN KAMAU`",
            parse_mode="Markdown",
        )
        return S.VALUER_NAME


# ──────────────────────────────────────────────────────────
# Step 2a — handle reassign choice
# ──────────────────────────────────────────────────────────
async def recv_reassign_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    choice = query.data.split(":")[1]
    sess   = get_sess(ctx)

    if choice == "cancel":
        await query.edit_message_text("❌ Assignment cancelled.")
        await query.message.reply_text("Use the menu to start again.", reply_markup=_main_menu())
        return ConversationHandler.END

    if choice == "skip":
        new_refs = [r for r in sess.refs if r not in {a["ref"] for a in sess.already_assigned}]
        if not new_refs:
            await query.edit_message_text(
                "ℹ️ All references are already assigned. Nothing to do."
            )
            await query.message.reply_text("Use the menu to start again.", reply_markup=_main_menu())
            return ConversationHandler.END
        sess.refs = new_refs
        await query.edit_message_text(
            f"⏭ Skipping already-assigned refs.\n"
            f"Proceeding with *{len(new_refs)}* new reference(s).",
            parse_mode="Markdown",
        )
    else:  # "all"
        await query.edit_message_text(
            f"🔄 Reassigning all *{len(sess.refs)}* reference(s).",
            parse_mode="Markdown",
        )

    return await _proceed_to_valuer_pick(query.message, sess, sess.refs)


# ──────────────────────────────────────────────────────────
# Step 2b — pick saved valuer or "search new"
# ──────────────────────────────────────────────────────────
async def recv_valuer_source(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    sess = get_sess(ctx)
    data = query.data.split(":")[1]

    if data == "new":
        await query.edit_message_text(
            "Step 2 — Enter the *valuer name* to search:\n"
            "_Partial names work, e.g._ `JOHN KAMAU`",
            parse_mode="Markdown",
        )
        return S.VALUER_NAME
    else:
        saved = load_saved_valuers()
        sv = saved[int(data)]
        sess.saved_valuer = sv
        await query.edit_message_text(
            f"✅ Valuer: *{sv['name']}*\n\n"
            "Step 3 — Choose *credential profile*:",
            parse_mode="Markdown",
            reply_markup=_cred_keyboard(),
        )
        return S.CHOOSE_CRED


# ──────────────────────────────────────────────────────────
# Step 2c — receive valuer name (search new path)
# ──────────────────────────────────────────────────────────
async def recv_valuer_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sess = get_sess(ctx)
    sess.valuer_name = update.message.text.strip()
    await update.message.reply_text(
        f"✅ Searching for: *{sess.valuer_name}*\n\n"
        "Step 3 — Choose *credential profile*:",
        parse_mode="Markdown",
        reply_markup=_cred_keyboard(),
    )
    return S.CHOOSE_CRED


# ──────────────────────────────────────────────────────────
# Step 3 — credential chosen → use cache or full login
# ──────────────────────────────────────────────────────────
async def recv_cred_choice(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    sess = get_sess(ctx)
    cred_type = query.data.split(":")[1]
    sess.cred_type = cred_type
    creds = CRED_MAP[cred_type]

    # ── Cached tokens path ────────────────────────────────
    cached = get_valid_tokens(cred_type)
    if cached:
        sess.tokens  = cached
        sess.session = build_session()
        logger.info("Using cached tokens for %s", cred_type)

        if sess.saved_valuer:
            # Saved valuer + cached tokens → jump straight to confirm
            sv = sess.saved_valuer
            refs_list = "\n".join(f"  • `{r}`" for r in sess.refs)
            await query.edit_message_text(
                f"🔑 Cached login: *{CRED_LABELS[cred_type]}*\n\n"
                f"📋 *Assignment Summary*\n\n"
                f"*Valuer:* {sv['name']}\n"
                f"*User ID:* `{sv['uid']}`\n\n"
                f"*References ({len(sess.refs)}):*\n{refs_list}\n\n"
                f"Proceed?",
                parse_mode="Markdown",
                reply_markup=_confirm_keyboard(),
            )
            return S.CONFIRM
        else:
            # Cached tokens + new search → search valuers
            await query.edit_message_text(
                f"🔑 Cached login: *{CRED_LABELS[cred_type]}*\n\n"
                f"🔍 Searching for valuer *{sess.valuer_name}*…",
                parse_mode="Markdown",
            )
            results = await _do_valuer_search(query.message, sess)
            if results is None:
                await query.message.reply_text(
                    "Use the menu to start again.", reply_markup=_main_menu()
                )
                return ConversationHandler.END
            if not results:
                await query.message.reply_text(
                    f"⚠️ No valuers found matching *{sess.valuer_name}*.",
                    parse_mode="Markdown",
                    reply_markup=_main_menu(),
                )
                return ConversationHandler.END
            return await _show_valuer_keyboard(query.message, sess, results)

    # ── Full login path ───────────────────────────────────
    await query.edit_message_text(
        f"✅ Credential: *{CRED_LABELS[cred_type]}* (`{creds['username']}`)\n\n"
        "Step 4 — 🔐 Sending login request…",
        parse_mode="Markdown",
    )

    sess.session = build_session()
    try:
        resp = sess.session.post(
            f"{AUTH_BASE_URL}/login",
            json={
                "username": creds["username"],
                "password": creds["password"],
                "usertype": creds["usertype"],
                "otpcode":  "",
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        if data.get("success") is False and "error" in data:
            raise RuntimeError(data.get("error") or data.get("message"))
    except Exception as e:
        await query.message.reply_text(
            f"❌ Login failed: `{e}`\n\nUse the menu to retry.",
            parse_mode="Markdown",
            reply_markup=_main_menu(),
        )
        return ConversationHandler.END

    await query.message.reply_text(
        "📲 OTP has been sent to the registered device.\n\n"
        "Step 4 — Please *reply with the OTP code* now:",
        parse_mode="Markdown",
    )
    return S.WAIT_OTP


# ──────────────────────────────────────────────────────────
# Step 4 — receive OTP → verify → cache tokens → search / confirm
# ──────────────────────────────────────────────────────────
async def recv_otp(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    sess  = get_sess(ctx)
    otp   = update.message.text.strip()
    creds = CRED_MAP[sess.cred_type]

    await update.message.reply_text("🔄 Verifying OTP…")

    try:
        resp = sess.session.post(
            f"{AUTH_BASE_URL}/otpverify",
            json={
                "username": creds["username"],
                "password": creds["password"],
                "otpcode":  otp,
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        details      = data.get("details", {})
        access_token = details.get("access_token")
        jwt          = details.get("jwt")
        refresh_token = details.get("refresh_token", "")
        if not access_token or not jwt:
            raise RuntimeError(f"Tokens missing. Keys: {list(data.keys())}")

        sess.tokens = AuthTokens(access_token=access_token, jwt=jwt)
        persist_tokens(sess.cred_type, access_token, jwt, refresh_token)

    except Exception as e:
        await update.message.reply_text(
            f"❌ OTP verification failed: `{e}`\n\nSend the OTP again or tap 🛑 Cancel.",
            parse_mode="Markdown",
        )
        return S.WAIT_OTP

    await update.message.reply_text("✅ Authenticated!")

    if sess.saved_valuer:
        # Saved valuer selected earlier → jump to confirm
        sv = sess.saved_valuer
        refs_list = "\n".join(f"  • `{r}`" for r in sess.refs)
        await update.message.reply_text(
            f"📋 *Assignment Summary*\n\n"
            f"*Valuer:* {sv['name']}\n"
            f"*User ID:* `{sv['uid']}`\n\n"
            f"*References ({len(sess.refs)}):*\n{refs_list}\n\n"
            f"Proceed?",
            parse_mode="Markdown",
            reply_markup=_confirm_keyboard(),
        )
        return S.CONFIRM

    await update.message.reply_text(
        f"🔍 Searching for valuer *{sess.valuer_name}*…", parse_mode="Markdown"
    )
    results = await _do_valuer_search(update.message, sess)
    if results is None:
        await update.message.reply_text(
            "Use the menu to start again.", reply_markup=_main_menu()
        )
        return ConversationHandler.END
    if not results:
        await update.message.reply_text(
            f"⚠️ No valuers found matching *{sess.valuer_name}*.",
            parse_mode="Markdown",
            reply_markup=_main_menu(),
        )
        return ConversationHandler.END
    return await _show_valuer_keyboard(update.message, sess, results)


# ──────────────────────────────────────────────────────────
# Step 5 — valuer selected → confirm
# ──────────────────────────────────────────────────────────
async def recv_valuer_select(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    sess = get_sess(ctx)
    idx  = int(query.data.split(":")[1])
    sess.selected_idx = idx
    v    = sess.valuers[idx]
    sd   = v.get("staff_details", {})
    name = " ".join(filter(None, [sd.get("firstname"), sd.get("middlename"), sd.get("lastname")]))
    uid  = sd.get("user_id", v.get("id", "?"))

    refs_list = "\n".join(f"  • `{r}`" for r in sess.refs)
    await query.edit_message_text(
        f"📋 *Assignment Summary*\n\n"
        f"*Valuer:* {name}\n"
        f"*User ID:* `{uid}`\n\n"
        f"*References ({len(sess.refs)}):*\n{refs_list}\n\n"
        f"Proceed?",
        parse_mode="Markdown",
        reply_markup=_confirm_keyboard(),
    )
    return S.CONFIRM


# ──────────────────────────────────────────────────────────
# Step 6 — confirmed → run assignments → save valuer → show results
# ──────────────────────────────────────────────────────────
async def recv_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "confirm:no":
        await query.edit_message_text("❌ Assignment cancelled.")
        await query.message.reply_text("Use the menu to start again.", reply_markup=_main_menu())
        return ConversationHandler.END

    sess = get_sess(ctx)

    if sess.saved_valuer:
        sv   = sess.saved_valuer
        name = sv["name"]
        uid  = sv["uid"]
        acct = sv["account_number"]
    else:
        v    = sess.valuers[sess.selected_idx]
        sd   = v.get("staff_details", {})
        uid  = sd.get("user_id", v.get("id"))
        name = " ".join(filter(None, [sd.get("firstname"), sd.get("middlename"), sd.get("lastname")]))
        acct = v.get("account_number", "?")

    await query.edit_message_text(
        f"⚙️ Assigning *{name}* to {len(sess.refs)} reference(s)…",
        parse_mode="Markdown",
    )

    url     = f"{BASE_URL}/valuationservice/api/v1/stamp-duty/fix_application_details"
    headers = {
        "Authorization": f"Bearer {sess.tokens.access_token}",
        "JWTAUTH":       f"Bearer {sess.tokens.jwt}",
    }

    ok_refs, fail_refs = [], []
    result_lines = []

    for ref in sess.refs:
        try:
            r = sess.session.post(
                url, headers=headers,
                json={"reference_number": ref, "valuation_officer": uid, "node": "VALUATION_STAMP_DUTY_VALUER_REPORT"},
                timeout=30,
            )
            r.raise_for_status()
            ok_refs.append(ref)
            result_lines.append(f"✅ `{ref}`")
        except Exception as e:
            fail_refs.append(ref)
            result_lines.append(f"❌ `{ref}` — {e}")

    if ok_refs:
        persist_valuer(name, uid, acct)          # auto-save valuer for future assignments
        for ref in ok_refs:
            persist_assignment(ref, name, uid)   # record ref → valuer mapping

    summary = (
        f"🏁 *Assignment Complete*\n\n"
        f"*Valuer:* {name}\n"
        f"*Success:* {len(ok_refs)} / {len(sess.refs)}\n"
        f"*Failed:*  {len(fail_refs)} / {len(sess.refs)}\n\n"
        + "\n".join(result_lines)
    )

    if len(summary) > 4000:
        summary = summary[:4000] + "\n…_(truncated)_"

    await query.message.reply_text(summary, parse_mode="Markdown", reply_markup=_main_menu())

    if fail_refs:
        await query.message.reply_text(
            "⚠️ Some assignments failed. Tap *📋 New Assignment* to retry failed refs.",
            parse_mode="Markdown",
        )

    return ConversationHandler.END


# ──────────────────────────────────────────────────────────
# Receive Tasks — persistence helpers
# ──────────────────────────────────────────────────────────

def load_task_batches() -> List[Dict]:
    try:
        with open(SAVED_TASK_BATCHES_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def persist_task_batch(batch: Dict):
    _ensure_data_dir()
    batches = load_task_batches()
    batches.append(batch)
    with open(SAVED_TASK_BATCHES_FILE, "w") as f:
        json.dump(batches, f, indent=2)
    logger.info("Saved task batch %s (%d tasks)", batch["batch_id"], len(batch["tasks"]))


def load_schedules() -> List[Dict]:
    try:
        with open(SAVED_SCHEDULES_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def _save_schedules(schedules: List[Dict]):
    _ensure_data_dir()
    with open(SAVED_SCHEDULES_FILE, "w") as f:
        json.dump(schedules, f, indent=2)


def persist_schedule(sched: Dict):
    schedules = load_schedules()
    schedules = [s for s in schedules if s["schedule_id"] != sched["schedule_id"]]
    schedules.append(sched)
    _save_schedules(schedules)
    logger.info("Saved schedule %s (%dmin interval)", sched["schedule_id"], sched["interval_minutes"])


# ──────────────────────────────────────────────────────────
# Receive Tasks — staff validation
# ──────────────────────────────────────────────────────────

def _validate_staff(user_data: Dict) -> Tuple[bool, str, str, str, str]:
    """
    Check eligibility for task receipt.
    All detail fields live inside staff_details, not at the top level.
    Returns (ok, error_msg, task_type, staff_registry, staff_county).
    task_type is "STAMP_DUTY", "COUNTY_STAMP_DUTY", or "BOTH" (user must choose).
    registry/county are populated only for COUNTY_STAMP_DUTY / BOTH.
    """
    if user_data.get("account_status", "").upper() != "ACTIVE":
        return False, "Account status is not ACTIVE.", "", "", ""

    sd = user_data.get("staff_details") or {}

    dept_code = (
        sd.get("department_details", {})
          .get("department", {})
          .get("code", "")
    )
    if dept_code.upper() != "DLV":
        return False, f"Department is '{dept_code}', expected DLV.", "", "", ""

    roles      = sd.get("roles") or []
    role_names = [r.get("rolename", "").upper() for r in roles]
    if "VALUER" not in role_names:
        return False, f"No VALUER role found. Roles: {role_names}", "", "", ""

    has_county_valuer = "COUNTY_VALUER" in role_names

    if has_county_valuer:
        county_units = sd.get("county_units") or []
        if not county_units:
            return False, "COUNTY_VALUER role but no county_units found on account.", "", "", ""
        primary  = next((u for u in county_units if u.get("is_primary")), county_units[0])
        registry = primary.get("registry", "").upper()
        county   = primary.get("county", "").upper()
        # Has both VALUER and COUNTY_VALUER — let the user pick the task pool
        return True, "", "BOTH", registry, county

    return True, "", "STAMP_DUTY", "", ""


# ──────────────────────────────────────────────────────────
# Receive Tasks — API helpers
# ──────────────────────────────────────────────────────────

def _rt_auth_headers(rt: RTSession) -> Dict:
    return {
        "Authorization": f"Bearer {rt.tokens.access_token}",
        "JWTAUTH":       f"Bearer {rt.tokens.jwt}",
        "cparams":       CPARAMS_DLV,
    }


def _fetch_tasks(rt: RTSession, needed: int) -> List[Dict]:
    """
    Paginate the task list endpoint, returning tasks that pass the
    node/status pre-filter (and county+registry filter for COUNTY_STAMP_DUTY).
    Stops once we have needed*5 candidates (to leave headroom for detail filtering).
    """
    headers = _rt_auth_headers(rt)
    if rt.task_type == "STAMP_DUTY":
        base_params: Dict = {
            "filter": "Pending", "role": "DLV",
            "request_type": "STAMP_DUTY", "search": "",
        }
    else:
        base_params = {
            "filter": "Ongoing", "from_ardhipay": "true",
            "role": "DLV", "request_type": "COUNTY_STAMP_DUTY", "search": "",
        }

    tasks: List[Dict] = []
    page = 1
    target = max(needed * 5, 50)

    while len(tasks) < target:
        resp = rt.session.get(
            f"{BASE_URL}/valuationservice/api/v1/stamp-duty/application",
            headers=headers, params={**base_params, "page": page}, timeout=30,
        )
        resp.raise_for_status()
        data    = resp.json()
        results = data.get("results", [])
        if not results:
            break

        for task in results:
            if (task.get("application_status") == "ONGOING"
                    and task.get("node") == "VALUATION_STAMP_DUTY_CREATED"):
                if rt.task_type == "COUNTY_STAMP_DUTY":
                    if (task.get("registry", "").upper() != rt.staff_registry
                            or task.get("county", "").upper() != rt.staff_county):
                        continue
                tasks.append(task)

        if not data.get("next"):
            break
        page += 1

    return tasks


def _fetch_task_detail(rt: RTSession, task_id: str) -> Optional[Dict]:
    """
    Call the detail-view endpoint and return the payload only if all three
    conditions are met: application_status=ONGOING, node=VALUATION_STAMP_DUTY_CREATED,
    node_code=APPLICATION_AWAITING_VALUATION.
    Returns None if any condition fails or the request errors.
    """
    try:
        resp = rt.session.get(
            f"{BASE_URL}/valuationservice/api/v1/stamp-duty/application/detail-view",
            headers=_rt_auth_headers(rt),
            params={"request_id": task_id},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        logger.warning("detail-view failed for %s: %s", task_id, e)
        return None

    if (data.get("application_status") != "ONGOING"
            or data.get("node") != "VALUATION_STAMP_DUTY_CREATED"):
        return None

    ext = data.get("external_process_details", {})
    if ext.get("node_code") != "APPLICATION_AWAITING_VALUATION":
        return None

    return data


def _verify_and_filter_tasks(
    rt: RTSession,
    candidates: List[Dict],
) -> List[Dict]:
    """
    For each candidate, call detail-view, apply the three mandatory checks,
    then apply the optional consideration_amount range filter.
    Stops once rt.task_count tasks are matched.
    """
    matched: List[Dict] = []
    for task in candidates:
        if len(matched) >= rt.task_count:
            break
        detail = _fetch_task_detail(rt, task["id"])
        if detail is None:
            continue
        ext    = detail.get("external_process_details", {})
        amount = float(ext.get("consideration_amount") or 0)
        if rt.amount_min is not None and amount < rt.amount_min:
            continue
        if rt.amount_max is not None and amount > rt.amount_max:
            continue
        matched.append({
            "id":                   task["id"],
            "reference_number":     task["reference_number"],
            "consideration_amount": amount,
            "parcel_number":        task.get("parcel_number", ""),
            "registry":             task.get("registry", ""),
            "county":               task.get("county", ""),
            "date_created":         task.get("date_created", ""),
        })
    return matched


async def _do_assign_tasks(
    bot,
    chat_id: int,
    http_sess,
    tokens: AuthTokens,
    tasks: List[Dict],
    staff_uid: str,
    staff_name: str,
    cred_type: str,
):
    """POST assignments and send a result summary to chat_id."""
    url     = f"{BASE_URL}/valuationservice/api/v1/stamp-duty/fix_application_details"
    headers = {
        "Authorization": f"Bearer {tokens.access_token}",
        "JWTAUTH":       f"Bearer {tokens.jwt}",
    }

    ok_refs: List[str]   = []
    fail_refs: List[str] = []
    result_lines: List[str] = []

    for task in tasks:
        ref = task["reference_number"]
        try:
            r = http_sess.post(
                url, headers=headers,
                json={
                    "reference_number":   ref,
                    "valuation_officer":  staff_uid,
                    "node":               "VALUATION_STAMP_DUTY_VALUER_REPORT",
                },
                timeout=30,
            )
            r.raise_for_status()
            ok_refs.append(ref)
            result_lines.append(f"✅ `{ref}` — KES {task['consideration_amount']:,.0f}")
            persist_assignment(ref, staff_name, staff_uid)
        except Exception as e:
            fail_refs.append(ref)
            result_lines.append(f"❌ `{ref}` — {e}")

    # Persist the batch
    batch = {
        "batch_id":   str(uuid.uuid4()),
        "staff_name": staff_name,
        "staff_uid":  staff_uid,
        "cred_type":  cred_type,
        "created_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "tasks":      [t for t in tasks if t["reference_number"] in ok_refs],
        "failed":     fail_refs,
    }
    persist_task_batch(batch)

    summary = (
        f"🏁 *Receive Tasks Complete*\n\n"
        f"*Valuer:* {staff_name}\n"
        f"*Assigned:* {len(ok_refs)} / {len(tasks)}\n"
        f"*Failed:*   {len(fail_refs)} / {len(tasks)}\n\n"
        + "\n".join(result_lines)
    )
    if len(summary) > 4000:
        summary = summary[:4000] + "\n…_(truncated)_"

    await bot.send_message(chat_id, summary, parse_mode="Markdown")


# ──────────────────────────────────────────────────────────
# Receive Tasks — scheduled job
# ──────────────────────────────────────────────────────────

async def _receive_tasks_job(context: ContextTypes.DEFAULT_TYPE):
    """JobQueue callback: runs receive-tasks automatically on a schedule."""
    job_data  = context.job.data
    chat_id   = job_data["chat_id"]
    sched     = job_data["schedule"]
    cred_type = sched["cred_type"]

    tokens = get_valid_tokens(cred_type)
    if not tokens:
        await context.bot.send_message(
            chat_id,
            "⚠️ *Scheduled receive-tasks failed:* cached tokens are expired.\n"
            "Use */receive* to re-authenticate and reschedule.",
            parse_mode="Markdown",
        )
        return

    await context.bot.send_message(chat_id, "⏰ *Scheduled receive-tasks running…*", parse_mode="Markdown")

    rt = RTSession()
    rt.tokens               = tokens
    rt.session              = build_session()
    rt.task_type            = sched["task_type"]
    rt.staff_registry       = sched.get("staff_registry", "")
    rt.staff_county         = sched.get("staff_county", "")
    rt.task_count           = sched["task_count"]
    rt.amount_min           = sched.get("amount_min")
    rt.amount_max           = sched.get("amount_max")

    try:
        candidates = _fetch_tasks(rt, rt.task_count)
    except Exception as e:
        await context.bot.send_message(chat_id, f"❌ Task fetch failed: `{e}`", parse_mode="Markdown")
        return

    if not candidates:
        await context.bot.send_message(chat_id, "ℹ️ Scheduled run: no eligible tasks found.")
        return

    matched = _verify_and_filter_tasks(rt, candidates)
    if not matched:
        await context.bot.send_message(chat_id, "ℹ️ Scheduled run: no tasks passed detail-view verification.")
        return

    await _do_assign_tasks(
        context.bot, chat_id,
        rt.session, tokens,
        matched,
        sched["staff_uid"], sched["staff_name"], cred_type,
    )


def _restore_schedules(app) -> None:
    """Re-register active scheduled jobs from persistent storage on startup."""
    schedules = load_schedules()
    restored  = 0
    for sched in schedules:
        if not sched.get("active", True):
            continue
        interval = sched["interval_minutes"] * 60
        app.job_queue.run_repeating(
            _receive_tasks_job,
            interval=interval,
            first=interval,
            data={"chat_id": sched["chat_id"], "schedule": sched},
            name=f"rt_{sched['schedule_id']}",
        )
        restored += 1
    if restored:
        logger.info("Restored %d scheduled receive-tasks job(s).", restored)


# ──────────────────────────────────────────────────────────
# Receive Tasks — conversation helpers
# ──────────────────────────────────────────────────────────

def _get_rt(ctx: ContextTypes.DEFAULT_TYPE) -> RTSession:
    if "rt_session" not in ctx.user_data:
        ctx.user_data["rt_session"] = RTSession()
    return ctx.user_data["rt_session"]


def _fetch_staff_detail(rt: RTSession, list_entry: Dict) -> Dict:
    """
    Fetch the full staff profile from the detail endpoint.
    The list-user-accounts response only returns summary fields;
    department_details / roles / ardhipay_roles / county_units come from here.
    Falls back to the list entry if all attempts fail.
    """
    headers = {
        "Authorization": f"Bearer {rt.tokens.access_token}",
        "JWTAUTH":       f"Bearer {rt.tokens.jwt}",
    }
    account_id = list_entry.get("id", "")
    user_id    = list_entry.get("staff_details", {}).get("user_id", account_id)

    candidates = [
        f"{BASE_URL}/acl/api/v1/accounts/get-user-detail?user_id={user_id}",
        f"{BASE_URL}/acl/api/v1/accounts/get-user-detail/{user_id}",
        f"{BASE_URL}/acl/api/v1/accounts/user-details?user_id={user_id}",
        f"{BASE_URL}/acl/api/v1/accounts/view-user?user_id={user_id}",
        f"{BASE_URL}/acl/api/v1/accounts/{account_id}",
    ]

    for url in candidates:
        try:
            resp = rt.session.get(url, headers=headers, timeout=15)
            logger.info("Staff detail probe %s → HTTP %s body: %.300s", url, resp.status_code, resp.text)
            if resp.status_code != 200:
                continue
            data = resp.json()
            if data.get("department_details") or data.get("roles") or data.get("ardhipay_roles"):
                logger.info("Staff detail fetched from: %s  keys: %s", url, list(data.keys()))
                return data
        except Exception as e:
            logger.warning("Detail attempt failed (%s): %s", url, e)

    logger.warning(
        "Could not fetch staff detail — falling back to list entry.\n"
        "list entry staff_details: %s",
        json.dumps(list_entry.get("staff_details"), default=str),
    )
    return list_entry


async def _rt_resolve_saved_valuer(message, rt: RTSession) -> int:
    """
    After auth, fetch and validate a pre-selected saved valuer.
    Searches by name and matches on account_number, then runs validation.
    """
    sv = rt.saved_valuer
    try:
        headers = {
            "Authorization": f"Bearer {rt.tokens.access_token}",
            "JWTAUTH":       f"Bearer {rt.tokens.jwt}",
        }
        resp = rt.session.get(
            f"{BASE_URL}/acl/api/v1/accounts/list-user-accounts",
            headers=headers,
            params={"account_type": "STAFF", "filter_type": "ACTIVE",
                    "page": 1, "search": sv["name"]},
            timeout=30,
        )
        resp.raise_for_status()
        results = resp.json().get("results", [])
    except Exception as e:
        await message.reply_text(
            f"❌ Could not fetch profile for *{sv['name']}*: `{e}`",
            parse_mode="Markdown", reply_markup=_main_menu(),
        )
        return ConversationHandler.END

    # Match by account_number first, fall back to uid, then first result
    match = (
        next((r for r in results if r.get("account_number") == sv["account_number"]), None)
        or next((r for r in results if r.get("id") == sv["uid"]), None)
        or (results[0] if len(results) == 1 else None)
    )
    if not match:
        await message.reply_text(
            f"⚠️ Could not uniquely identify *{sv['name']}* from search results.\n"
            "Use 🔍 Search new valuer to select manually.",
            parse_mode="Markdown", reply_markup=_main_menu(),
        )
        return ConversationHandler.END

    user_data = _fetch_staff_detail(rt, match)

    sd   = user_data.get("staff_details", {})
    name = " ".join(filter(None, [sd.get("firstname"), sd.get("middlename"), sd.get("lastname")]))

    ok, err_msg, task_type, registry, county = _validate_staff(user_data)
    if not ok:
        await message.reply_text(
            f"❌ *Validation failed for {name}:*\n{err_msg}",
            parse_mode="Markdown", reply_markup=_main_menu(),
        )
        return ConversationHandler.END

    rt.staff_data     = user_data
    rt.staff_registry = registry
    rt.staff_county   = county

    sd_roles    = sd.get("roles") or []
    role_labels = ", ".join(r.get("rolename", "") for r in sd_roles) or "None"

    if task_type == "BOTH":
        await message.reply_text(
            f"✅ *Staff Validated*\n\n"
            f"*Name:* {name}\n"
            f"*Roles:* {role_labels}\n"
            f"*County:* {county}  |  *Registry:* {registry}\n\n"
            "Step 3 — Which task pool do you want to assign from?",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🏛 Stamp Duty (national, from_ardhipay=false)",
                                      callback_data="rt_type:STAMP_DUTY")],
                [InlineKeyboardButton("🏙 County Stamp Duty (from_ardhipay=true)",
                                      callback_data="rt_type:COUNTY_STAMP_DUTY")],
            ]),
        )
        return RS.TASK_TYPE

    rt.task_type = task_type
    type_label   = "County Stamp Duty" if task_type == "COUNTY_STAMP_DUTY" else "Stamp Duty"
    await message.reply_text(
        f"✅ *Staff Validated*\n\n"
        f"*Name:* {name}\n"
        f"*Task Type:* {type_label}\n"
        f"*Roles:* {role_labels}\n\n"
        "Step 3 — How many tasks do you want to assign?",
        parse_mode="Markdown",
    )
    return RS.TASK_COUNT


async def _rt_do_staff_search(message, rt: RTSession) -> int:
    """Search the accounts endpoint and show a selection keyboard."""
    try:
        headers = {
            "Authorization": f"Bearer {rt.tokens.access_token}",
            "JWTAUTH":       f"Bearer {rt.tokens.jwt}",
        }
        resp = rt.session.get(
            f"{BASE_URL}/acl/api/v1/accounts/list-user-accounts",
            headers=headers,
            params={"account_type": "STAFF", "filter_type": "ACTIVE",
                    "page": 1, "search": rt.staff_name},
            timeout=30,
        )
        resp.raise_for_status()
        results = resp.json().get("results", [])
    except Exception as e:
        await message.reply_text(
            f"❌ Staff search failed: `{e}`", parse_mode="Markdown", reply_markup=_main_menu()
        )
        return ConversationHandler.END

    if not results:
        await message.reply_text(
            f"⚠️ No staff found matching *{rt.staff_name}*.",
            parse_mode="Markdown", reply_markup=_main_menu(),
        )
        return ConversationHandler.END

    rt.staff_results = results
    rows = []
    for i, v in enumerate(results):
        sd   = v.get("staff_details", {})
        name = " ".join(filter(None, [sd.get("firstname"), sd.get("middlename"), sd.get("lastname")]))
        rows.append([InlineKeyboardButton(name or f"Staff {i+1}", callback_data=f"rt_staff:{i}")])

    await message.reply_text(
        f"Found *{len(results)}* staff member(s). Select one:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(rows),
    )
    return RS.SELECT_STAFF


async def _rt_fetch_and_show(message, rt: RTSession) -> int:
    """Fetch + verify tasks, display them, and ask for confirmation."""
    await message.reply_text("⏳ Fetching eligible tasks from the queue…")

    # Peek at the total count first
    headers = _rt_auth_headers(rt)
    if rt.task_type == "STAMP_DUTY":
        peek_params: Dict = {
            "filter": "Pending", "role": "DLV",
            "request_type": "STAMP_DUTY", "search": "", "page": 1,
        }
    else:
        peek_params = {
            "filter": "Ongoing", "from_ardhipay": "true",
            "role": "DLV", "request_type": "COUNTY_STAMP_DUTY", "search": "", "page": 1,
        }

    try:
        peek = rt.session.get(
            f"{BASE_URL}/valuationservice/api/v1/stamp-duty/application",
            headers=headers, params=peek_params, timeout=30,
        )
        peek.raise_for_status()
        total_count = peek.json().get("count", "?")
    except Exception:
        total_count = "?"

    await message.reply_text(
        f"📊 Total tasks in queue: *{total_count}*\n"
        f"🔍 Scanning for up to *{rt.task_count}* eligible task(s)…",
        parse_mode="Markdown",
    )

    try:
        candidates = _fetch_tasks(rt, rt.task_count)
    except Exception as e:
        await message.reply_text(f"❌ Task fetch failed: `{e}`", parse_mode="Markdown", reply_markup=_main_menu())
        return ConversationHandler.END

    if not candidates:
        await message.reply_text(
            "ℹ️ No eligible tasks found matching your filters.", reply_markup=_main_menu()
        )
        return ConversationHandler.END

    await message.reply_text(f"🔍 Verifying *{len(candidates)}* candidate(s) via detail-view…", parse_mode="Markdown")

    matched = _verify_and_filter_tasks(rt, candidates)
    if not matched:
        await message.reply_text(
            "ℹ️ No tasks passed the detail-view verification checks.", reply_markup=_main_menu()
        )
        return ConversationHandler.END

    rt.matched_tasks = matched

    # Build display list
    sd   = rt.staff_data.get("staff_details", {})
    name = " ".join(filter(None, [sd.get("firstname"), sd.get("middlename"), sd.get("lastname")]))
    lines = []
    for i, t in enumerate(matched, 1):
        lines.append(
            f"{i}. `{t['reference_number']}`\n"
            f"   💰 KES {t['consideration_amount']:,.0f}\n"
            f"   📍 {t['parcel_number']} | 🏢 {t['registry']}\n"
            f"   📅 {t['date_created'][:10]}"
        )

    summary = (
        f"📋 *Tasks for {name}* ({len(matched)} task(s))\n\n"
        + "\n\n".join(lines)
        + "\n\nConfirm assignment?"
    )
    if len(summary) > 4000:
        summary = (
            f"📋 *{len(matched)} tasks* ready to assign to *{name}*.\n"
            "_(List too long to display in full)_\n\nConfirm assignment?"
        )

    await message.reply_text(
        summary,
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Confirm & Assign", callback_data="rt_confirm:yes")],
            [InlineKeyboardButton("❌ Cancel",           callback_data="rt_confirm:no")],
        ]),
    )
    return RS.RT_CONFIRM


# ──────────────────────────────────────────────────────────
# Receive Tasks — conversation handlers
# ──────────────────────────────────────────────────────────

async def cmd_receive(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    ctx.user_data["rt_session"] = RTSession()
    saved = load_saved_valuers()

    await update.message.reply_text(
        "📥 *Receive Tasks Flow*\n\nStep 1 — Select a valuer or search for a new one:",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove(),
    )

    if saved:
        rows = [
            [InlineKeyboardButton(f"👤 {sv['name']}", callback_data=f"rt_src:{i}")]
            for i, sv in enumerate(saved)
        ]
        rows.append([InlineKeyboardButton("🔍 Search new valuer", callback_data="rt_src:new")])
        await update.message.reply_text(
            "Choose a saved valuer or search new:",
            reply_markup=InlineKeyboardMarkup(rows),
        )
        return RS.PICK_STAFF_SOURCE

    # No saved valuers — go straight to name search
    await update.message.reply_text(
        "Enter the staff member's name to search:",
        parse_mode="Markdown",
    )
    return RS.STAFF_NAME


async def recv_rt_pick_source(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    rt   = _get_rt(ctx)
    data = query.data.split(":")[1]

    if data == "new":
        await query.edit_message_text(
            "🔍 Enter the staff member's name to search:",
            parse_mode="Markdown",
        )
        return RS.STAFF_NAME

    saved = load_saved_valuers()
    sv    = saved[int(data)]
    rt.saved_valuer = sv
    rt.staff_name   = sv["name"]

    await query.edit_message_text(
        f"👤 Selected: *{sv['name']}*\n\n"
        "Step 2 — Choose *credential profile*:",
        parse_mode="Markdown",
        reply_markup=_cred_keyboard(),
    )
    return RS.CHOOSE_CRED


async def recv_rt_staff_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    rt = _get_rt(ctx)
    rt.staff_name = update.message.text.strip()
    await update.message.reply_text(
        f"🔍 Searching for: *{rt.staff_name}*\n\n"
        "Step 2 — Choose *credential profile*:",
        parse_mode="Markdown",
        reply_markup=_cred_keyboard(),
    )
    return RS.CHOOSE_CRED


async def recv_rt_cred_choice(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    rt        = _get_rt(ctx)
    cred_type = query.data.split(":")[1]
    rt.cred_type = cred_type
    creds     = CRED_MAP[cred_type]

    cached = get_valid_tokens(cred_type)
    if cached:
        rt.tokens  = cached
        rt.session = build_session()
        if rt.saved_valuer:
            await query.edit_message_text(
                f"🔑 Cached login: *{CRED_LABELS[cred_type]}*\n\n"
                f"🔍 Fetching profile for *{rt.saved_valuer['name']}*…",
                parse_mode="Markdown",
            )
            return await _rt_resolve_saved_valuer(query.message, rt)
        await query.edit_message_text(
            f"🔑 Cached login: *{CRED_LABELS[cred_type]}*\n\n"
            f"🔍 Searching for *{rt.staff_name}*…",
            parse_mode="Markdown",
        )
        return await _rt_do_staff_search(query.message, rt)

    # Full login
    rt.session = build_session()
    await query.edit_message_text(
        f"✅ Credential: *{CRED_LABELS[cred_type]}*\n\n🔐 Sending login request…",
        parse_mode="Markdown",
    )
    try:
        resp = rt.session.post(
            f"{AUTH_BASE_URL}/login",
            json={"username": creds["username"], "password": creds["password"],
                  "usertype": creds["usertype"], "otpcode": ""},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        if data.get("success") is False:
            raise RuntimeError(data.get("error") or data.get("message"))
    except Exception as e:
        await query.message.reply_text(
            f"❌ Login failed: `{e}`", parse_mode="Markdown", reply_markup=_main_menu()
        )
        return ConversationHandler.END

    await query.message.reply_text(
        "📲 OTP sent to registered device.\n\nPlease *reply with the OTP code*:",
        parse_mode="Markdown",
    )
    return RS.WAIT_OTP


async def recv_rt_otp(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    rt    = _get_rt(ctx)
    otp   = update.message.text.strip()
    creds = CRED_MAP[rt.cred_type]

    await update.message.reply_text("🔄 Verifying OTP…")
    try:
        resp = rt.session.post(
            f"{AUTH_BASE_URL}/otpverify",
            json={"username": creds["username"], "password": creds["password"], "otpcode": otp},
            timeout=30,
        )
        resp.raise_for_status()
        data          = resp.json()
        details       = data.get("details", {})
        access_token  = details.get("access_token")
        jwt           = details.get("jwt")
        refresh_token = details.get("refresh_token", "")
        if not access_token or not jwt:
            raise RuntimeError(f"Tokens missing. Keys: {list(data.keys())}")
        rt.tokens = AuthTokens(access_token=access_token, jwt=jwt)
        persist_tokens(rt.cred_type, access_token, jwt, refresh_token)
    except Exception as e:
        await update.message.reply_text(
            f"❌ OTP failed: `{e}`\n\nSend the OTP again or tap 🛑 Cancel.",
            parse_mode="Markdown",
        )
        return RS.WAIT_OTP

    await update.message.reply_text("✅ Authenticated!")
    if rt.saved_valuer:
        return await _rt_resolve_saved_valuer(update.message, rt)
    return await _rt_do_staff_search(update.message, rt)


async def recv_rt_select_staff(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    rt        = _get_rt(ctx)
    idx       = int(query.data.split(":")[1])
    list_entry = rt.staff_results[idx]

    sd   = list_entry.get("staff_details", {})
    name = " ".join(filter(None, [sd.get("firstname"), sd.get("middlename"), sd.get("lastname")]))

    await query.edit_message_text(f"🔍 Fetching full profile for *{name}*…", parse_mode="Markdown")
    user_data = _fetch_staff_detail(rt, list_entry)

    ok, err_msg, task_type, registry, county = _validate_staff(user_data)
    if not ok:
        await query.edit_message_text(
            f"❌ *Validation failed for {name}:*\n{err_msg}",
            parse_mode="Markdown",
        )
        await query.message.reply_text("Use the menu to start again.", reply_markup=_main_menu())
        return ConversationHandler.END

    rt.staff_data     = user_data
    rt.staff_registry = registry
    rt.staff_county   = county

    sd_roles = (user_data.get("staff_details") or {}).get("roles") or []
    role_labels = ", ".join(r.get("rolename", "") for r in sd_roles) or "None"

    if task_type == "BOTH":
        # Staff has both VALUER and COUNTY_VALUER — let user pick the task pool
        await query.edit_message_text(
            f"✅ *Staff Validated*\n\n"
            f"*Name:* {name}\n"
            f"*Roles:* {role_labels}\n"
            f"*County:* {county}  |  *Registry:* {registry}\n\n"
            "Step 3 — Which task pool do you want to assign from?",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🏛 Stamp Duty (national, from_ardhipay=false)",
                                      callback_data="rt_type:STAMP_DUTY")],
                [InlineKeyboardButton("🏙 County Stamp Duty (from_ardhipay=true)",
                                      callback_data="rt_type:COUNTY_STAMP_DUTY")],
            ]),
        )
        return RS.TASK_TYPE

    # Only VALUER — go straight to task count
    rt.task_type = task_type
    type_label   = "County Stamp Duty" if task_type == "COUNTY_STAMP_DUTY" else "Stamp Duty"
    await query.edit_message_text(
        f"✅ *Staff Validated*\n\n"
        f"*Name:* {name}\n"
        f"*Task Type:* {type_label}\n"
        f"*Roles:* {role_labels}\n\n"
        "Step 3 — How many tasks do you want to assign?",
        parse_mode="Markdown",
    )
    return RS.TASK_COUNT


async def recv_rt_task_type(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handles the Stamp Duty / County Stamp Duty choice."""
    query = update.callback_query
    await query.answer()
    rt    = _get_rt(ctx)
    rt.task_type = query.data.split(":")[1]   # "STAMP_DUTY" or "COUNTY_STAMP_DUTY"

    type_label  = "County Stamp Duty" if rt.task_type == "COUNTY_STAMP_DUTY" else "Stamp Duty"
    county_line = (
        f"\n*County:* {rt.staff_county}  |  *Registry:* {rt.staff_registry}"
        if rt.task_type == "COUNTY_STAMP_DUTY" else ""
    )
    await query.edit_message_text(
        f"✅ *Task pool:* {type_label}{county_line}\n\n"
        "Step 3 — How many tasks do you want to assign?",
        parse_mode="Markdown",
    )
    return RS.TASK_COUNT


async def recv_rt_task_count(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    rt = _get_rt(ctx)
    try:
        count = int(update.message.text.strip())
        if count <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("⚠️ Please enter a positive whole number.")
        return RS.TASK_COUNT

    rt.task_count = count
    await update.message.reply_text(
        f"✅ *{count} task(s)* requested.\n\nStep 4 — Filter by consideration amount?",
        parse_mode="Markdown",
        reply_markup=_ft_amount_keyboard(),
    )
    return RS.AMOUNT_RANGE


async def recv_rt_amount_choice(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handles amount preset button on the amount range step."""
    query  = update.callback_query
    await query.answer()
    rt     = _get_rt(ctx)
    choice = query.data  # e.g. "ft_amount:1m_5m" or "ft_amount:custom"

    if choice == "ft_amount:custom":
        await query.edit_message_text(
            "✏️ Enter custom amount range as *min max* (e.g. `500000 5000000`).\n"
            "Or send just one number as max.",
            parse_mode="Markdown",
        )
        return RS.AMOUNT_TEXT

    ranges = {
        "ft_amount:0_1m":    (0.0,           1_000_000.0),
        "ft_amount:1m_5m":   (1_000_000.0,   5_000_000.0),
        "ft_amount:5m_10m":  (5_000_000.0,  10_000_000.0),
        "ft_amount:20m_50m": (20_000_000.0, 50_000_000.0),
        "ft_amount:50m_100m":(50_000_000.0,100_000_000.0),
        "ft_amount:80m_300m":(80_000_000.0,300_000_000.0),
        "ft_amount:80m_3b":  (80_000_000.0,  3_000_000_000.0),
        "ft_amount:all":     (None,           None),
    }
    rt.amount_min, rt.amount_max = ranges.get(choice, (None, None))

    await query.edit_message_text(
        "Step 5 — Run now or set up a recurring schedule?",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("▶️ Run Now",              callback_data="rt_sched:now")],
            [InlineKeyboardButton("⏰ Schedule (repeating)", callback_data="rt_sched:schedule")],
        ]),
    )
    return RS.SCHEDULE_CHOICE


async def recv_rt_amount_range(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handles the typed custom amount range."""
    rt    = _get_rt(ctx)
    text  = update.message.text.strip()
    parts = text.split()
    try:
        if len(parts) == 2:
            rt.amount_min = float(parts[0].replace(",", ""))
            rt.amount_max = float(parts[1].replace(",", ""))
        elif len(parts) == 1:
            rt.amount_min = None
            rt.amount_max = float(parts[0].replace(",", ""))
        else:
            raise ValueError
    except ValueError:
        await update.message.reply_text(
            "❌ Could not parse. Enter two numbers e.g. `500000 5000000`, or one number as max.",
            parse_mode="Markdown",
        )
        return RS.AMOUNT_TEXT

    await update.message.reply_text(
        "Step 5 — Run now or set up a recurring schedule?",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("▶️ Run Now",              callback_data="rt_sched:now")],
            [InlineKeyboardButton("⏰ Schedule (repeating)", callback_data="rt_sched:schedule")],
        ]),
    )
    return RS.SCHEDULE_CHOICE


async def recv_rt_schedule_choice(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    rt     = _get_rt(ctx)
    choice = query.data.split(":")[1]

    if choice == "now":
        await query.edit_message_text("▶️ Running now…")
        return await _rt_fetch_and_show(query.message, rt)

    await query.edit_message_text(
        "⏰ *Schedule Setup*\n\n"
        "Enter the repeat interval in minutes:\n"
        "_(e.g._ `60` _= every hour,_ `1440` _= daily)_",
        parse_mode="Markdown",
    )
    return RS.SCHEDULE_INTERVAL


async def recv_rt_schedule_interval(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    rt = _get_rt(ctx)
    try:
        minutes = int(update.message.text.strip())
        if minutes < 1:
            raise ValueError
    except ValueError:
        await update.message.reply_text("⚠️ Enter a positive number of minutes.")
        return RS.SCHEDULE_INTERVAL

    rt.schedule_interval_minutes = minutes
    await update.message.reply_text(
        f"⏰ Will repeat every *{minutes} minute(s)*.\n\nFetching tasks for preview…",
        parse_mode="Markdown",
    )
    return await _rt_fetch_and_show(update.message, rt)


async def recv_rt_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "rt_confirm:no":
        await query.edit_message_text("❌ Receive tasks cancelled.")
        await query.message.reply_text("Use the menu to start again.", reply_markup=_main_menu())
        return ConversationHandler.END

    rt   = _get_rt(ctx)
    sd   = rt.staff_data.get("staff_details", {})
    name = " ".join(filter(None, [sd.get("firstname"), sd.get("middlename"), sd.get("lastname")]))
    uid  = sd.get("user_id", rt.staff_data.get("id", "?"))

    await query.edit_message_text(
        f"⚙️ Assigning *{len(rt.matched_tasks)}* task(s) to *{name}*…",
        parse_mode="Markdown",
    )

    await _do_assign_tasks(
        ctx.bot, query.message.chat_id,
        rt.session, rt.tokens,
        rt.matched_tasks, uid, name, rt.cred_type,
    )

    # Register repeating schedule if requested
    if rt.schedule_interval_minutes:
        sched = {
            "schedule_id":      str(uuid.uuid4()),
            "staff_uid":        uid,
            "staff_name":       name,
            "cred_type":        rt.cred_type,
            "task_count":       rt.task_count,
            "amount_min":       rt.amount_min,
            "amount_max":       rt.amount_max,
            "task_type":        rt.task_type,
            "staff_registry":   rt.staff_registry,
            "staff_county":     rt.staff_county,
            "interval_minutes": rt.schedule_interval_minutes,
            "chat_id":          query.message.chat_id,
            "active":           True,
        }
        persist_schedule(sched)
        ctx.job_queue.run_repeating(
            _receive_tasks_job,
            interval=rt.schedule_interval_minutes * 60,
            first=rt.schedule_interval_minutes * 60,
            data={"chat_id": query.message.chat_id, "schedule": sched},
            name=f"rt_{sched['schedule_id']}",
        )
        await query.message.reply_text(
            f"⏰ *Schedule saved:* runs every *{rt.schedule_interval_minutes} minute(s)*.\n"
            f"Use */schedules* to view active schedules.",
            parse_mode="Markdown",
            reply_markup=_main_menu(),
        )
    else:
        await query.message.reply_text("Use the menu to start again.", reply_markup=_main_menu())

    return ConversationHandler.END


# ──────────────────────────────────────────────────────────
# /schedules — list saved schedules
# ──────────────────────────────────────────────────────────

async def cmd_schedules(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    schedules = [s for s in load_schedules() if s.get("active", True)]
    if not schedules:
        await update.message.reply_text("📭 No active schedules.", reply_markup=_main_menu())
        return
    lines = []
    for s in schedules:
        range_str = (
            f"KES {s['amount_min']:,.0f}–{s['amount_max']:,.0f}"
            if s.get("amount_min") is not None else "any amount"
        )
        lines.append(
            f"• *{s['staff_name']}* — every *{s['interval_minutes']}min*\n"
            f"  Tasks: {s['task_count']} | {s['task_type']} | {range_str}\n"
            f"  ID: `{s['schedule_id'][:8]}…`"
        )
    await update.message.reply_text(
        "⏰ *Active Schedules:*\n\n" + "\n\n".join(lines),
        parse_mode="Markdown",
        reply_markup=_main_menu(),
    )


# ──────────────────────────────────────────────────────────
# /task_batches — view saved task batch history
# ──────────────────────────────────────────────────────────

async def cmd_task_batches(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    batches = load_task_batches()
    if not batches:
        await update.message.reply_text("📭 No saved task batches yet.", reply_markup=_main_menu())
        return
    lines = []
    for b in batches[-10:]:   # most recent 10
        lines.append(
            f"• *{b['staff_name']}*  —  {b['created_at']}\n"
            f"  Assigned: {len(b['tasks'])}  |  Failed: {len(b.get('failed', []))}"
        )
    await update.message.reply_text(
        "📦 *Recent Task Batches (last 10):*\n\n" + "\n\n".join(lines),
        parse_mode="Markdown",
        reply_markup=_main_menu(),
    )


# ──────────────────────────────────────────────────────────
# AF Results — view auto fetch run history
# ──────────────────────────────────────────────────────────

async def cmd_af_results(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    results = load_af_results()
    if not results:
        await update.message.reply_text(
            "📭 No Auto Fetch runs recorded yet. The history is saved once the Auto Fetch schedule runs.",
            reply_markup=_main_menu(),
        )
        return

    # Show last 10 runs as inline buttons (newest first)
    rows = []
    for r in reversed(results[-10:]):
        label = f"{r['run_at']}  ({r['count']} task{'s' if r['count'] != 1 else ''})"
        rows.append([InlineKeyboardButton(label, callback_data=f"af_result:{r['run_id']}")])

    await update.message.reply_text(
        "🗂 *Auto Fetch History* (last 10 runs)\n\nTap a run to see its tasks:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(rows),
    )


async def recv_af_result_detail(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Show tasks for a specific AF run, marking already-assigned refs."""
    query = update.callback_query
    await query.answer()
    run_id = query.data.split(":", 1)[1]

    results  = load_af_results()
    run      = next((r for r in results if r["run_id"] == run_id), None)
    if not run:
        await query.edit_message_text("❌ Run not found (may have been pruned).")
        return

    assigned = load_saved_assignments()   # {ref: {valuer_name, ...}}
    tasks    = run["tasks"]
    f        = run["filters"]

    lo_s  = f"KES {int(f['amount_min']):,}" if f.get("amount_min") is not None else "0"
    hi_s  = f"KES {int(f['amount_max']):,}" if f.get("amount_max") is not None else "∞"
    sec_label = {"exclude": "No Sectional", "only": "Sectional Only", "all": "All"}.get(
        f.get("sectional_filter", "exclude"), f.get("sectional_filter", "exclude")
    )
    header = (
        f"🗂 *AF Run — {run['run_at']}*\n"
        f"Tasks found: *{run['count']}*\n"
        f"Filters: County={f.get('county','All') or 'All'} | Registry={f.get('registry','All') or 'All'}\n"
        f"Amount: {lo_s}–{hi_s} | {sec_label}\n\n"
    )

    if not tasks:
        await query.edit_message_text(header + "_No tasks found in this run._", parse_mode="Markdown")
        return

    lines = []
    pending = 0
    for i, t in enumerate(tasks, 1):
        ref    = t.get("ref", "—")
        parcel = t.get("parcel", "—")
        cnty   = (t.get("county") or "—").upper()
        reg    = (t.get("registry") or "—").upper()
        date   = t.get("date_created", "")
        try:
            raw = t.get("consideration")
            cons = f"KES {int(float(str(raw).replace(',','').strip())):,}" if raw else "—"
        except (ValueError, TypeError):
            cons = str(t.get("consideration") or "—")

        if ref in assigned:
            status = f"✅ assigned to {assigned[ref].get('valuer_name', '?')}"
        else:
            status = "⏳ pending"
            pending += 1

        lines.append(
            f"{i}. `{ref}` — {status}\n"
            f"   {cnty}/{reg} | {date} | {cons}\n"
            f"   {parcel}"
        )

    summary_line = f"*Pending: {pending}  |  Assigned: {len(tasks) - pending}*\n\n"

    # Split into chunks if needed
    chunks, chunk = [], header + summary_line
    for line in lines:
        candidate = chunk + "\n\n" + line
        if len(candidate) > 4000:
            chunks.append(chunk)
            chunk = line
        else:
            chunk = candidate
    if chunk:
        chunks.append(chunk)

    for c in chunks:
        await query.message.reply_text(c, parse_mode="Markdown")


# ──────────────────────────────────────────────────────────
# Token refresh daemon — process management
# ──────────────────────────────────────────────────────────

def _daemon_read_pid() -> Optional[int]:
    """Return the PID stored in the pid file, or None if missing/invalid."""
    try:
        with open(DAEMON_PID_FILE) as f:
            return int(f.read().strip())
    except (FileNotFoundError, ValueError):
        return None


def _daemon_running() -> bool:
    """Return True if the daemon process is alive."""
    pid = _daemon_read_pid()
    if pid is None:
        return False
    try:
        os.kill(pid, 0)   # signal 0 = probe only, no actual signal sent
        return True
    except (ProcessLookupError, PermissionError):
        return False


def _daemon_start() -> Tuple[bool, str]:
    """
    Launch token_refresh_daemon.py in a detached process.
    Uses start_new_session=True so it is not killed when the bot receives SIGTERM.
    stdout/stderr are appended to DAEMON_LOG_FILE.
    Returns (success, message).
    """
    if _daemon_running():
        return False, f"Already running (PID {_daemon_read_pid()})."

    if not os.path.exists(DAEMON_SCRIPT):
        return False, f"Script not found: {DAEMON_SCRIPT}"

    _ensure_data_dir()
    log_fh = open(DAEMON_LOG_FILE, "a")
    proc = subprocess.Popen(
        [sys.executable, "-u", DAEMON_SCRIPT],
        stdout=log_fh,
        stderr=log_fh,
        stdin=subprocess.DEVNULL,
        start_new_session=True,   # detach from bot's process group
        close_fds=True,
    )
    with open(DAEMON_PID_FILE, "w") as f:
        f.write(str(proc.pid))
    logger.info("Token refresh daemon started (PID %d)", proc.pid)
    return True, f"Started (PID {proc.pid}). Logs → `{DAEMON_LOG_FILE}`"


def _daemon_stop() -> Tuple[bool, str]:
    """Send SIGTERM to the daemon. Returns (success, message)."""
    if not _daemon_running():
        return False, "Daemon is not running."
    pid = _daemon_read_pid()
    try:
        os.kill(pid, signal.SIGTERM)
        # Give it a moment, then confirm it's gone
        for _ in range(10):
            time.sleep(0.3)
            try:
                os.kill(pid, 0)
            except ProcessLookupError:
                break
        try:
            os.remove(DAEMON_PID_FILE)
        except FileNotFoundError:
            pass
        logger.info("Token refresh daemon stopped (PID %d)", pid)
        return True, f"Daemon (PID {pid}) stopped."
    except Exception as e:
        return False, f"Failed to stop daemon: {e}"


def _daemon_status_text() -> str:
    running = _daemon_running()
    pid     = _daemon_read_pid()
    if running:
        return f"🟢 *Running* (PID {pid})"
    elif pid:
        return "🔴 *Not running* (stale PID file — process died)"
    else:
        return "🔴 *Not running*"


def _daemon_keyboard() -> InlineKeyboardMarkup:
    running = _daemon_running()
    rows = []
    if running:
        rows.append([InlineKeyboardButton("⏹ Stop Daemon",   callback_data="daemon:stop")])
    else:
        rows.append([InlineKeyboardButton("▶️ Start Daemon",  callback_data="daemon:start")])
    rows.append([InlineKeyboardButton("🔁 Refresh Status", callback_data="daemon:status")])
    return InlineKeyboardMarkup(rows)


async def cmd_daemon(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    await update.message.reply_text(
        f"🔄 *Token Refresh Daemon*\n\n"
        f"Status: {_daemon_status_text()}\n\n"
        f"The daemon watches the token cache and silently refreshes "
        f"each token *5 minutes before it expires*.\n"
        f"Logs are written to `data/daemon.log`.",
        parse_mode="Markdown",
        reply_markup=_daemon_keyboard(),
    )


async def cmd_token_status(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)

    raw   = _load_tokens_raw()
    now   = time.time()
    lines = []

    for cred_type, label in CRED_LABELS.items():
        entry = raw.get(cred_type)
        if not entry:
            lines.append(f"{label}\n  ⚫ No token cached")
            continue

        exp = entry.get("expires_at") or _jwt_exp(entry.get("jwt", ""))
        if not exp:
            lines.append(f"{label}\n  ⚠️ Expiry unreadable")
            continue

        secs_left = exp - now
        exp_str   = datetime.fromtimestamp(exp, tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        if secs_left <= 0:
            lines.append(f"{label}\n  🔴 Expired — {exp_str}")
        elif secs_left < 10 * 60:
            mins = int(secs_left // 60)
            lines.append(f"{label}\n  🟡 Expires in {mins}m — {exp_str}")
        else:
            hrs  = int(secs_left // 3600)
            mins = int((secs_left % 3600) // 60)
            time_str = f"{hrs}h {mins}m" if hrs else f"{mins}m"
            lines.append(f"{label}\n  🟢 Valid — expires in {time_str} ({exp_str})")

    text = "🔒 *Token Status*\n\n" + "\n\n".join(lines)
    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=_main_menu())


async def recv_daemon_action(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query  = update.callback_query
    await query.answer()
    action = query.data.split(":")[1]

    if action == "start":
        ok, msg = _daemon_start()
    elif action == "stop":
        ok, msg = _daemon_stop()
    else:
        ok, msg = True, "Status refreshed."

    status = _daemon_status_text()
    await query.edit_message_text(
        f"🔄 *Token Refresh Daemon*\n\n"
        f"Status: {status}\n\n"
        f"{'✅' if ok else '❌'} {msg}",
        parse_mode="Markdown",
        reply_markup=_daemon_keyboard(),
    )


# ──────────────────────────────────────────────────────────
# /cancel
# ──────────────────────────────────────────────────────────
async def cmd_cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.pop("session", None)
    await update.message.reply_text(
        "🛑 Flow cancelled.",
        reply_markup=_main_menu(),
    )
    return ConversationHandler.END


async def cmd_restart(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    await update.message.reply_text("🔁 Restarting bot… back in a moment.")
    await asyncio.sleep(2)   # let the message deliver before replacing the process
    os.execv(sys.executable, [sys.executable] + sys.argv)


# ──────────────────────────────────────────────────────────
# Fallback (unexpected input during conversation)
# ──────────────────────────────────────────────────────────
async def fallback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🤔 I didn't understand that. Follow the steps above, or tap 🛑 Cancel to abort.",
        reply_markup=_main_menu(),
    )


# ──────────────────────────────────────────────────────────
# ──────────────────────────────────────────────────────────
# DLV Batch — persistent queue helpers
# ──────────────────────────────────────────────────────────

def load_dlv_batch() -> List[Dict]:
    try:
        with open(SAVED_DLV_BATCH_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def save_dlv_batch(items: List[Dict]) -> None:
    _ensure_data_dir()
    with open(SAVED_DLV_BATCH_FILE, "w") as f:
        json.dump(items, f, indent=2)


def clear_dlv_batch() -> None:
    save_dlv_batch([])


def _parse_batch_input(text: str) -> List[Dict]:
    """
    Parse lines of format:  "REF1, REF2 : Valuer Name"
    Returns [{refs: [...], valuer_name_raw: "..."}]
    """
    groups = []
    for line in text.strip().splitlines():
        line = line.strip()
        if not line or ":" not in line:
            continue
        ref_part, valuer_part = line.rsplit(":", 1)
        refs = [r.strip().upper() for r in ref_part.split(",") if r.strip()]
        valuer_name_raw = valuer_part.strip()
        if refs and valuer_name_raw:
            groups.append({"refs": refs, "valuer_name_raw": valuer_name_raw})
    return groups


def _resolve_valuer_from_saved(name: str) -> Optional[Dict]:
    """Case-insensitive substring match against saved valuers."""
    name_lower = name.lower()
    for v in load_saved_valuers():
        if name_lower in v["name"].lower():
            return v
    return None


def _search_valuer_api(name: str, tokens: AuthTokens) -> List[Dict]:
    http_sess = build_session()
    resp = http_sess.get(
        f"{BASE_URL}/acl/api/v1/accounts/list-user-accounts",
        headers={"Authorization": f"Bearer {tokens.access_token}", "JWTAUTH": f"Bearer {tokens.jwt}"},
        params={"account_type": "STAFF", "filter_type": "ACTIVE", "page": 1, "search": name},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json().get("results", [])


def _search_ref_dlv(tokens: AuthTokens, ref: str) -> Optional[Dict]:
    """Search for a reference number via DLV endpoint. Returns the task dict or None."""
    http_sess = build_session()
    resp = http_sess.get(
        f"{BASE_URL}/valuationservice/api/v1/stamp-duty/application",
        headers={
            "Authorization": f"Bearer {tokens.access_token}",
            "JWTAUTH":       f"Bearer {tokens.jwt}",
            "cparams":       CPARAMS_DLV,
        },
        params={
            "filter": "Pending", "role": "DLV", "request_type": "STAMP_DUTY",
            "search": ref, "page": 1,
        },
        timeout=30,
    )
    resp.raise_for_status()
    for task in resp.json().get("results", []):
        if task.get("reference_number") == ref:
            return task
    return None


def _fetch_ref_detail_dlv(tokens: AuthTokens, request_id: str) -> Optional[Dict]:
    """Fetch the detail view for a DLV application."""
    http_sess = build_session()
    resp = http_sess.get(
        f"{BASE_URL}/valuationservice/api/v1/stamp-duty/application/detail-view",
        headers={
            "Authorization": f"Bearer {tokens.access_token}",
            "JWTAUTH":       f"Bearer {tokens.jwt}",
            "cparams":       CPARAMS_DLV,
        },
        params={"request_id": request_id},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def _process_dlv_batch_items(tokens: AuthTokens) -> str:
    """
    Process the flat batch queue (list of {ref, valuer_name, valuer_uid, valuer_acct}).
    Refs not found in DLV are kept in the queue for the next 5-minute retry cycle.
    Returns a report string of completed items only, or "" if nothing was processed.
    """
    items = load_dlv_batch()
    if not items:
        return ""

    http_sess = build_session()
    assign_url = f"{BASE_URL}/valuationservice/api/v1/stamp-duty/fix_application_details"
    auth_hdrs = {
        "Authorization": f"Bearer {tokens.access_token}",
        "JWTAUTH":       f"Bearer {tokens.jwt}",
    }

    completed_lines: List[str] = []
    remaining:       List[Dict] = []

    for item in items:
        ref         = item.get("ref", "")
        valuer_name = item.get("valuer_name", "")
        valuer_uid  = item.get("valuer_uid", "")

        try:
            task = _search_ref_dlv(tokens, ref)
            if not task:
                # Not in DLV yet — keep for retry
                item["last_error"] = "Not found in DLV endpoint"
                remaining.append(item)
                continue

            detail = _fetch_ref_detail_dlv(tokens, task["id"])
            if not detail:
                # Transient failure — keep for retry
                item["last_error"] = "Detail fetch returned empty response"
                remaining.append(item)
                continue

            node = detail.get("node", "")

            if node == "VALUATION_STAMP_DUTY_CREATED":
                r = http_sess.post(
                    assign_url, headers=auth_hdrs,
                    json={
                        "reference_number":  ref,
                        "valuation_officer": valuer_uid,
                        "node":              "VALUATION_STAMP_DUTY_VALUER_REPORT",
                    },
                    timeout=30,
                )
                r.raise_for_status()
                completed_lines.append(f"✅ `{ref}` — assigned to *{valuer_name}*")
                persist_assignment(ref, valuer_name, valuer_uid)

            elif node == "VALUATION_STAMP_DUTY_VALUER_REPORT":
                actors = detail.get("actors", [])
                if actors:
                    actor_name = actors[0].get("user_details", {}).get("names", "Unknown")
                    completed_lines.append(f"📋 `{ref}` — already with *{actor_name}*")
                else:
                    completed_lines.append(f"📋 `{ref}` — at VALUER_REPORT stage, no actor listed")

            else:
                completed_lines.append(f"❓ `{ref}` — unexpected node: `{node}`")

        except Exception as e:
            # Keep for retry on error
            item["last_error"] = str(e)[:120]
            remaining.append(item)
            logger.warning("DLV batch error for %s: %s", ref, e)

    # Save only refs that still need processing
    save_dlv_batch(remaining)

    if remaining:
        pending_refs = ", ".join(f"`{i['ref']}`" for i in remaining)
        completed_lines.append(f"\n⏳ Still pending (retry in 5 min): {pending_refs}")

    return "\n".join(completed_lines)


async def _dlv_batch_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    """PTB repeating job: every 5 minutes, process the DLV batch queue if non-empty."""
    items = load_dlv_batch()
    if not items:
        return

    tokens = _any_valid_tokens()
    if not tokens:
        logger.warning("DLV batch job: no valid tokens — skipping cycle.")
        return

    before_count = len(items)
    report = _process_dlv_batch_items(tokens)
    after_count  = len(load_dlv_batch())

    # Only notify if something was actually completed (queue shrank)
    if report and after_count < before_count:
        msg = f"📋 *DLV Batch (auto)*\n{report}"
        if len(msg) > 4000:
            msg = msg[:4000] + "\n…_(truncated)_"
        for chat_id in ALLOWED_IDS:
            try:
                await context.bot.send_message(chat_id, msg, parse_mode="Markdown")
            except Exception as e:
                logger.warning("DLV batch job notify error for %s: %s", chat_id, e)


# ──────────────────────────────────────────────────────────
# Auto Fetch — scheduled periodic fetch + notify
# ──────────────────────────────────────────────────────────

_AF_INTERVALS = [
    ("15 min",  15),
    ("30 min",  30),
    ("1 hr",    60),
    ("2 hr",   120),
    ("4 hr",   240),
    ("6 hr",   360),
    ("12 hr",  720),
    ("24 hr", 1440),
]


def load_auto_fetch_schedule() -> Optional[Dict]:
    try:
        with open(SAVED_AUTO_FETCH_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return None


def save_auto_fetch_schedule(cfg: Dict) -> None:
    _ensure_data_dir()
    with open(SAVED_AUTO_FETCH_FILE, "w") as f:
        json.dump(cfg, f, indent=2)


def clear_auto_fetch_schedule() -> None:
    try:
        os.remove(SAVED_AUTO_FETCH_FILE)
    except FileNotFoundError:
        pass


# ── Auto Fetch result history ──────────────────────────────
_AF_RESULTS_KEEP = 20   # keep last N runs


def load_af_results() -> List[Dict]:
    try:
        with open(SAVED_AF_RESULTS_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def persist_af_result(run_id: str, run_at: str, tasks: List[Dict], cfg: Dict) -> None:
    _ensure_data_dir()
    results = load_af_results()
    results.append({
        "run_id":   run_id,
        "run_at":   run_at,
        "count":    len(tasks),
        "filters": {
            "county":           cfg.get("county_filter", ""),
            "registry":         cfg.get("registry_filter", ""),
            "amount_min":       cfg.get("amount_min"),
            "amount_max":       cfg.get("amount_max"),
            "days_back":        cfg.get("days_back", 2),
            "sectional_filter": cfg.get("sectional_filter", "exclude"),
        },
        "tasks": [
            {
                "ref":          t.get("reference_number", ""),
                "parcel":       t.get("parcel_number", ""),
                "county":       t.get("county", ""),
                "registry":     t.get("registry", ""),
                "consideration": t.get("consideration"),
                "date_created": (t.get("date_created") or "")[:10],
                "source":       t.get("source", ""),
            }
            for t in tasks
        ],
    })
    # Trim to keep only the most recent runs
    if len(results) > _AF_RESULTS_KEEP:
        results = results[-_AF_RESULTS_KEEP:]
    with open(SAVED_AF_RESULTS_FILE, "w") as f:
        json.dump(results, f, indent=2)


def _af_interval_keyboard() -> InlineKeyboardMarkup:
    rows = []
    row  = []
    for label, mins in _AF_INTERVALS:
        row.append(InlineKeyboardButton(label, callback_data=f"af:interval:{mins}"))
        if len(row) == 4:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton("❌ Cancel / Stop schedule", callback_data="af:cancel")])
    return InlineKeyboardMarkup(rows)


async def cmd_auto_fetch(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)

    cfg = load_auto_fetch_schedule()
    if cfg:
        mins    = cfg.get("interval_minutes", 0)
        lo      = cfg.get("amount_min")
        hi      = cfg.get("amount_max")
        lo_s    = f"KES {int(lo):,}" if lo is not None else "0"
        hi_s    = f"KES {int(hi):,}" if hi is not None else "∞"
        county    = cfg.get("county_filter", "") or "All"
        reg       = cfg.get("registry_filter", "") or "All"
        days      = cfg.get("days_back", 2)
        sec       = {"exclude": "Exclude Sectional", "only": "Sectional Only", "all": "All"}.get(
                        cfg.get("sectional_filter", "exclude"), "Exclude Sectional")
        email_s   = cfg.get("email") or "Telegram only"
        status  = (
            f"⏰ *Auto Fetch is active*\n"
            f"Interval: every {mins} min | Days back: {days}\n"
            f"County: {county.title()} | Registry: {reg.title()}\n"
            f"Amount: {lo_s} – {hi_s} | Sectional: {sec}\n"
            f"Email: {email_s}\n\n"
            f"Choose a new interval to update, or cancel to stop."
        )
    else:
        status = (
            "⏰ *Auto Fetch*\n\n"
            "Periodically fetches tasks and sends results here.\n"
            "Choose how often to run:"
        )

    await update.message.reply_text(
        status,
        parse_mode="Markdown",
        reply_markup=_af_interval_keyboard(),
    )
    return AF.INTERVAL


async def recv_af_interval(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "af:cancel":
        for job in ctx.job_queue.get_jobs_by_name("auto_fetch_job"):
            job.schedule_removal()
        clear_auto_fetch_schedule()
        await query.edit_message_text("🛑 Auto Fetch schedule cancelled.")
        await query.message.reply_text("Main menu:", reply_markup=_main_menu())
        return ConversationHandler.END

    try:
        interval_minutes = int(query.data.split(":")[-1])
    except ValueError:
        return AF.INTERVAL

    ctx.user_data["af_interval_minutes"] = interval_minutes
    await query.edit_message_text(
        f"✅ Interval: *{interval_minutes} min*\n\nHow many days back to fetch?",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [
                InlineKeyboardButton("1 day",  callback_data="af_days:1"),
                InlineKeyboardButton("2 days", callback_data="af_days:2"),
                InlineKeyboardButton("3 days", callback_data="af_days:3"),
            ],
            [
                InlineKeyboardButton("5 days",  callback_data="af_days:5"),
                InlineKeyboardButton("7 days",  callback_data="af_days:7"),
                InlineKeyboardButton("10 days", callback_data="af_days:10"),
            ],
        ]),
    )
    return AF.DAYS_BACK


async def recv_af_days(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    ctx.user_data["af_days_back"] = int(query.data.split(":")[-1])
    await query.edit_message_text(
        f"✅ Days back: *{ctx.user_data['af_days_back']}*\n\nFilter by county?",
        parse_mode="Markdown",
        reply_markup=_ft_county_keyboard(),
    )
    return AF.COUNTY


async def recv_af_county(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    ctx.user_data["af_county"] = "" if query.data == "ft_county:all" else query.data.split(":")[1]
    label = ctx.user_data["af_county"].title() or "All Counties"
    await query.edit_message_text(
        f"✅ County: *{label}*\n\nFilter by registry?",
        parse_mode="Markdown",
        reply_markup=_ft_registry_keyboard(),
    )
    return AF.REGISTRY


async def recv_af_registry(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    ctx.user_data["af_registry"] = "" if query.data == "ft_registry:all" else query.data.split(":")[1]
    reg_label    = ctx.user_data["af_registry"].title() or "All Registries"
    county_label = ctx.user_data.get("af_county", "").title() or "All Counties"
    await query.edit_message_text(
        f"✅ County: *{county_label}* | Registry: *{reg_label}*\n\nFilter by amount?",
        parse_mode="Markdown",
        reply_markup=_ft_amount_keyboard(),
    )
    return AF.AMOUNT


async def recv_af_amount(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Callback handler — user picked an amount preset or Custom."""
    query = update.callback_query
    await query.answer()
    choice = query.data  # e.g. "ft_amount:1m_5m" or "ft_amount:custom"

    if choice == "ft_amount:custom":
        await query.edit_message_text(
            "✏️ Enter custom amount range as *min max* (e.g. `500000 5000000`).\n"
            "Or send just one number as max.",
            parse_mode="Markdown",
        )
        return AF.AMOUNT_TEXT

    ranges = {
        "ft_amount:0_1m":    (0.0,           1_000_000.0),
        "ft_amount:1m_5m":   (1_000_000.0,   5_000_000.0),
        "ft_amount:5m_10m":  (5_000_000.0,  10_000_000.0),
        "ft_amount:20m_50m": (20_000_000.0, 50_000_000.0),
        "ft_amount:50m_100m":(50_000_000.0,100_000_000.0),
        "ft_amount:80m_300m":(80_000_000.0,300_000_000.0),
        "ft_amount:80m_3b":  (80_000_000.0,  3_000_000_000.0),
        "ft_amount:all":     (None,           None),
    }
    amount_min, amount_max = ranges.get(choice, (None, None))
    ctx.user_data["af_amount_min"] = amount_min
    ctx.user_data["af_amount_max"] = amount_max

    await query.edit_message_text(
        "Include sectional properties?\n_(Sectional: parcel has 4 parts e.g. Nairobi/Block12/345/888)_",
        parse_mode="Markdown",
        reply_markup=_sectional_keyboard(),
    )
    return AF.SECTIONAL


async def recv_af_amount_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Text handler for custom amount entry in Auto Fetch."""
    text = update.message.text.strip()
    parts = text.split()
    try:
        if len(parts) == 2:
            amount_min = float(parts[0].replace(",", ""))
            amount_max = float(parts[1].replace(",", ""))
        elif len(parts) == 1:
            amount_min = None
            amount_max = float(parts[0].replace(",", ""))
        else:
            raise ValueError
    except ValueError:
        await update.message.reply_text(
            "❌ Could not parse. Enter two numbers e.g. `500000 5000000`, or one number as max.",
            parse_mode="Markdown",
        )
        return AF.AMOUNT_TEXT

    ctx.user_data["af_amount_min"] = amount_min
    ctx.user_data["af_amount_max"] = amount_max
    await update.message.reply_text(
        "Include sectional properties?\n_(Sectional: parcel has 4 parts e.g. Nairobi/Block12/345/888)_",
        parse_mode="Markdown",
        reply_markup=_sectional_keyboard(),
    )
    return AF.SECTIONAL


async def recv_af_sectional(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    ctx.user_data["af_sectional"] = query.data.split(":")[1]  # "exclude" | "only" | "all"

    existing_email = ""
    cfg = load_auto_fetch_schedule()
    if cfg:
        existing_email = cfg.get("email", "")

    hint = f"Current: `{existing_email}`\n\n" if existing_email else ""
    await query.edit_message_text(
        f"📧 *Send results to email?*\n\n"
        f"{hint}"
        f"Enter an email address, or send `skip` to notify via Telegram only.",
        parse_mode="Markdown",
    )
    return AF.EMAIL


async def recv_af_email(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()

    if text.lower() == "skip":
        email = ""
    else:
        # Basic email validation
        if "@" not in text or "." not in text.split("@")[-1]:
            await update.message.reply_text(
                "❌ Invalid email address. Enter a valid email or send `skip`.",
                parse_mode="Markdown",
            )
            return AF.EMAIL
        email = text

    interval   = ctx.user_data.get("af_interval_minutes", 60)
    days       = ctx.user_data.get("af_days_back", 2)
    county     = ctx.user_data.get("af_county", "")
    registry   = ctx.user_data.get("af_registry", "")
    amount_min = ctx.user_data.get("af_amount_min")
    amount_max = ctx.user_data.get("af_amount_max")
    sectional  = ctx.user_data.get("af_sectional", "exclude")

    cfg = {
        "interval_minutes": interval,
        "days_back":        days,
        "county_filter":    county,
        "registry_filter":  registry,
        "amount_min":       amount_min,
        "amount_max":       amount_max,
        "sectional_filter": sectional,
        "email":            email,
    }
    save_auto_fetch_schedule(cfg)

    for job in ctx.job_queue.get_jobs_by_name("auto_fetch_job"):
        job.schedule_removal()
    ctx.job_queue.run_repeating(
        _auto_fetch_job,
        interval=interval * 60,
        first=interval * 60,
        name="auto_fetch_job",
    )

    lo_s      = f"KES {int(amount_min):,}" if amount_min is not None else "0"
    hi_s      = f"KES {int(amount_max):,}" if amount_max is not None else "∞"
    co_label  = county.title() or "All"
    re_label  = registry.title() or "All"
    sec_label = {"exclude": "Exclude Sectional", "only": "Sectional Only", "all": "All"}.get(sectional, sectional)
    email_label = email or "Telegram only"
    await update.message.reply_text(
        f"✅ *Auto Fetch scheduled*\n"
        f"Every *{interval} min* | Days back: *{days}*\n"
        f"County: *{co_label}* | Registry: *{re_label}*\n"
        f"Amount: {lo_s} – {hi_s} | Sectional: *{sec_label}*\n"
        f"Email: *{email_label}*\n"
        f"First run in {interval} min.",
        parse_mode="Markdown",
        reply_markup=_main_menu(),
    )
    return ConversationHandler.END


def _send_auto_fetch_email(to_email: str, subject: str, body: str) -> None:
    """Send Auto Fetch results via SMTP. Raises on failure."""
    if not SMTP_USER or not SMTP_PASS:
        raise RuntimeError("SMTP_USER / SMTP_PASS not configured in .env")

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = SMTP_USER
    msg["To"]      = to_email

    # Plain text part
    msg.attach(MIMEText(body, "plain"))

    # Simple HTML version
    html_body = "<pre style='font-family:monospace'>" + body.replace("&", "&amp;").replace("<", "&lt;") + "</pre>"
    msg.attach(MIMEText(html_body, "html"))

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.ehlo()
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_USER, to_email, msg.as_string())


async def _auto_fetch_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Background job: fetch tasks with saved schedule settings and notify."""
    cfg = load_auto_fetch_schedule()
    if not cfg:
        return

    tokens = _any_valid_tokens()
    if not tokens:
        logger.warning("Auto Fetch job: no valid tokens — skipping cycle.")
        return

    days_back       = cfg.get("days_back", 2)
    county_filter   = cfg.get("county_filter", "")
    registry_filter = cfg.get("registry_filter", "")
    amount_min      = cfg.get("amount_min")
    amount_max      = cfg.get("amount_max")

    try:
        tasks, stats = _load_fetch_tasks(tokens, days_back)
    except Exception as e:
        logger.warning("Auto Fetch job: fetch failed: %s", e)
        return

    # County filter
    if county_filter:
        tasks = [t for t in tasks if county_filter in (t.get("county") or "").strip().lower()]

    # Registry filter
    if registry_filter:
        tasks = [t for t in tasks if registry_filter in (t.get("registry") or "").strip().lower()]

    # Amount filter
    if amount_min is not None or amount_max is not None:
        def _in_range(t):
            raw = t.get("consideration")
            if raw is None:
                return False
            try:
                val = float(str(raw).replace(",", "").strip())
            except (ValueError, TypeError):
                return False
            if amount_min is not None and val < amount_min:
                return False
            if amount_max is not None and val > amount_max:
                return False
            return True
        tasks = [t for t in tasks if _in_range(t)]

    # Sectional filter
    sf = cfg.get("sectional_filter", "exclude")
    if sf == "exclude":
        tasks = [t for t in tasks if str(t.get("parcel_number") or "").count("/") < 3]
    elif sf == "only":
        tasks = [t for t in tasks if str(t.get("parcel_number") or "").count("/") >= 3]
    # "all" → no filter

    # Exclude already-queued refs
    queued_refs = {item.get("ref", "") for item in load_dlv_batch()}
    tasks = [t for t in tasks if t.get("reference_number", "") not in queued_refs]

    # Persist result history (record even if empty so the run appears in AF Results)
    _af_run_id = str(uuid.uuid4())[:8]
    _af_run_at = time.strftime("%Y-%m-%d %H:%M:%S")
    persist_af_result(_af_run_id, _af_run_at, tasks, cfg)

    if not tasks:
        logger.info("Auto Fetch job: no new tasks after filters.")
        return

    lo_s     = f"KES {int(amount_min):,}" if amount_min is not None else "0"
    hi_s     = f"KES {int(amount_max):,}" if amount_max is not None else "∞"
    co_label  = county_filter.title() or "All"
    re_label  = registry_filter.title() or "All"
    sec_label = {"exclude": "No Sectional", "only": "Sectional Only", "all": "All"}.get(sf, sf)
    header    = (
        f"⏰ *Auto Fetch — {len(tasks)} task(s)*\n"
        f"Days: {days_back} | County: {co_label} | Registry: {re_label} | Amount: {lo_s}–{hi_s} | {sec_label}\n\n"
    )

    lines = []
    for i, t in enumerate(tasks, 1):
        src    = t.get("source", "")
        ref    = t.get("reference_number", "—")
        cnty   = (t.get("county") or "—").upper()
        reg    = (t.get("registry") or "—").upper()
        date   = (t.get("date_created") or "")[:10]
        parcel = t.get("parcel_number") or "—"
        try:
            raw_cons = t.get("consideration")
            cons = f"KES {int(float(str(raw_cons).replace(',','').strip())):,}" if raw_cons else "—"
        except (ValueError, TypeError):
            cons = str(t.get("consideration") or "—")
        lines.append(
            f"{i}. [{src}] {ref}\n"
            f"   {cnty} / {reg} | {date}\n"
            f"   {cons} | {parcel}"
        )

    # Split into chunks
    chunks = []
    chunk  = header
    for line in lines:
        candidate = (chunk + "\n\n" + line).strip()
        if len(candidate) > 4000:
            chunks.append(chunk)
            chunk = line
        else:
            chunk = candidate
    if chunk:
        chunks.append(chunk)

    for chat_id in ALLOWED_IDS:
        for c in chunks:
            try:
                await context.bot.send_message(chat_id, c, parse_mode="Markdown")
            except Exception as e:
                logger.warning("Auto Fetch notify error for %s: %s", chat_id, e)

    # Email notification
    email = cfg.get("email", "")
    if email:
        # Build plain-text body (strip Markdown asterisks)
        plain_header = (
            f"Auto Fetch — {len(tasks)} task(s)\n"
            f"Days: {days_back} | County: {co_label} | Registry: {re_label} | Amount: {lo_s}–{hi_s} | {sec_label}\n"
            + "─" * 60 + "\n\n"
        )
        plain_lines = []
        for i, t in enumerate(tasks, 1):
            src    = t.get("source", "")
            ref    = t.get("reference_number", "—")
            cnty   = (t.get("county") or "—").upper()
            reg    = (t.get("registry") or "—").upper()
            date   = (t.get("date_created") or "")[:10]
            parcel = t.get("parcel_number") or "—"
            try:
                raw_cons = t.get("consideration")
                cons = f"KES {int(float(str(raw_cons).replace(',','').strip())):,}" if raw_cons else "—"
            except (ValueError, TypeError):
                cons = str(t.get("consideration") or "—")
            plain_lines.append(
                f"{i}. [{src}] {ref}\n"
                f"   {cnty} / {reg} | {date}\n"
                f"   {cons} | {parcel}"
            )
        plain_body   = plain_header + "\n\n".join(plain_lines)
        email_subject = f"Auto Fetch — {len(tasks)} task(s) found"
        try:
            _send_auto_fetch_email(email, email_subject, plain_body)
            logger.info("Auto Fetch email sent to %s", email)
        except Exception as e:
            logger.warning("Auto Fetch email failed: %s", e)


# ──────────────────────────────────────────────────────────
# DLV Queue viewer
# ──────────────────────────────────────────────────────────

# Current DLV batch job interval in seconds (default 5 min); updated by user choice
_dlv_batch_interval: int = 300


def _dlv_queue_keyboard(current_interval: int) -> InlineKeyboardMarkup:
    options = [
        ("1 min",       60),
        ("2 min",       120),
        ("3 min",       180),
        ("Default (5 min)", 300),
    ]
    interval_row = [
        InlineKeyboardButton(
            f"{'✅ ' if current_interval == secs else ''}{label}",
            callback_data=f"dlvq:interval:{secs}",
        )
        for label, secs in options
    ]
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("▶️ Query Now", callback_data="dlvq:now")],
        interval_row,
        [InlineKeyboardButton("❌ Cancel",   callback_data="dlvq:cancel")],
    ])


async def cmd_dlv_queue(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)

    items = load_dlv_batch()
    if not items:
        await update.message.reply_text(
            "✅ DLV Queue is empty — no pending assignments.",
            reply_markup=_main_menu(),
        )
        return

    lines = [f"🔍 *DLV Queue* — {len(items)} pending ref(s)\n"]
    for item in items:
        ref         = item.get("ref", "?")
        valuer_name = item.get("valuer_name", "?")
        last_error  = item.get("last_error", "")
        line = f"• `{ref}` → *{valuer_name}*"
        if last_error:
            line += f"\n  ⚠️ _{last_error}_"
        lines.append(line)

    interval_label = {60: "1 min", 120: "2 min", 180: "3 min", 300: "5 min"}.get(
        _dlv_batch_interval, f"{_dlv_batch_interval}s"
    )
    lines.append(f"\n_Current check interval: {interval_label}_")

    msg = "\n".join(lines)
    if len(msg) > 4000:
        msg = msg[:4000] + "\n…_(truncated)_"

    await update.message.reply_text(
        msg,
        parse_mode="Markdown",
        reply_markup=_dlv_queue_keyboard(_dlv_batch_interval),
    )


async def recv_dlv_queue_action(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    global _dlv_batch_interval
    query = update.callback_query
    await query.answer()
    data = query.data  # "dlvq:now" | "dlvq:interval:N" | "dlvq:cancel"

    if data == "dlvq:cancel":
        await query.edit_message_reply_markup(reply_markup=None)
        return

    if data == "dlvq:now":
        tokens = _any_valid_tokens()
        if not tokens:
            await query.edit_message_text(
                "⚠️ No valid tokens — please authenticate first.",
            )
            return
        items_before = load_dlv_batch()
        if not items_before:
            await query.edit_message_text("✅ Queue is empty — nothing to process.")
            return
        report = _process_dlv_batch_items(tokens)
        remaining = load_dlv_batch()
        msg = f"📋 *DLV Queue — Query Result*\n{report}" if report else "ℹ️ Nothing processed."
        if remaining:
            msg += f"\n\n⏳ *{len(remaining)} ref(s) still pending*"
        if len(msg) > 4000:
            msg = msg[:4000] + "\n…_(truncated)_"
        await query.edit_message_text(msg, parse_mode="Markdown")
        return

    if data.startswith("dlvq:interval:"):
        try:
            new_interval = int(data.split(":")[-1])
        except ValueError:
            return
        _dlv_batch_interval = new_interval

        # Reschedule the repeating job with the new interval
        job_queue = ctx.job_queue
        if job_queue:
            existing = job_queue.get_jobs_by_name("dlv_batch_job")
            for job in existing:
                job.schedule_removal()
            job_queue.run_repeating(
                _dlv_batch_job,
                interval=new_interval,
                first=new_interval,
                name="dlv_batch_job",
            )

        label = {60: "1 min", 120: "2 min", 180: "3 min", 300: "5 min"}.get(
            new_interval, f"{new_interval}s"
        )
        await query.edit_message_reply_markup(
            reply_markup=_dlv_queue_keyboard(_dlv_batch_interval),
        )
        await query.message.reply_text(
            f"✅ DLV check interval set to *{label}*.",
            parse_mode="Markdown",
            reply_markup=_main_menu(),
        )


# ──────────────────────────────────────────────────────────
# Implementor Tasks & DLV Tasks — view-only checklist
# ──────────────────────────────────────────────────────────

def _any_valid_tokens() -> Optional[AuthTokens]:
    """Return the first valid cached AuthTokens across all credential profiles."""
    for cred_type in ("staff_valuer", "staff2", "staff", "publicuser"):
        t = get_valid_tokens(cred_type)
        if t:
            return t
    return None


def _date_cutoff_str(days: int) -> str:
    """Return a YYYY-MM-DD cutoff string for N days ago (UTC)."""
    return (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")


def _within_days(date_created: str, cutoff: str) -> bool:
    """
    Compare date_created from the API response against a YYYY-MM-DD cutoff.
    ISO date strings are lexicographically sortable, so simple string compare works.
    If date_created is missing or unparseable, include the task (fail-open).
    """
    if not date_created:
        return True
    # Take only the date part (first 10 chars: YYYY-MM-DD) regardless of time/timezone suffix
    return date_created[:10] >= cutoff


def _ft_headers(tokens: AuthTokens) -> dict:
    return {
        "Authorization": f"Bearer {tokens.access_token}",
        "JWTAUTH":       f"Bearer {tokens.jwt}",
        "cparams":       CPARAMS_SUPPORT,
    }


def _has_stamp_duty_invoice(invoices: list) -> bool:
    """Return True if any invoice is for stamp duty (task already processed)."""
    for inv in invoices:
        pf = str(inv.get("payment_for", "")).lower()
        if "stamp" in pf:
            return True
    return False


def _fetch_hq_list(http_sess, tokens: AuthTokens, cutoff: str) -> List[Dict]:
    """Fetch HQ digitised TRANSFER applications up to cutoff date."""
    headers = _ft_headers(tokens)
    candidates: List[Dict] = []
    page = 1
    stop = False
    while not stop:
        try:
            resp = http_sess.get(
                f"{BASE_URL}/stampdutyservice/api/v1/stamp-duty/hod-or-clr",
                headers=headers,
                params={"filter": "Ongoing", "page": page, "search": ""},
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            logger.warning("HQ list page %d failed: %s", page, e)
            break

        results = data.get("results", [])
        if not results:
            break
        for task in results:
            if task.get("date_created", "")[:10] < cutoff:
                stop = True
                break
            if (task.get("application_type") == "TRANSFER"
                    and task.get("from_ardhipay") is False):
                candidates.append(task)
        if not data.get("next") or stop:
            break
        page += 1

    return candidates


def _fetch_county_list(http_sess, tokens: AuthTokens, cutoff: str) -> List[Dict]:
    """Fetch County undigitised TRANSFER applications up to cutoff date."""
    headers = _ft_headers(tokens)
    candidates: List[Dict] = []
    page = 1
    stop = False
    while not stop:
        try:
            resp = http_sess.get(
                f"{BASE_URL}/stampdutyservice/api/v1/stamp-duty/hod-or-clr",
                headers=headers,
                params={"filter": "Ongoing", "from_ardhipay": "true", "page": page, "search": ""},
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            logger.warning("County list page %d failed: %s", page, e)
            break

        results = data.get("results", [])
        if not results:
            break
        for task in results:
            if task.get("date_created", "")[:10] < cutoff:
                stop = True
                break
            if task.get("application_type") == "TRANSFER":
                candidates.append(task)
        if not data.get("next") or stop:
            break
        page += 1

    return candidates


def _fetch_hq_detail_2a(http_sess, tokens: AuthTokens, application_id: str) -> Optional[Dict]:
    """Fetch registration detail (2a) using application_id."""
    try:
        resp = http_sess.get(
            f"{BASE_URL}/registrationservice/api/v1/transfer/transfer-request-staff-detailed-view",
            headers=_ft_headers(tokens),
            params={"request_id": application_id},
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        logger.warning("HQ 2a detail failed for %s: %s", application_id, e)
        return None


def _fetch_hq_detail_2b(http_sess, tokens: AuthTokens, task_id: str) -> Optional[Dict]:
    """Fetch stamp-duty detail (2b) using task id — has officer assignments."""
    try:
        resp = http_sess.get(
            f"{BASE_URL}/stampdutyservice/api/v1/stamp-duty/detail-view",
            headers=_ft_headers(tokens),
            params={"request_id": task_id},
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        logger.warning("HQ 2b detail failed for %s: %s", task_id, e)
        return None


def _fetch_county_detail(http_sess, tokens: AuthTokens, task_id: str) -> Optional[Dict]:
    """Fetch county stamp-duty detail view."""
    try:
        resp = http_sess.get(
            f"{BASE_URL}/stampdutyservice/api/v1/stamp-duty/detail-view",
            headers=_ft_headers(tokens),
            params={"request_id": task_id},
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        logger.warning("County detail failed for %s: %s", task_id, e)
        return None


def _load_fetch_tasks(tokens: AuthTokens, days_back: int) -> Tuple[List[Dict], Dict]:
    """
    Fetch qualifying HQ + County TRANSFER applications in the last days_back days.
    Returns (tasks, stats).
    Each task: source, reference_number, date_created, county, registry,
               consideration, currency_code, parcel_number, officers.
    """
    http_sess = build_session()
    cutoff = _date_cutoff_str(days_back)

    # Both list fetches run in parallel
    with ThreadPoolExecutor(max_workers=2) as pool:
        hq_future     = pool.submit(_fetch_hq_list,    http_sess, tokens, cutoff)
        county_future = pool.submit(_fetch_county_list, http_sess, tokens, cutoff)
        hq_candidates     = hq_future.result()
        county_candidates = county_future.result()

    stats: Dict = {
        "hq_raw":     len(hq_candidates),
        "county_raw": len(county_candidates),
        "hq_kept":    0,
        "county_kept": 0,
    }
    tasks: List[Dict] = []

    # ── HQ: 2a + 2b in parallel per task ─────────────────────
    if hq_candidates:
        with ThreadPoolExecutor(max_workers=8) as pool:
            fut2a = {
                pool.submit(_fetch_hq_detail_2a, http_sess, tokens, t["application_id"]): t
                for t in hq_candidates
            }
            fut2b = {
                pool.submit(_fetch_hq_detail_2b, http_sess, tokens, t["id"]): t
                for t in hq_candidates
            }
            res2a: Dict[str, Optional[Dict]] = {}
            res2b: Dict[str, Optional[Dict]] = {}
            for f in _futures_as_completed(fut2a):
                res2a[fut2a[f]["id"]] = f.result()
            for f in _futures_as_completed(fut2b):
                res2b[fut2b[f]["id"]] = f.result()

        for t in hq_candidates:
            d2a = res2a.get(t["id"])
            if not d2a:
                continue
            if _has_stamp_duty_invoice(d2a.get("invoices", [])):
                continue
            if (d2a.get("stamp_duty_status") != "SENT_TO_COLLECTOR"
                    or d2a.get("application_status", "").upper() != "ONGOING"):
                continue

            d2b = res2b.get(t["id"])
            officers = []
            if d2b:
                officers = [
                    {"name": o.get("names", ""), "role": o.get("role", "")}
                    for o in d2b.get("details", {}).get("officers", [])
                ]

            tasks.append({
                "source":             "HQ",
                "reference_number":   t.get("reference_number", ""),
                "date_created":       t.get("date_created", ""),
                "county":             d2a.get("county") or t.get("county", ""),
                "registry":           d2a.get("registry") or t.get("registry", ""),
                "consideration":      str(d2a.get("consideration", "")),
                "consideration_type": d2a.get("consideration_type", ""),
                "currency_code":      d2a.get("currency_code", "KES"),
                "parcel_number":      t.get("parcel_number", ""),
                "officers":           officers,
            })
            stats["hq_kept"] += 1

    # ── County: detail per task ───────────────────────────────
    if county_candidates:
        with ThreadPoolExecutor(max_workers=8) as pool:
            county_fut = {
                pool.submit(_fetch_county_detail, http_sess, tokens, t["id"]): t
                for t in county_candidates
            }
            for f in _futures_as_completed(county_fut):
                t   = county_fut[f]
                det = (f.result() or {}).get("details", {})
                if not det:
                    continue
                if (det.get("node") != "STAMP_DUTY_PAYMENT_DEFINITION"
                        or det.get("application_status", "").upper() != "ONGOING"):
                    continue
                ext = det.get("external_process_details") or {}
                if _has_stamp_duty_invoice(ext.get("invoice", [])):
                    continue

                officers = [
                    {"name": o.get("names", ""), "role": o.get("role", "")}
                    for o in det.get("officers", [])
                ]
                tasks.append({
                    "source":             "County",
                    "reference_number":   det.get("reference_number") or t.get("reference_number", ""),
                    "date_created":       t.get("date_created", ""),
                    "county":             det.get("county") or t.get("county", ""),
                    "registry":           det.get("registry") or t.get("registry", ""),
                    "consideration":      str(ext.get("consideration_amount", "")),
                    "consideration_type": ext.get("process_type", ""),
                    "currency_code":      ext.get("currency_code", "KES"),
                    "parcel_number":      ext.get("parcel_number") or t.get("parcel_number", ""),
                    "officers":           officers,
                })
                stats["county_kept"] += 1

    tasks.sort(key=lambda x: x["date_created"], reverse=True)
    return tasks, stats


def _fetch_dlv_detail_one(http_sess, tokens: AuthTokens, task_id: str) -> Optional[Dict]:
    """Fetch DLV application detail view for one task."""
    headers = {
        "Authorization": f"Bearer {tokens.access_token}",
        "JWTAUTH":       f"Bearer {tokens.jwt}",
        "cparams":       CPARAMS_VALUER_ROLE,
    }
    try:
        resp = http_sess.get(
            f"{BASE_URL}/valuationservice/api/v1/stamp-duty/application/detail-view",
            headers=headers,
            params={"request_id": task_id},
            timeout=30,
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        logger.warning("DLV detail failed for %s: %s", task_id, e)
        return None


def _days_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("1 day",  callback_data="ft_days:1"),
            InlineKeyboardButton("2 days", callback_data="ft_days:2"),
            InlineKeyboardButton("3 days", callback_data="ft_days:3"),
        ],
        [
            InlineKeyboardButton("5 days",  callback_data="ft_days:5"),
            InlineKeyboardButton("7 days",  callback_data="ft_days:7"),
            InlineKeyboardButton("10 days", callback_data="ft_days:10"),
        ],
        [InlineKeyboardButton("✏️ Enter custom", callback_data="ft_days:custom")],
    ])


def _load_dlv_tasks(tokens: AuthTokens) -> Tuple[List[Dict], int]:
    """
    Fetch all pending DLV/VALUER tasks created in the last 2 days,
    enrich each with detail fields, return (enriched_tasks, raw_count).
    raw_count = total results seen before date filtering.
    """
    http_sess = build_session()
    cutoff = _date_cutoff_str(2)
    headers = {
        "Authorization": f"Bearer {tokens.access_token}",
        "JWTAUTH":       f"Bearer {tokens.jwt}",
        "cparams":       CPARAMS_VALUER_ROLE,
    }

    candidates: List[Dict] = []
    raw_count = 0
    page = 1
    while True:
        resp = http_sess.get(
            f"{BASE_URL}/valuationservice/api/v1/stamp-duty/application",
            headers=headers,
            params={
                "filter": "Pending", "role": "VALUER",
                "request_type": "STAMP_DUTY", "search": "", "page": page,
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        results = data.get("results", [])
        if not results:
            break

        raw_count += len(results)
        for task in results:
            if _within_days(task.get("date_created", ""), cutoff):
                candidates.append(task)

        if not data.get("next"):
            break
        page += 1

    enriched: List[Dict] = []
    with ThreadPoolExecutor(max_workers=8) as pool:
        future_map = {
            pool.submit(_fetch_dlv_detail_one, http_sess, tokens, t["id"]): t
            for t in candidates
        }
        for future in _futures_as_completed(future_map):
            list_task = future_map[future]
            detail    = future.result()
            consideration = ""
            if detail:
                ext = detail.get("external_process_details") or {}
                consideration = str(ext.get("consideration_amount", ""))
            enriched.append({
                "id":               list_task["id"],
                "reference_number": list_task.get("reference_number", ""),
                "date_created":     list_task.get("date_created", ""),
                "consideration":    consideration,
                "county":           list_task.get("county", ""),
                "registry":         list_task.get("registry", ""),
            })

    enriched.sort(key=lambda x: x["date_created"], reverse=True)
    return enriched, raw_count


def _tasks_keyboard(tasks: List[Dict], checked: set, prefix: str) -> InlineKeyboardMarkup:
    """Build a checklist inline keyboard for a task list."""
    rows = []
    for i, t in enumerate(tasks):
        mark  = "☑️" if i in checked else "⬜"
        ref   = t.get("reference_number", f"#{i+1}")
        cons  = t.get("consideration", "")
        try:
            cons_str = f"KES {int(float(cons)):,}" if cons else "—"
        except (ValueError, TypeError):
            cons_str = cons or "—"
        county = t.get("county", "")
        date   = (t.get("date_created") or "")[:10]
        label  = f"{mark} {ref} | {cons_str} | {county} | {date}"
        rows.append([InlineKeyboardButton(label, callback_data=f"{prefix}:{i}")])

    action_row = [InlineKeyboardButton("🔄 Refresh", callback_data=f"{prefix}:refresh")]
    if checked:
        action_row.append(
            InlineKeyboardButton(f"📤 Assign Selected ({len(checked)})", callback_data=f"{prefix}:assign")
        )
    rows.append(action_row)
    return InlineKeyboardMarkup(rows)


async def cmd_fetch_tasks(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    ctx.user_data["ft_session"] = FTSession()
    await update.message.reply_text(
        "📊 *Fetch Tasks* — Select the account to use:",
        parse_mode="Markdown",
        reply_markup=_cred_keyboard(),
    )
    return FT.CHOOSE_CRED


async def recv_ft_cred(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    cred_type = query.data.split(":")[1]
    sess = _get_ft_sess(ctx)
    sess.cred_type = cred_type

    cached = get_valid_tokens(cred_type)
    if cached:
        sess.tokens = cached
        await query.edit_message_text(
            f"✅ Using cached tokens for *{CRED_LABELS[cred_type]}*.\n\n"
            "How many *days back* should I look?\nTap a button or type a number:",
            parse_mode="Markdown",
            reply_markup=_days_keyboard(),
        )
        return FT.DAYS_BACK

    sess.http_session = build_session()
    creds = CRED_MAP[cred_type]
    await query.edit_message_text(
        f"🔐 Sending login request for *{CRED_LABELS[cred_type]}*…",
        parse_mode="Markdown",
    )
    try:
        resp = sess.http_session.post(
            f"{AUTH_BASE_URL}/login",
            json={"username": creds["username"], "password": creds["password"],
                  "usertype": creds["usertype"], "otpcode": ""},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        if data.get("success") is False and "error" in data:
            raise RuntimeError(data.get("error") or data.get("message"))
    except Exception as e:
        await query.message.reply_text(
            f"❌ Login failed: `{e}`", parse_mode="Markdown", reply_markup=_main_menu()
        )
        return ConversationHandler.END

    await query.message.reply_text(
        "📲 OTP sent to the registered device.\nPlease *reply with the OTP code*:",
        parse_mode="Markdown",
    )
    return FT.WAIT_OTP


async def recv_ft_otp(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    otp   = update.message.text.strip()
    sess  = _get_ft_sess(ctx)
    creds = CRED_MAP[sess.cred_type]

    await update.message.reply_text("🔄 Verifying OTP…")
    try:
        resp = sess.http_session.post(
            f"{AUTH_BASE_URL}/otpverify",
            json={"username": creds["username"], "password": creds["password"], "otpcode": otp},
            timeout=30,
        )
        resp.raise_for_status()
        data          = resp.json()
        details       = data.get("details", {})
        access_token  = details.get("access_token")
        jwt           = details.get("jwt")
        refresh_token = details.get("refresh_token", "")
        if not access_token or not jwt:
            raise RuntimeError(f"Tokens missing. Keys: {list(data.keys())}")
        persist_tokens(sess.cred_type, access_token, jwt, refresh_token)
        sess.tokens = AuthTokens(access_token=access_token, jwt=jwt)
    except Exception as e:
        await update.message.reply_text(
            f"❌ OTP verification failed: `{e}`\n\nSend the OTP again or tap 🛑 Cancel.",
            parse_mode="Markdown",
        )
        return FT.WAIT_OTP

    await update.message.reply_text(
        f"✅ Authenticated as *{CRED_LABELS[sess.cred_type]}*.\n\n"
        "How many *days back* should I look?\nTap a button or type a number:",
        parse_mode="Markdown",
        reply_markup=_days_keyboard(),
    )
    return FT.DAYS_BACK


async def recv_ft_days_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    val  = query.data.split(":")[1]
    sess = _get_ft_sess(ctx)

    if val == "custom":
        await query.edit_message_text(
            "Enter the number of days to look back (1–90):",
        )
        return FT.DAYS_BACK

    sess.days_back = int(val)
    await query.edit_message_text(
        f"✅ *{sess.days_back}* day(s) selected.\n\nFilter by *county*?",
        parse_mode="Markdown",
        reply_markup=_ft_county_keyboard(),
    )
    return FT.COUNTY_FILTER


async def recv_ft_days_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    sess = _get_ft_sess(ctx)
    try:
        days = int(text)
        if days < 1 or days > 90:
            raise ValueError
    except ValueError:
        await update.message.reply_text(
            "Please enter a number between 1 and 90:",
            reply_markup=_days_keyboard(),
        )
        return FT.DAYS_BACK

    sess.days_back = days
    await update.message.reply_text(
        f"✅ *{sess.days_back}* day(s) selected.\n\nFilter by *county*?",
        parse_mode="Markdown",
        reply_markup=_ft_county_keyboard(),
    )
    return FT.COUNTY_FILTER


def _ft_county_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🌆 Nairobi",      callback_data="ft_county:nairobi"),
            InlineKeyboardButton("📋 All Counties", callback_data="ft_county:all"),
        ],
    ])


def _ft_registry_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("📁 Central",        callback_data="ft_registry:central"),
            InlineKeyboardButton("📁 Nairobi",         callback_data="ft_registry:nairobi"),
        ],
        [
            InlineKeyboardButton("📋 All Registries", callback_data="ft_registry:all"),
        ],
    ])


def _ft_amount_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("0 – 1M",        callback_data="ft_amount:0_1m"),
            InlineKeyboardButton("1M – 5M",        callback_data="ft_amount:1m_5m"),
        ],
        [
            InlineKeyboardButton("5M – 10M",       callback_data="ft_amount:5m_10m"),
            InlineKeyboardButton("20M – 50M",      callback_data="ft_amount:20m_50m"),
        ],
        [
            InlineKeyboardButton("50M – 100M",     callback_data="ft_amount:50m_100m"),
            InlineKeyboardButton("80M – 300M",     callback_data="ft_amount:80m_300m"),
        ],
        [
            InlineKeyboardButton("80M – 3B",       callback_data="ft_amount:80m_3b"),
            InlineKeyboardButton("✏️ Custom",       callback_data="ft_amount:custom"),
        ],
        [
            InlineKeyboardButton("📋 No filter",   callback_data="ft_amount:all"),
        ],
    ])


def _sectional_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("🚫 Exclude Sectional", callback_data="ft_sectional:exclude"),
        InlineKeyboardButton("🏢 Sectional Only",    callback_data="ft_sectional:only"),
        InlineKeyboardButton("📋 All",               callback_data="ft_sectional:all"),
    ]])


async def recv_ft_county_filter(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess = _get_ft_sess(ctx)

    sess.county_filter = "" if query.data == "ft_county:all" else query.data.split(":")[1]
    label = f"*{sess.county_filter.title()}*" if sess.county_filter else "*All Counties*"
    await query.edit_message_text(
        f"County: {label}\n\nFilter by *registry*?",
        parse_mode="Markdown",
        reply_markup=_ft_registry_keyboard(),
    )
    return FT.REGISTRY_FILTER


async def recv_ft_registry_filter(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess = _get_ft_sess(ctx)

    sess.registry_filter = "" if query.data == "ft_registry:all" else query.data.split(":")[1]
    reg_label    = f"*{sess.registry_filter.title()}*" if sess.registry_filter else "*All Registries*"
    county_label = f"*{sess.county_filter.title()}*"   if sess.county_filter   else "*All Counties*"
    await query.edit_message_text(
        f"County: {county_label} | Registry: {reg_label}\n\nFilter by *amount*?",
        parse_mode="Markdown",
        reply_markup=_ft_amount_keyboard(),
    )
    return FT.AMOUNT_FILTER


async def recv_ft_amount_filter(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess = _get_ft_sess(ctx)

    choice = query.data

    if choice == "ft_amount:custom":
        await query.edit_message_text(
            "Enter the amount range as *min max* (space-separated), in KES:\n"
            "e.g. `500000 2000000`",
            parse_mode="Markdown",
        )
        return FT.AMOUNT_TEXT

    ranges = {
        "ft_amount:0_1m":    (0.0,           1_000_000.0),
        "ft_amount:1m_5m":   (1_000_000.0,   5_000_000.0),
        "ft_amount:5m_10m":  (5_000_000.0,  10_000_000.0),
        "ft_amount:20m_50m": (20_000_000.0, 50_000_000.0),
        "ft_amount:50m_100m":(50_000_000.0,100_000_000.0),
        "ft_amount:80m_300m":(80_000_000.0,300_000_000.0),
        "ft_amount:80m_3b":  (80_000_000.0,  3_000_000_000.0),
        "ft_amount:all":     (None,           None),
    }
    sess.amount_min, sess.amount_max = ranges.get(choice, (None, None))
    await query.edit_message_text(
        "Include sectional properties?\n_(Sectional: parcel has 4 parts e.g. Nairobi/Block12/345/888)_",
        parse_mode="Markdown",
        reply_markup=_sectional_keyboard(),
    )
    return FT.SECTIONAL_FILTER


async def recv_ft_amount_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text  = update.message.text.strip()
    sess  = _get_ft_sess(ctx)
    parts = text.replace(",", "").split()
    try:
        if len(parts) == 2:
            sess.amount_min, sess.amount_max = float(parts[0]), float(parts[1])
        elif len(parts) == 1:
            sess.amount_min, sess.amount_max = 0.0, float(parts[0])
        else:
            raise ValueError
    except (ValueError, IndexError):
        await update.message.reply_text(
            "❌ Could not parse that. Enter two numbers separated by a space, e.g. `500000 2000000`",
            parse_mode="Markdown",
        )
        return FT.AMOUNT_TEXT

    await update.message.reply_text(
        "Include sectional properties?\n_(Sectional: parcel has 4 parts e.g. Nairobi/Block12/345/888)_",
        parse_mode="Markdown",
        reply_markup=_sectional_keyboard(),
    )
    return FT.SECTIONAL_FILTER


async def recv_ft_sectional_filter(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    sess = _get_ft_sess(ctx)
    sess.sectional_filter = query.data.split(":")[1]  # "exclude" | "only" | "all"
    await query.edit_message_text(
        f"⏳ Fetching tasks from the last *{sess.days_back}* day(s)…",
        parse_mode="Markdown",
    )
    return await _ft_do_fetch(query.message, ctx, sess)


async def _ft_do_fetch(message, ctx: ContextTypes.DEFAULT_TYPE, sess: FTSession):
    """Fetch tasks, apply county / registry / amount filters, then show results."""
    try:
        tasks, stats = _load_fetch_tasks(sess.tokens, sess.days_back)
    except Exception as e:
        await message.reply_text(
            f"❌ Fetch failed: `{e}`", parse_mode="Markdown", reply_markup=_main_menu()
        )
        return ConversationHandler.END

    sess.stats = stats
    hq_raw  = stats.get("hq_raw",     0)
    c_raw   = stats.get("county_raw", 0)
    hq_kept = stats.get("hq_kept",    0)
    c_kept  = stats.get("county_kept", 0)

    # County filter (partial, case-insensitive)
    if sess.county_filter:
        tasks = [
            t for t in tasks
            if sess.county_filter in (t.get("county") or "").strip().lower()
        ]

    # Registry filter (partial, case-insensitive)
    if sess.registry_filter:
        tasks = [
            t for t in tasks
            if sess.registry_filter in (t.get("registry") or "").strip().lower()
        ]

    # Amount filter — strip commas/spaces so "1,500,000.00" parses correctly
    if sess.amount_min is not None or sess.amount_max is not None:
        def _in_range(t):
            raw = t.get("consideration")
            if raw is None:
                return False
            try:
                val = float(str(raw).replace(",", "").strip())
            except (ValueError, TypeError):
                return False
            if sess.amount_min is not None and val < sess.amount_min:
                return False
            if sess.amount_max is not None and val > sess.amount_max:
                return False
            return True
        tasks = [t for t in tasks if _in_range(t)]

    sess.tasks = tasks

    # Build active-filter summary for the header
    filter_parts = []
    if sess.county_filter:
        filter_parts.append(f"County: {sess.county_filter.title()}")
    if sess.registry_filter:
        filter_parts.append(f"Registry: {sess.registry_filter.title()}")
    if sess.amount_min is not None or sess.amount_max is not None:
        lo = f"KES {int(sess.amount_min):,}" if sess.amount_min is not None else "0"
        hi = f"KES {int(sess.amount_max):,}" if sess.amount_max is not None else "∞"
        filter_parts.append(f"Amount: {lo} – {hi}")
    filter_line = ("_Filters: " + " | ".join(filter_parts) + "_\n") if filter_parts else ""

    if not tasks:
        await message.reply_text(
            f"ℹ️ No qualifying tasks in the last {sess.days_back} day(s).\n"
            f"{filter_line}"
            f"(HQ: {hq_raw} seen → {hq_kept} matched | "
            f"County: {c_raw} seen → {c_kept} matched)",
            reply_markup=_main_menu(),
        )
        return ConversationHandler.END

    # Sectional filter (parcel with 4+ parts = sectional)
    sf = sess.sectional_filter
    if sf == "exclude":
        tasks = [t for t in tasks if str(t.get("parcel_number") or "").count("/") < 3]
    elif sf == "only":
        tasks = [t for t in tasks if str(t.get("parcel_number") or "").count("/") >= 3]
    # "all" → no filter

    # Remove tasks already queued in the DLV batch
    queued_refs = {item.get("ref", "") for item in load_dlv_batch()}
    tasks_before = len(tasks)
    tasks = [t for t in tasks if t.get("reference_number", "") not in queued_refs]
    queued_removed = tasks_before - len(tasks)

    queued_note = f" ({queued_removed} already in DLV queue — excluded)" if queued_removed else ""

    if not tasks:
        await message.reply_text(
            f"ℹ️ All {tasks_before} task(s) are already in the DLV queue.",
            reply_markup=_main_menu(),
        )
        return ConversationHandler.END

    await message.reply_text(
        f"✅ {len(tasks)} task(s) found — last {sess.days_back} day(s).{queued_note}\n"
        f"{filter_line}"
        f"(HQ: {hq_raw} → {hq_kept} | County: {c_raw} → {c_kept})",
    )
    try:
        await _ft_show_results(message, tasks)
    except Exception as e:
        logger.error("_ft_show_results error: %s", e)
        await message.reply_text(
            f"❌ Failed to display results: {e}", reply_markup=_main_menu()
        )
    return ConversationHandler.END


async def _ft_show_results(message, tasks: List[Dict]):
    """Send tasks as formatted text, splitting at Telegram's 4096-char limit."""
    if not tasks:
        await message.reply_text("No tasks to display.", reply_markup=_main_menu())
        return

    lines = []
    for i, t in enumerate(tasks, 1):
        src    = t.get("source", "")
        ref    = t.get("reference_number", "—")
        cnty   = (t.get("county") or "—").upper()
        reg    = (t.get("registry") or "—").upper()
        date   = (t.get("date_created") or "")[:10]
        parcel = t.get("parcel_number") or "—"
        try:
            raw_cons = t.get("consideration")
            cons = f"KES {int(float(str(raw_cons).replace(',', '').strip())):,}" if raw_cons else "—"
        except (ValueError, TypeError):
            cons = str(t.get("consideration") or "—")
        officers_str = ", ".join(
            f"{o['name']} ({o['role']})" for o in t.get("officers", []) if o.get("name")
        ) or "none"

        lines.append(
            f"{i}. [{src}] {ref}\n"
            f"   {cnty} / {reg} | {date}\n"
            f"   {cons} | {parcel}\n"
            f"   {officers_str}"
        )

    chunks = []
    chunk = ""
    for line in lines:
        candidate = (chunk + "\n\n" + line).strip()
        if len(candidate) > 4000:
            chunks.append(chunk)
            chunk = line
        else:
            chunk = candidate
    if chunk:
        chunks.append(chunk)

    for i, c in enumerate(chunks):
        is_last = (i == len(chunks) - 1)
        text = c + (f"\n\nTotal: {len(tasks)} task(s)" if is_last else "")
        try:
            await message.reply_text(
                text,
                reply_markup=_main_menu() if is_last else None,
            )
        except Exception as e:
            logger.warning("_ft_show_results send failed: %s", e)
            await message.reply_text(
                text[:4000],
                reply_markup=_main_menu() if is_last else None,
            )


async def cmd_dlv_tasks(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    tokens = _any_valid_tokens()
    if not tokens:
        await update.message.reply_text(
            "⚠️ No valid cached tokens found.\n"
            "Please authenticate first via *New Assignment* or *Receive Tasks*, then try again.",
            parse_mode="Markdown",
            reply_markup=_main_menu(),
        )
        return

    await update.message.reply_text("⏳ Fetching DLV Tasks (last 2 days)…")
    try:
        tasks, raw_count = _load_dlv_tasks(tokens)
    except Exception as e:
        await update.message.reply_text(
            f"❌ Failed to fetch tasks: `{e}`", parse_mode="Markdown", reply_markup=_main_menu()
        )
        return

    if not tasks:
        await update.message.reply_text(
            f"ℹ️ No pending DLV tasks in the last 2 days.\n"
            f"_(API returned {raw_count} total result(s) before date filter)_",
            parse_mode="Markdown",
            reply_markup=_main_menu(),
        )
        return

    ctx.user_data["dlv_tasks"]   = tasks
    ctx.user_data["dlv_checked"] = set()
    await update.message.reply_text(
        f"📋 *DLV Tasks* — {len(tasks)} task(s) in last 2 days\n"
        f"_(fetched {raw_count} total from API)_\n"
        "Tap any row to check/uncheck:",
        parse_mode="Markdown",
        reply_markup=_tasks_keyboard(tasks, set(), "dlv_ck"),
    )


async def recv_dlv_check(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    idx_str = query.data.split(":")[1]
    tasks   = ctx.user_data.get("dlv_tasks", [])
    checked = ctx.user_data.get("dlv_checked", set())

    if idx_str == "refresh":
        tokens = _any_valid_tokens()
        if not tokens:
            await query.answer("No valid tokens — authenticate first.", show_alert=True)
            return
        await query.edit_message_text("⏳ Refreshing DLV Tasks…")
        try:
            tasks, raw_count = _load_dlv_tasks(tokens)
        except Exception as e:
            await query.edit_message_text(f"❌ Refresh failed: {e}")
            return
        checked = set()
        ctx.user_data["dlv_tasks"]   = tasks
        ctx.user_data["dlv_checked"] = checked
        await query.edit_message_text(
            f"📋 *DLV Tasks* — {len(tasks)} task(s) in last 2 days\n"
            f"_(fetched {raw_count} total from API)_\n"
            "Tap any row to check/uncheck:",
            parse_mode="Markdown",
            reply_markup=_tasks_keyboard(tasks, checked, "dlv_ck"),
        )
        return

    if idx_str == "assign":
        if not checked:
            await query.answer("No tasks selected.", show_alert=True)
            return
        selected = [tasks[i] for i in sorted(checked) if i < len(tasks)]
        ctx.user_data["ta_pending"] = {"type": "dlv", "tasks": selected, "search_results": []}
        await query.message.reply_text(
            f"📤 Assigning *{len(selected)}* DLV task(s).\n\n"
            "Select the valuer to assign to:",
            parse_mode="Markdown",
            reply_markup=_ta_valuer_picker_keyboard("dlv"),
        )
        return

    idx = int(idx_str)
    if idx in checked:
        checked.discard(idx)
    else:
        checked.add(idx)
    ctx.user_data["dlv_checked"] = checked
    await query.edit_message_reply_markup(
        reply_markup=_tasks_keyboard(tasks, checked, "dlv_ck"),
    )


# ──────────────────────────────────────────────────────────
# DLV Batch — conversation handlers
# ──────────────────────────────────────────────────────────

async def cmd_dlv_batch(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    ctx.user_data["db_session"] = DBSession()
    await update.message.reply_text(
        "📥 *DLV Batch Assignment*\n\n"
        "Send your batch — one group per line:\n"
        "`REF1, REF2, REF3 : Valuer Name`\n\n"
        "*Example:*\n"
        "`REG/TSFR/5A0B3E1VLS, REG/TSFR/5A0B3E1VLQ : Byron`\n"
        "`REG/TSFR/XXXXXXXX : John Kamau`\n\n"
        "Multiple lines are processed together.",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove(),
    )
    return DB.INPUT_BATCH


async def recv_db_input(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text   = update.message.text.strip()
    groups = _parse_batch_input(text)

    if not groups:
        await update.message.reply_text(
            "⚠️ Could not parse any entries. Use format:\n`REF1, REF2 : Valuer Name`",
            parse_mode="Markdown",
        )
        return DB.INPUT_BATCH

    await update.message.reply_text("🔍 Resolving valuers…")
    tokens  = _any_valid_tokens()
    sess    = _get_db_sess(ctx)
    resolved = []

    for group in groups:
        name_raw = group["valuer_name_raw"]
        refs     = group["refs"]

        # 1. Check saved valuers
        saved = _resolve_valuer_from_saved(name_raw)
        if saved:
            resolved.append({
                "refs":        refs,
                "valuer_name": saved["name"],
                "valuer_uid":  saved["uid"],
                "valuer_acct": saved["account_number"],
                "status":      "resolved",
            })
            continue

        # 2. Search API
        if tokens:
            try:
                results = _search_valuer_api(name_raw, tokens)
                if results:
                    v    = results[0]
                    sd   = v.get("staff_details", {})
                    name = " ".join(filter(None, [sd.get("firstname"), sd.get("middlename"), sd.get("lastname")]))
                    resolved.append({
                        "refs":        refs,
                        "valuer_name": name,
                        "valuer_uid":  str(v.get("id", "")),
                        "valuer_acct": str(v.get("account_number", "")),
                        "status":      "resolved",
                    })
                    continue
            except Exception as e:
                logger.warning("Valuer search error for %s: %s", name_raw, e)

        # 3. Unresolvable
        resolved.append({
            "refs":        refs,
            "valuer_name": name_raw,
            "valuer_uid":  "",
            "valuer_acct": "",
            "status":      "unresolved",
        })

    sess.groups = resolved

    # Build confirmation summary
    lines = ["📋 *Batch Summary — please confirm:*\n"]
    has_unresolved = False
    for g in resolved:
        refs_str = ", ".join(f"`{r}`" for r in g["refs"])
        if g["status"] == "resolved":
            lines.append(f"✅ {refs_str}\n   → *{g['valuer_name']}*")
        else:
            lines.append(f"⚠️ {refs_str}\n   → _{g['valuer_name']}_ (NOT FOUND — will be skipped)")
            has_unresolved = True

    if has_unresolved:
        lines.append("\n_Unresolved valuers will be skipped._")

    msg = "\n".join(lines)
    if len(msg) > 4000:
        msg = msg[:4000] + "\n…_(truncated)_"

    await update.message.reply_text(
        msg,
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Confirm & Run", callback_data="db:confirm"),
            InlineKeyboardButton("❌ Cancel",        callback_data="db:cancel"),
        ]]),
    )
    return DB.CONFIRM_BATCH


async def recv_db_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "db:cancel":
        await query.edit_message_text("❌ DLV Batch cancelled.")
        await query.message.reply_text("Main menu:", reply_markup=_main_menu())
        return ConversationHandler.END

    sess    = _get_db_sess(ctx)
    to_save = [g for g in sess.groups if g["status"] == "resolved"]

    if not to_save:
        await query.edit_message_text("⚠️ No resolved valuers — nothing to process.")
        await query.message.reply_text("Main menu:", reply_markup=_main_menu())
        return ConversationHandler.END

    # Flatten groups to individual ref+valuer items for per-ref retry tracking
    existing     = load_dlv_batch()
    existing_refs = {item["ref"] for item in existing}
    new_items = []
    for g in to_save:
        for ref in g["refs"]:
            if ref not in existing_refs:
                new_items.append({
                    "ref":         ref,
                    "valuer_name": g["valuer_name"],
                    "valuer_uid":  g["valuer_uid"],
                    "valuer_acct": g["valuer_acct"],
                })
    flat_items = existing + new_items
    save_dlv_batch(flat_items)
    await query.edit_message_text(
        f"✅ *{len(new_items)} new ref(s)* added to queue ({len(flat_items)} total). Processing now…",
        parse_mode="Markdown",
    )

    tokens = _any_valid_tokens()
    if not tokens:
        await query.message.reply_text(
            "⚠️ No valid tokens — authenticate first.\n"
            "Batch saved; will retry on the next 5-minute cycle.",
            reply_markup=_main_menu(),
        )
        return ConversationHandler.END

    report = _process_dlv_batch_items(tokens)
    msg    = f"📋 *DLV Batch Report*\n{report}" if report else "ℹ️ Batch was already empty."
    if len(msg) > 4000:
        msg = msg[:4000] + "\n…_(truncated)_"
    await query.message.reply_text(msg, parse_mode="Markdown", reply_markup=_main_menu())
    return ConversationHandler.END


# ──────────────────────────────────────────────────────────
# Task-assign sub-flow (valuer picker + assignment execution)
# ──────────────────────────────────────────────────────────

def _ta_valuer_picker_keyboard(task_type: str) -> InlineKeyboardMarkup:
    """Inline keyboard: saved valuers + search-new option."""
    rows = []
    for i, v in enumerate(load_saved_valuers()):
        rows.append([InlineKeyboardButton(
            f"👤 {v['name']}", callback_data=f"ta:{task_type}:sv:{i}"
        )])
    rows.append([InlineKeyboardButton("🔍 Search New Valuer", callback_data=f"ta:{task_type}:ns")])
    rows.append([InlineKeyboardButton("❌ Cancel",            callback_data=f"ta:{task_type}:cancel")])
    return InlineKeyboardMarkup(rows)


async def _ta_run_assign(message, ctx: ContextTypes.DEFAULT_TYPE,
                         task_type: str, valuer_uid: str, valuer_name: str, acct: str):
    """Execute assignment for all tasks in ta_pending and report results."""
    pending = ctx.user_data.get("ta_pending", {})
    tasks   = pending.get("tasks", [])
    tokens  = _any_valid_tokens()
    if not tokens:
        await message.reply_text(
            "⚠️ Tokens expired. Please re-authenticate via New Assignment or Receive Tasks.",
            reply_markup=_main_menu(),
        )
        return

    http_sess = build_session()
    url     = f"{BASE_URL}/valuationservice/api/v1/stamp-duty/fix_application_details"
    headers = {
        "Authorization": f"Bearer {tokens.access_token}",
        "JWTAUTH":       f"Bearer {tokens.jwt}",
    }

    await message.reply_text(f"⏳ Assigning {len(tasks)} task(s) to *{valuer_name}*…", parse_mode="Markdown")

    ok_refs, fail_refs, result_lines = [], [], []
    for task in tasks:
        ref = task.get("reference_number", "")
        try:
            r = http_sess.post(
                url, headers=headers,
                json={
                    "reference_number":  ref,
                    "valuation_officer": valuer_uid,
                    "node":              "VALUATION_STAMP_DUTY_VALUER_REPORT",
                },
                timeout=30,
            )
            r.raise_for_status()
            ok_refs.append(ref)
            result_lines.append(f"✅ `{ref}`")
            persist_assignment(ref, valuer_name, valuer_uid)
        except Exception as e:
            fail_refs.append(ref)
            result_lines.append(f"❌ `{ref}` — {e}")

    if ok_refs:
        persist_valuer(valuer_name, valuer_uid, acct)

    summary = (
        f"🏁 *Assignment Complete*\n\n"
        f"*Valuer:* {valuer_name}\n"
        f"*Success:* {len(ok_refs)} / {len(tasks)}\n"
        f"*Failed:*  {len(fail_refs)} / {len(tasks)}\n\n"
        + "\n".join(result_lines)
    )
    if len(summary) > 4000:
        summary = summary[:4000] + "\n…_(truncated)_"
    await message.reply_text(summary, parse_mode="Markdown", reply_markup=_main_menu())

    # Clear pending state
    ctx.user_data.pop("ta_pending", None)


async def recv_ta_valuer(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Handle all ta:* callback data for the task-assign sub-flow."""
    query = update.callback_query
    await query.answer()

    # callback format: ta:{task_type}:{action}[:{idx}]
    parts     = query.data.split(":")
    task_type = parts[1]   # "impl" or "dlv"
    action    = parts[2]   # "sv", "ns", "cancel", "sr"

    pending = ctx.user_data.get("ta_pending", {})

    if action == "cancel":
        ctx.user_data.pop("ta_pending", None)
        await query.edit_message_text("❌ Assignment cancelled.")
        return

    if action == "sv":
        idx    = int(parts[3])
        saved  = load_saved_valuers()
        if idx >= len(saved):
            await query.edit_message_text("⚠️ Valuer not found in saved list.")
            return
        v = saved[idx]
        await query.edit_message_text(
            f"✅ Valuer: *{v['name']}*\nStarting assignment…",
            parse_mode="Markdown",
        )
        await _ta_run_assign(query.message, ctx, task_type, v["uid"], v["name"], v["account_number"])
        return

    if action == "ns":
        pending["awaiting_search"] = True
        ctx.user_data["ta_pending"] = pending
        await query.edit_message_text(
            "🔍 Enter the *valuer name* to search:\n_(Partial names work, e.g._ `JOHN KAMAU`_)_",
            parse_mode="Markdown",
        )
        return

    if action == "sr":
        idx     = int(parts[3])
        results = pending.get("search_results", [])
        if idx >= len(results):
            await query.edit_message_text("⚠️ Search result no longer available.")
            return
        v    = results[idx]
        sd   = v.get("staff_details", {})
        name = " ".join(filter(None, [sd.get("firstname"), sd.get("middlename"), sd.get("lastname")]))
        uid  = str(v.get("id", ""))
        acct = str(v.get("account_number", ""))
        await query.edit_message_text(
            f"✅ Valuer: *{name}*\nStarting assignment…",
            parse_mode="Markdown",
        )
        await _ta_run_assign(query.message, ctx, task_type, uid, name, acct)
        return


async def handle_ta_search(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Global text handler: catches valuer search input during task-assign flow."""
    pending = ctx.user_data.get("ta_pending")
    if not pending or not pending.get("awaiting_search"):
        return   # nothing to do — not in task-assign search mode

    valuer_name = update.message.text.strip()
    task_type   = pending.get("type", "impl")
    pending["awaiting_search"] = False

    tokens = _any_valid_tokens()
    if not tokens:
        await update.message.reply_text(
            "⚠️ Tokens expired. Please re-authenticate first.",
            reply_markup=_main_menu(),
        )
        ctx.user_data.pop("ta_pending", None)
        return

    await update.message.reply_text(f"🔍 Searching for *{valuer_name}*…", parse_mode="Markdown")
    try:
        http_sess = build_session()
        resp = http_sess.get(
            f"{BASE_URL}/acl/api/v1/accounts/list-user-accounts",
            headers={
                "Authorization": f"Bearer {tokens.access_token}",
                "JWTAUTH":       f"Bearer {tokens.jwt}",
            },
            params={"account_type": "STAFF", "filter_type": "ACTIVE", "page": 1, "search": valuer_name},
            timeout=30,
        )
        resp.raise_for_status()
        results = resp.json().get("results", [])
    except Exception as e:
        await update.message.reply_text(f"❌ Search failed: `{e}`", parse_mode="Markdown")
        return

    if not results:
        await update.message.reply_text(
            f"⚠️ No staff found matching *{valuer_name}*. Try again:",
            parse_mode="Markdown",
            reply_markup=_ta_valuer_picker_keyboard(task_type),
        )
        pending["awaiting_search"] = False
        return

    pending["search_results"] = results
    ctx.user_data["ta_pending"] = pending

    rows = []
    for i, v in enumerate(results):
        sd   = v.get("staff_details", {})
        name = " ".join(filter(None, [sd.get("firstname"), sd.get("middlename"), sd.get("lastname")]))
        rows.append([InlineKeyboardButton(
            name or f"Valuer {i+1}", callback_data=f"ta:{task_type}:sr:{i}"
        )])
    rows.append([InlineKeyboardButton("❌ Cancel", callback_data=f"ta:{task_type}:cancel")])
    await update.message.reply_text(
        f"Found *{len(results)}* match(es). Select one:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(rows),
    )


# ──────────────────────────────────────────────────────────
# Refresh Auth — conversation handlers
# ──────────────────────────────────────────────────────────

def _auth_cred_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(CRED_LABELS["publicuser"],   callback_data="auth_cred:publicuser")],
        [InlineKeyboardButton(CRED_LABELS["staff"],        callback_data="auth_cred:staff")],
        [InlineKeyboardButton(CRED_LABELS["staff2"],       callback_data="auth_cred:staff2")],
        [InlineKeyboardButton(CRED_LABELS["staff_valuer"], callback_data="auth_cred:staff_valuer")],
        [InlineKeyboardButton("❌ Cancel",                 callback_data="auth_cred:cancel")],
    ])


async def cmd_auth(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    ctx.user_data["auth_session"] = AuthSession()
    await update.message.reply_text(
        "🔑 *Refresh Auth*\n\nSelect a credential profile to authenticate:",
        parse_mode="Markdown",
        reply_markup=_auth_cred_keyboard(),
    )
    return AS.CHOOSE_CRED


async def recv_auth_cred(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    cred_type = query.data.split(":")[1]

    if cred_type == "cancel":
        await query.edit_message_text("❌ Cancelled.")
        await query.message.reply_text("Use the menu to continue.", reply_markup=_main_menu())
        return ConversationHandler.END

    auth_sess           = _get_auth_sess(ctx)
    auth_sess.cred_type = cred_type

    cached = get_valid_tokens(cred_type)
    if cached:
        entry   = _load_tokens_raw().get(cred_type, {})
        exp_ts  = entry.get("expires_at", 0)
        exp_str = (
            datetime.fromtimestamp(exp_ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
            if exp_ts else "unknown"
        )
        await query.edit_message_text(
            f"✅ *{CRED_LABELS[cred_type]}* already has valid cached tokens.\n"
            f"*Expires:* {exp_str}\n\n"
            "Force a fresh login anyway?",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Yes, re-authenticate", callback_data="auth_force:yes")],
                [InlineKeyboardButton("✅ No, keep current",     callback_data="auth_force:no")],
            ]),
        )
        return AS.FORCE_CONFIRM

    # No valid tokens — go straight to login
    return await _auth_trigger_login(query, auth_sess)


async def recv_auth_force(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    auth_sess = _get_auth_sess(ctx)

    if query.data.split(":")[1] == "no":
        await query.edit_message_text(
            f"✅ Keeping existing tokens for *{CRED_LABELS[auth_sess.cred_type]}*.",
            parse_mode="Markdown",
        )
        await query.message.reply_text("Use the menu to continue.", reply_markup=_main_menu())
        return ConversationHandler.END

    return await _auth_trigger_login(query, auth_sess)


async def _auth_trigger_login(query, auth_sess: AuthSession) -> int:
    """Send the login request and transition to WAIT_OTP, or END on failure."""
    creds = CRED_MAP[auth_sess.cred_type]
    auth_sess.http_session = build_session()

    await query.edit_message_text(
        f"🔐 Sending login request for *{CRED_LABELS[auth_sess.cred_type]}*…",
        parse_mode="Markdown",
    )
    try:
        resp = auth_sess.http_session.post(
            f"{AUTH_BASE_URL}/login",
            json={"username": creds["username"], "password": creds["password"],
                  "usertype": creds["usertype"], "otpcode": ""},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        if data.get("success") is False and "error" in data:
            raise RuntimeError(data.get("error") or data.get("message"))
    except Exception as e:
        await query.message.reply_text(
            f"❌ Login failed: `{e}`\n\nUse the menu to retry.",
            parse_mode="Markdown",
            reply_markup=_main_menu(),
        )
        return ConversationHandler.END

    await query.message.reply_text(
        "📲 OTP sent to the registered device.\n\nPlease *reply with the OTP code*:",
        parse_mode="Markdown",
    )
    return AS.WAIT_OTP


async def recv_auth_otp(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    auth_sess = _get_auth_sess(ctx)
    otp       = update.message.text.strip()
    creds     = CRED_MAP[auth_sess.cred_type]

    await update.message.reply_text("🔄 Verifying OTP…")
    try:
        resp = auth_sess.http_session.post(
            f"{AUTH_BASE_URL}/otpverify",
            json={"username": creds["username"], "password": creds["password"], "otpcode": otp},
            timeout=30,
        )
        resp.raise_for_status()
        data          = resp.json()
        details       = data.get("details", {})
        access_token  = details.get("access_token")
        jwt           = details.get("jwt")
        refresh_token = details.get("refresh_token", "")
        if not access_token or not jwt:
            raise RuntimeError(f"Tokens missing. Keys: {list(data.keys())}")

        persist_tokens(auth_sess.cred_type, access_token, jwt, refresh_token)

        exp_ts  = _jwt_exp(jwt)
        exp_str = (
            datetime.fromtimestamp(exp_ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
            if exp_ts else "unknown"
        )
    except Exception as e:
        await update.message.reply_text(
            f"❌ OTP verification failed: `{e}`\n\nSend the OTP again or tap 🛑 Cancel.",
            parse_mode="Markdown",
        )
        return AS.WAIT_OTP

    await update.message.reply_text(
        f"✅ *Authenticated successfully!*\n\n"
        f"*Profile:* {CRED_LABELS[auth_sess.cred_type]}\n"
        f"*Token expires:* {exp_str}",
        parse_mode="Markdown",
        reply_markup=_main_menu(),
    )
    return ConversationHandler.END


# ──────────────────────────────────────────────────────────
# Bulk Export — NAIROBI Completed stamp-duty applications
# ──────────────────────────────────────────────────────────

_BE_LIST_URL   = f"{BASE_URL}/valuationservice/api/v1/stamp-duty/application"
_BE_DETAIL_URL = f"{BASE_URL}/valuationservice/api/v1/stamp-duty/application/detail-view"
_BE_PAGE_SIZE  = 10   # API default
_BE_LIST_WORKERS        = 5
_BE_DETAIL_WORKERS      = 5
_BE_MAX_RETRIES         = 3
_BE_TOKEN_ROTATE_DELAY  = 10   # seconds to wait before retrying with a new token


class _AllTokensExhausted(Exception):
    pass


class _TokenRotator:
    """Thread-safe token rotator — advances to the next valid credential on 403."""

    def __init__(self, token_pairs: List[tuple]):
        # token_pairs: [(cred_type, AuthTokens), ...]
        self._tokens = list(token_pairs)
        self._idx    = 0
        self._lock   = threading.Lock()

    def current(self) -> Optional["AuthTokens"]:
        with self._lock:
            return self._tokens[self._idx][1] if self._idx < len(self._tokens) else None

    def current_label(self) -> str:
        with self._lock:
            if self._idx < len(self._tokens):
                return CRED_LABELS.get(self._tokens[self._idx][0], self._tokens[self._idx][0])
            return "none"

    def rotate(self, failed: "AuthTokens") -> Optional["AuthTokens"]:
        """Advance past `failed` if it is still the current token. Returns new token or None."""
        with self._lock:
            if self._idx < len(self._tokens) and self._tokens[self._idx][1] is failed:
                self._idx += 1
            return self._tokens[self._idx][1] if self._idx < len(self._tokens) else None

    @property
    def exhausted(self) -> bool:
        with self._lock:
            return self._idx >= len(self._tokens)


_EXCEL_COLUMNS = [
    "Filter",
    "Reference Number",
    "Parcel Number",
    "Registry",
    "County",
    "Valuation Request Type",
    "Application Status",
    "Application Date Created",
    "Valuation Officer",
    "Date of Valuation",
    "Valuer Total Land Value (KES)",
    "Harmonized Total Land Value (KES)",
    "Document URL",
    "Enrich Error",
]


def _be_headers(tokens: AuthTokens) -> dict:
    return {
        "Authorization": f"Bearer {tokens.access_token}",
        "JWTAUTH":       f"Bearer {tokens.jwt}",
    }


def load_be_schedule() -> Optional[Dict]:
    try:
        with open(SAVED_BULK_EXPORT_SCHED_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return None


def save_be_schedule(cfg: Dict) -> None:
    _ensure_data_dir()
    with open(SAVED_BULK_EXPORT_SCHED_FILE, "w") as f:
        json.dump(cfg, f, indent=2)


def clear_be_schedule() -> None:
    try:
        os.remove(SAVED_BULK_EXPORT_SCHED_FILE)
    except FileNotFoundError:
        pass


def load_be_partial() -> Optional[Dict]:
    try:
        with open(SAVED_BULK_EXPORT_PARTIAL_FILE) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return None


def save_be_partial(county: str, registries: List[str], rows: List[dict], done_ids: List[str]) -> None:
    _ensure_data_dir()
    with open(SAVED_BULK_EXPORT_PARTIAL_FILE, "w") as f:
        json.dump({
            "saved_at":  datetime.now().isoformat(),
            "county":    county,
            "registries": registries,
            "rows":      rows,
            "done_ids":  done_ids,
        }, f)


def clear_be_partial() -> None:
    try:
        os.remove(SAVED_BULK_EXPORT_PARTIAL_FILE)
    except FileNotFoundError:
        pass


def _be_schedule_keyboard() -> InlineKeyboardMarkup:
    rows = []
    for label, secs in _BE_SCHEDULE_OPTIONS:
        rows.append([InlineKeyboardButton(label, callback_data=f"be_sched:{secs}")])
    return InlineKeyboardMarkup(rows)


def _be_county_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(label, callback_data=f"be_county:{key}")]
        for key, label in _BE_COUNTY_LABELS.items()
    ])


def _be_cred_keyboard() -> Optional[InlineKeyboardMarkup]:
    """Return an inline keyboard of credential profiles that currently have valid tokens.
    Returns None if no credentials are valid."""
    rows = [
        [InlineKeyboardButton(label, callback_data=f"be_cred:{key}")]
        for key, label in CRED_LABELS.items()
        if get_valid_tokens(key)
    ]
    return InlineKeyboardMarkup(rows) if rows else None


def _be_fetch_page(sess: requests.Session, headers: dict, page: int) -> dict:
    """Fetch one list page. Returns the parsed JSON dict."""
    params = {
        "filter":       "Completed",
        "role":         "DLV",
        "request_type": "STAMP_DUTY",
        "search":       "",
        "page":         page,
    }
    resp = sess.get(_BE_LIST_URL, headers=headers, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


def _be_fetch_detail(sess: requests.Session, rotator: "_TokenRotator", app_id: str) -> dict:
    """
    Fetch detail view for one application ID.
    On 403 rotates to the next valid token (with a short pause).
    Raises _AllTokensExhausted when no more tokens remain.
    """
    while True:
        tokens = rotator.current()
        if tokens is None:
            raise _AllTokensExhausted(f"All tokens exhausted fetching {app_id}")
        headers = _be_headers(tokens)
        try:
            resp = sess.get(_BE_DETAIL_URL, headers=headers, params={"request_id": app_id}, timeout=30)
            if resp.status_code == 403:
                new_tokens = rotator.rotate(tokens)
                if new_tokens is None:
                    raise _AllTokensExhausted(f"All tokens exhausted (403) fetching {app_id}")
                time.sleep(_BE_TOKEN_ROTATE_DELAY)
                continue
            if resp.status_code in (429, 502, 503, 504):
                time.sleep(2)
                continue
            resp.raise_for_status()
            return resp.json()
        except _AllTokensExhausted:
            raise
        except Exception as exc:
            raise exc


def _be_extract_row(detail: dict) -> dict:
    """Extract the target columns from a detail-view response dict."""
    # Valuation Officer from actors[]
    vo_name = ""
    vo_date = ""
    for actor in (detail.get("actors") or []):
        role = (actor.get("role") or "").upper()
        if role in ("VALUATION OFFICER", "VO"):
            vo_name = (actor.get("user_details") or {}).get("names", "")
            vo_date = actor.get("date_assigned", "")
            if role == "VALUATION OFFICER":
                break   # prefer exact match

    # Consideration amount from external_process_details
    ext = detail.get("external_process_details") or {}
    land_value = ext.get("consideration_amount", "")

    # Document URL: prefer VALUATION CERTIFICATE in process_documents
    doc_url = ""
    for pdoc in (detail.get("process_documents") or []):
        if (pdoc.get("document_name") or "").upper() == "VALUATION CERTIFICATE":
            doc_url = pdoc.get("document", "")
            break
    if not doc_url:
        app_docs = detail.get("application_documents") or []
        if app_docs:
            doc_url = app_docs[0].get("document", "")

    return {
        "Filter":                          "Completed",
        "Reference Number":                detail.get("reference_number", ""),
        "Parcel Number":                   detail.get("parcel_number", ""),
        "Registry":                        detail.get("registry", ""),
        "County":                          detail.get("county", ""),
        "Valuation Request Type":          detail.get("valuation_request_type", ""),
        "Application Status":              detail.get("application_status", ""),
        "Application Date Created":        detail.get("date_created", ""),
        "Valuation Officer":               vo_name,
        "Date of Valuation":               vo_date,
        "Valuer Total Land Value (KES)":   land_value,
        "Harmonized Total Land Value (KES)": detail.get("harmonized_total_land_value", ""),
        "Document URL":                    doc_url,
        "Enrich Error":                    "",
    }


def _be_build_excel(rows: List[dict]) -> bytes:
    """Build the formatted Excel workbook and return the raw bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stamp Duty Valuations"

    header_font  = Font(bold=True)
    header_fill  = PatternFill("solid", fgColor="BDD7EE")
    date_fmt     = "YYYY-MM-DD HH:MM:SS"
    number_fmt   = "#,##0"
    date_cols    = {"Application Date Created", "Date of Valuation"}
    number_cols  = {"Valuer Total Land Value (KES)", "Harmonized Total Land Value (KES)"}

    # Write header
    ws.append(_EXCEL_COLUMNS)
    for col_idx, col_name in enumerate(_EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font  = header_font
        cell.fill  = header_fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"

    # Write data rows
    for row_data in rows:
        row_vals = [row_data.get(col, "") for col in _EXCEL_COLUMNS]
        ws.append(row_vals)
        row_idx = ws.max_row
        for col_idx, col_name in enumerate(_EXCEL_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name in date_cols and cell.value:
                cell.number_format = date_fmt
            elif col_name in number_cols and cell.value not in ("", None, "FETCH_ERROR"):
                try:
                    cell.value         = float(str(cell.value).replace(",", "").strip())
                    cell.number_format = number_fmt
                except (ValueError, TypeError):
                    pass

    # Auto-fit column widths (min 15, max 50)
    for col_idx, col_name in enumerate(_EXCEL_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len    = len(col_name)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            val = str(row[0].value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(15, min(50, max_len + 2))

    # ── Summary sheet ─────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")

    def _write_section_header(ws, row: int, text: str):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill("solid", fgColor="BDD7EE")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    def _write_col_headers(ws, row: int, *headers):
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = PatternFill("solid", fgColor="D9E1F2")

    # ── Collect stats from rows ───────────────────────────────
    from collections import defaultdict

    monthly: dict  = defaultdict(int)    # "YYYY-MM" → count
    yearly:  dict  = defaultdict(int)    # "YYYY"    → count
    vo_tasks: dict = defaultdict(int)    # officer name → count
    total_land_value = 0.0
    total_harmonized = 0.0

    for r in rows:
        # Date of Valuation — monthly/yearly distribution
        dov = str(r.get("Date of Valuation") or "")
        if len(dov) >= 7:
            monthly[dov[:7]] += 1
        if len(dov) >= 4:
            yearly[dov[:4]] += 1

        # Valuation Officer task count
        vo = (r.get("Valuation Officer") or "").strip()
        if vo and vo != "FETCH_ERROR":
            vo_tasks[vo] += 1

        # Sum land values
        for key, target in (
            ("Valuer Total Land Value (KES)", "land"),
            ("Harmonized Total Land Value (KES)", "harm"),
        ):
            raw = r.get(key, "")
            if raw not in ("", None, "FETCH_ERROR"):
                try:
                    val = float(str(raw).replace(",", "").strip())
                    if key == "Valuer Total Land Value (KES)":
                        total_land_value += val
                    else:
                        total_harmonized += val
                except (ValueError, TypeError):
                    pass

    cur_row = 1

    # ── Section 1: Totals ─────────────────────────────────────
    _write_section_header(ws2, cur_row, "Overall Totals")
    cur_row += 1
    for label, value in (
        ("Total Records",                        len(rows)),
        ("Valuer Total Land Value (KES)",        total_land_value),
        ("Harmonized Total Land Value (KES)",    total_harmonized),
    ):
        ws2.cell(row=cur_row, column=1, value=label).font = Font(bold=True)
        c = ws2.cell(row=cur_row, column=2, value=value)
        if isinstance(value, float):
            c.number_format = number_fmt
        cur_row += 1

    cur_row += 1  # blank row

    # ── Section 2: Monthly distribution ──────────────────────
    _write_section_header(ws2, cur_row, "Monthly Distribution (Date of Valuation)")
    cur_row += 1
    _write_col_headers(ws2, cur_row, "Month (YYYY-MM)", "Count")
    cur_row += 1
    for month in sorted(monthly):
        ws2.cell(row=cur_row, column=1, value=month)
        ws2.cell(row=cur_row, column=2, value=monthly[month])
        cur_row += 1

    cur_row += 1  # blank row

    # ── Section 3: Yearly distribution ───────────────────────
    _write_section_header(ws2, cur_row, "Yearly Distribution (Date of Valuation)")
    cur_row += 1
    _write_col_headers(ws2, cur_row, "Year", "Count")
    cur_row += 1
    for year in sorted(yearly):
        ws2.cell(row=cur_row, column=1, value=year)
        ws2.cell(row=cur_row, column=2, value=yearly[year])
        cur_row += 1

    cur_row += 1  # blank row

    # ── Section 4: Valuation Officer task counts ──────────────
    _write_section_header(ws2, cur_row, "Tasks per Valuation Officer")
    cur_row += 1
    _write_col_headers(ws2, cur_row, "Valuation Officer", "Tasks")
    cur_row += 1
    for officer, count in sorted(vo_tasks.items(), key=lambda x: -x[1]):
        ws2.cell(row=cur_row, column=1, value=officer)
        ws2.cell(row=cur_row, column=2, value=count)
        cur_row += 1

    # Auto-fit Summary sheet columns
    for col_idx in (1, 2):
        col_letter = get_column_letter(col_idx)
        max_len = 20
        for row in ws2.iter_rows(min_col=col_idx, max_col=col_idx):
            val = str(row[0].value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws2.column_dimensions[col_letter].width = max(20, min(50, max_len + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _send_bulk_export_email(to_email: str, filename: str, xlsx_bytes: bytes) -> None:
    """Send the Excel file as an email attachment. Raises on failure."""
    if not SMTP_USER or not SMTP_PASS:
        raise RuntimeError("SMTP_USER / SMTP_PASS not configured in .env")

    msg            = MIMEMultipart()
    msg["Subject"] = f"Ardhisasa Export Valuation Report — {filename}"
    msg["From"]    = SMTP_USER
    msg["To"]      = to_email

    body = (
        f"Please find attached the Ardhisasa stamp-duty bulk export.\n\n"
        f"File: {filename}\n"
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
    )
    msg.attach(MIMEText(body, "plain"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(xlsx_bytes)
    _email_encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.ehlo()
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_USER, to_email, msg.as_string())


def _bulk_export_run(tokens: AuthTokens, chat_id: int, email: str, bot, loop,
                     registries: List[str] = None, county: str = "") -> None:
    """
    Full synchronous export worker — runs in a background thread.

    Behaviour:
    - Resumes from a saved partial checkpoint when one exists for the same county/registries.
    - Sorts all records by Application Date Created (ascending) so resume is date-ordered.
    - Rotates tokens on 403: tries every valid credential before giving up.
    - On full token exhaustion saves a partial checkpoint and notifies the user.
    - On success clears the partial checkpoint and sends the Excel file.
    """
    def _tg(text: str):
        asyncio.run_coroutine_threadsafe(
            bot.send_message(chat_id, text, parse_mode="Markdown"),
            loop,
        ).result(timeout=15)

    def _set_status(**kwargs):
        _BE_STATUS.setdefault(chat_id, {}).update(kwargs)

    _BE_STATUS[chat_id] = {
        "phase":          "fetching pages",
        "started_at":     datetime.now(),
        "total":          None,
        "pages_done":     1,
        "total_pages":    None,
        "details_done":   0,
        "details_total":  None,
        "errors":         0,
        "completed_at":   None,
        "rows":           None,
        "error_msg":      None,
    }

    # ── Build token rotator from all currently-valid credentials ──
    token_pairs = [
        (ct, get_valid_tokens(ct))
        for ct in CRED_MAP
        if get_valid_tokens(ct)
    ]
    # Put the originally-chosen token first
    token_pairs.sort(key=lambda p: 0 if p[1] is tokens else 1)
    rotator = _TokenRotator(token_pairs)

    sess = build_session()
    adapter = requests.adapters.HTTPAdapter(
        pool_connections=_BE_DETAIL_WORKERS,
        pool_maxsize=_BE_DETAIL_WORKERS,
        max_retries=0,
    )
    sess.mount("https://", adapter)
    sess.mount("http://",  adapter)
    list_headers = _be_headers(tokens)

    try:
        # ── Step 1: fetch list pages ───────────────────────────────
        first_page  = _be_fetch_page(sess, list_headers, 1)
        total       = first_page.get("count", 0)
        results     = list(first_page.get("results") or [])
        page_size   = len(results) if results else _BE_PAGE_SIZE
        if page_size == 0:
            page_size = _BE_PAGE_SIZE
        total_pages = max(1, -(-total // page_size))

        logger.info("Bulk export: count=%d, page_size=%d, total_pages=%d", total, page_size, total_pages)
        _set_status(total=total, total_pages=total_pages, pages_done=1)

        if total_pages > 1:
            with ThreadPoolExecutor(max_workers=_BE_LIST_WORKERS) as pool:
                futures = {pool.submit(_be_fetch_page, sess, list_headers, p): p
                           for p in range(2, total_pages + 1)}
                for fut in _futures_as_completed(futures):
                    results.extend(fut.result().get("results") or [])
                    _set_status(pages_done=_BE_STATUS[chat_id]["pages_done"] + 1)

        # ── Client-side registry filter ────────────────────────────
        reg_set = {r.upper() for r in (registries or [])}
        if reg_set:
            filtered = [r for r in results if (r.get("registry") or "").upper() in reg_set]
        else:
            filtered = results

        # Sort by Application Date Created ascending so resume is date-ordered
        filtered.sort(key=lambda r: (r.get("date_created") or ""))

        # ── Resume: load partial checkpoint if it matches ──────────
        partial     = load_be_partial()
        resume_rows: List[dict] = []
        done_ids:    set        = set()

        if partial and partial.get("county") == county and \
                set(partial.get("registries", [])) == set(registries or []):
            resume_rows = partial.get("rows") or []
            done_ids    = set(partial.get("done_ids") or [])
            _tg(
                f"♻️ Resuming from checkpoint — {len(resume_rows):,} rows already saved, "
                f"{len(done_ids):,} IDs done."
            )

        id_list = [r["id"] for r in filtered if r.get("id") and r["id"] not in done_ids]

        if not id_list and not resume_rows:
            _set_status(phase="done", completed_at=datetime.now(), rows=0)
            _tg("ℹ️ Bulk export complete — no records found.")
            return

        if not id_list:
            # All IDs already done from checkpoint — skip straight to Excel
            rows = resume_rows
        else:
            # ── Step 2: parallel detail fetch with token rotation ──────
            _set_status(phase="fetching details", details_total=len(id_list))
            rows: List[dict] = list(resume_rows)
            current_done_ids = list(done_ids)
            exhausted_flag   = threading.Event()

            with ThreadPoolExecutor(max_workers=_BE_DETAIL_WORKERS) as pool:
                futures = {pool.submit(_be_fetch_detail, sess, rotator, app_id): app_id
                           for app_id in id_list}
                for fut in _futures_as_completed(futures):
                    app_id = futures[fut]
                    try:
                        if exhausted_flag.is_set():
                            # Tokens already gone — skip remaining futures
                            fut.cancel()
                            continue
                        detail = fut.result()
                        rows.append(_be_extract_row(detail))
                        current_done_ids.append(app_id)
                        _set_status(details_done=_BE_STATUS[chat_id]["details_done"] + 1)
                    except _AllTokensExhausted:
                        exhausted_flag.set()
                        logger.warning("Bulk export: all tokens exhausted at id=%s", app_id)
                    except Exception as exc:
                        logger.warning("Bulk export detail failed id=%s: %s", app_id, exc)
                        _set_status(
                            details_done=_BE_STATUS[chat_id]["details_done"] + 1,
                            errors=_BE_STATUS[chat_id]["errors"] + 1,
                        )

            if exhausted_flag.is_set():
                # Save whatever we managed to fetch and bail out
                save_be_partial(county, list(registries or []), rows, current_done_ids)
                remaining = len(id_list) - len(current_done_ids) + len(done_ids)
                _set_status(phase="paused — tokens exhausted", completed_at=datetime.now(), rows=len(rows))
                _tg(
                    f"⚠️ All tokens returned 403 — export paused.\n\n"
                    f"*Saved so far:* {len(rows):,} rows\n"
                    f"*Remaining:* ~{remaining:,} records\n\n"
                    "Refresh your tokens and run the export again to continue from this checkpoint."
                )
                return

        # ── Step 3: sort final rows by date, build Excel, send ────
        _set_status(phase="building excel")
        rows.sort(key=lambda r: (r.get("Application Date Created") or ""))

        clear_be_partial()
        filename   = f"Ardhisasa_Valuation_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        xlsx_bytes = _be_build_excel(rows)

        asyncio.run_coroutine_threadsafe(
            bot.send_document(
                chat_id,
                document=io.BytesIO(xlsx_bytes),
                filename=filename,
                caption=f"📊 Bulk export complete — {len(rows):,} rows",
            ),
            loop,
        ).result(timeout=60)

        _set_status(phase="done", completed_at=datetime.now(), rows=len(rows))

        if email:
            try:
                _send_bulk_export_email(email, filename, xlsx_bytes)
                _tg(f"📧 File also sent to *{email}*.")
            except Exception as exc:
                logger.warning("Bulk export email failed: %s", exc)
                _tg(f"⚠️ Email delivery failed: `{exc}`")

    except Exception as exc:
        logger.error("Bulk export worker crashed: %s", exc, exc_info=True)
        _set_status(phase="failed", completed_at=datetime.now(), error_msg=str(exc))
        _tg(f"❌ Export failed: `{exc}`")


# ──────────────────────────────────────────────────────────
# Job Distribution Analysis
# ──────────────────────────────────────────────────────────

_TEAMS_URL        = f"{BASE_URL}/acl/api/v1/list-teams"
_TEAM_MEMBERS_URL = f"{BASE_URL}/acl/api/v1/staff-teams/get-team-members"
_JD_ONGOING_URL   = f"{BASE_URL}/valuationservice/api/v1/stamp-duty/application"
_JD_DETAIL_URL    = f"{BASE_URL}/valuationservice/api/v1/stamp-duty/application/detail-view"
_JD_WORKERS       = 5


@dataclass
class JDSession:
    cred_type: str = ""


def _get_jd_sess(ctx: ContextTypes.DEFAULT_TYPE) -> JDSession:
    if "jd_session" not in ctx.user_data:
        ctx.user_data["jd_session"] = JDSession()
    return ctx.user_data["jd_session"]


def _jd_headers(tokens: AuthTokens) -> dict:
    return {
        "Authorization": f"Bearer {tokens.access_token}",
        "JWTAUTH":       f"Bearer {tokens.jwt}",
        "cparams":       CPARAMS_DLV,
    }


def _jd_fetch_teams(sess: requests.Session, headers: dict) -> List[dict]:
    """Fetch all teams (not paginated in practice — count is small)."""
    resp = sess.get(_TEAMS_URL, headers=headers, params={"page": 1, "search": ""}, timeout=30)
    resp.raise_for_status()
    return resp.json().get("results") or []


def _jd_fetch_team_members(sess: requests.Session, headers: dict, team_id: str) -> List[dict]:
    """Fetch all members of a team across all pages."""
    members: List[dict] = []
    page = 1
    while True:
        resp = sess.get(
            _TEAM_MEMBERS_URL,
            headers=headers,
            params={"team_id": team_id, "page": page, "search": ""},
            timeout=30,
        )
        resp.raise_for_status()
        data    = resp.json()
        results = data.get("results") or []
        members.extend(results)
        if not data.get("next"):
            break
        page += 1
    return members


def _jd_fetch_ongoing_page(sess: requests.Session, headers: dict, page: int) -> dict:
    """Fetch one page of Ongoing stamp-duty tasks."""
    resp = sess.get(
        _JD_ONGOING_URL,
        headers=headers,
        params={
            "filter":       "Ongoing",
            "role":         "DLV",
            "request_type": "STAMP_DUTY",
            "search":       "",
            "page":         page,
        },
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def _jd_fetch_task_detail(sess: requests.Session, rotator: "_TokenRotator", task_id: str) -> dict:
    """Fetch detail for one task with token rotation on 403."""
    while True:
        tokens = rotator.current()
        if tokens is None:
            raise _AllTokensExhausted(f"All tokens exhausted fetching task {task_id}")
        headers = _jd_headers(tokens)
        try:
            resp = sess.get(_JD_DETAIL_URL, headers=headers, params={"request_id": task_id}, timeout=30)
            if resp.status_code == 403:
                new_tokens = rotator.rotate(tokens)
                if new_tokens is None:
                    raise _AllTokensExhausted(f"All tokens exhausted (403) fetching task {task_id}")
                time.sleep(_BE_TOKEN_ROTATE_DELAY)
                continue
            if resp.status_code in (429, 502, 503, 504):
                time.sleep(2)
                continue
            resp.raise_for_status()
            return resp.json()
        except _AllTokensExhausted:
            raise
        except Exception as exc:
            raise exc


def _jd_build_excel(
    teams: List[dict],
    members_by_team: Dict[str, List[dict]],
    tasks_by_userid: Dict[str, List[dict]],
    unassigned_tasks: List[dict],
) -> bytes:
    """Build the Job Distribution Excel workbook and return raw bytes."""
    from collections import defaultdict
    wb       = openpyxl.Workbook()
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="BDD7EE")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")

    def _header_row(ws, cols):
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell       = ws.cell(row=1, column=c)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes    = "A2"

    def _autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = max(12, min(55, max_len + 2))

    # ── Sheet 1: Team Summary ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Team Summary"
    _header_row(ws1, [
        "Team Name", "Min Amount (KES)", "Max Amount (KES)",
        "Total Members", "Available", "Not Available",
        "Assigned Tasks", "Unassigned Tasks", "Assigned %",
    ])

    total_assigned   = sum(len(v) for v in tasks_by_userid.values())
    total_unassigned = len(unassigned_tasks)

    for i, team in enumerate(teams, start=2):
        tid     = team["id"]
        members = members_by_team.get(tid, [])
        available     = sum(1 for m in members if m.get("availability") == "AVAILABLE")
        not_available = len(members) - available
        assigned      = sum(len(tasks_by_userid.get(m["userid"], [])) for m in members)
        unassigned    = sum(
            1 for t in unassigned_tasks
            if True  # unassigned tasks have no team association in list; show global
        ) if i == 2 else 0   # only once on first team row for global total
        total_for_team = assigned + (total_unassigned if i == 2 else 0)
        pct = f"{assigned / total_for_team * 100:.1f}%" if total_for_team else "N/A"

        row = [
            team.get("team_name", ""),
            team.get("min_amount", ""),
            team.get("max_amount", ""),
            len(members),
            available,
            not_available,
            assigned,
            total_unassigned if i == 2 else "",
            pct,
        ]
        ws1.append(row)
        if i % 2 == 0:
            for c in range(1, 10):
                ws1.cell(row=i, column=c).fill = alt_fill
    _autofit(ws1)

    # ── Sheet 2: Member Distribution ──────────────────────
    ws2 = wb.create_sheet("Member Distribution")
    _header_row(ws2, [
        "Team", "Name", "Account Number", "Availability",
        "Registry", "Tasks Assigned", "Reference Numbers", "Analysis",
    ])

    warn_fill = PatternFill("solid", fgColor="FFE0B2")   # amber for out-of-range rows

    row_idx = 2
    for team in teams:
        tid     = team["id"]
        members = members_by_team.get(tid, [])
        members_sorted = sorted(
            members,
            key=lambda m: -len(tasks_by_userid.get(m.get("userid", ""), [])),
        )
        team_min = float(team.get("min_amount") or 0)
        team_max = float(team.get("max_amount") or float("inf"))

        for m in members_sorted:
            uid   = m.get("userid", "")
            tasks = tasks_by_userid.get(uid, [])

            in_range:  List[str] = []
            out_range: List[str] = []

            for t in tasks:
                ref    = t.get("reference_number", t.get("id", ""))
                amount = t.get("consideration_amount", "")
                if amount == "":
                    label = ref
                else:
                    try:
                        amt   = float(amount)
                        label = f"{ref}({int(amt):,})"
                        if team_min <= amt <= team_max:
                            in_range.append(label)
                        else:
                            out_range.append(label)
                        continue
                    except (ValueError, TypeError):
                        label = ref
                in_range.append(label)   # no amount → assume in range

            refs = ", ".join(in_range + out_range)

            if out_range:
                analysis = f"⚠️ Out of range: {', '.join(out_range)}"
                if in_range:
                    analysis += f" | ✅ In range: {len(in_range)}"
            elif in_range:
                analysis = f"✅ All {len(in_range)} in range"
            else:
                analysis = ""

            ws2.append([
                team.get("team_name", ""),
                m.get("name", ""),
                m.get("account_number", ""),
                m.get("availability", ""),
                m.get("registry", ""),
                len(tasks),
                refs,
                analysis,
            ])
            fill = warn_fill if out_range else (alt_fill if row_idx % 2 == 0 else None)
            if fill:
                for c in range(1, 9):
                    ws2.cell(row=row_idx, column=c).fill = fill
            row_idx += 1
    _autofit(ws2)

    # ── Sheet 3: Unassigned / No Valuer Tasks ─────────────
    ws3 = wb.create_sheet("Unassigned Tasks")
    _header_row(ws3, [
        "Reference Number", "Parcel Number", "Registry",
        "County", "Date Created", "Status",
    ])
    for i, t in enumerate(unassigned_tasks, start=2):
        ws3.append([
            t.get("reference_number", ""),
            t.get("parcel_number", ""),
            t.get("registry", ""),
            t.get("county", ""),
            t.get("date_created", ""),
            t.get("status", ""),
        ])
        if i % 2 == 0:
            for c in range(1, 7):
                ws3.cell(row=i, column=c).fill = alt_fill
    _autofit(ws3)

    # ── Sheet 4: Team Roster & Cross-Team Members ──────────
    ws4 = wb.create_sheet("Team Roster")

    # Build userid → list of team names (to detect multi-team members)
    userid_to_teams: Dict[str, List[str]] = {}
    for team in teams:
        for m in members_by_team.get(team["id"], []):
            uid = m.get("userid", "")
            userid_to_teams.setdefault(uid, []).append(team.get("team_name", ""))

    multi_team_fill = PatternFill("solid", fgColor="FFF176")   # yellow for multi-team rows

    # ── Section A: per-team roster ─────────────────────────
    ws4.append(["TEAM ROSTER"])
    ws4.cell(row=ws4.max_row, column=1).font = Font(bold=True, size=13)
    ws4.append([])

    current_row = 3
    for team in teams:
        # Team header
        ws4.append([team.get("team_name", ""), f"({len(members_by_team.get(team['id'], []))} members)"])
        for c in range(1, 3):
            cell      = ws4.cell(row=ws4.max_row, column=c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="BDD7EE")
        current_row = ws4.max_row + 1

        # Column headers
        ws4.append(["#", "Name", "Account Number", "Availability", "Registry", "Also In Teams"])
        for c in range(1, 7):
            cell      = ws4.cell(row=ws4.max_row, column=c)
            cell.font = hdr_font
            cell.fill = PatternFill("solid", fgColor="D9E1F2")

        members = sorted(members_by_team.get(team["id"], []), key=lambda m: m.get("name", ""))
        for idx, m in enumerate(members, start=1):
            uid        = m.get("userid", "")
            other_teams = [t for t in userid_to_teams.get(uid, []) if t != team.get("team_name", "")]
            also_in    = ", ".join(other_teams) if other_teams else ""
            ws4.append([
                idx,
                m.get("name", ""),
                m.get("account_number", ""),
                m.get("availability", ""),
                m.get("registry", ""),
                also_in,
            ])
            if other_teams:
                for c in range(1, 7):
                    ws4.cell(row=ws4.max_row, column=c).fill = multi_team_fill

        ws4.append([])   # blank row between teams

    # ── Section B: members in multiple teams ──────────────
    ws4.append([])
    ws4.append(["MEMBERS IN MULTIPLE TEAMS"])
    ws4.cell(row=ws4.max_row, column=1).font = Font(bold=True, size=13)
    ws4.cell(row=ws4.max_row, column=1).fill = PatternFill("solid", fgColor="FFF176")

    ws4.append(["Name", "Account Number", "Teams"])
    for c in range(1, 4):
        cell      = ws4.cell(row=ws4.max_row, column=c)
        cell.font = hdr_font
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    # collect all members once (avoid duplicates from members_by_team)
    seen_multi: set = set()
    for team in teams:
        for m in members_by_team.get(team["id"], []):
            uid        = m.get("userid", "")
            team_names = userid_to_teams.get(uid, [])
            if len(team_names) > 1 and uid not in seen_multi:
                seen_multi.add(uid)
                ws4.append([
                    m.get("name", ""),
                    m.get("account_number", ""),
                    ", ".join(team_names),
                ])
                for c in range(1, 4):
                    ws4.cell(row=ws4.max_row, column=c).fill = multi_team_fill

    if not seen_multi:
        ws4.append(["No members belong to more than one team."])

    _autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _jd_run(tokens: AuthTokens, chat_id: int, bot, loop) -> None:
    """Background worker: fetch teams + task distribution, build Excel, send."""
    def _tg(text: str):
        asyncio.run_coroutine_threadsafe(
            bot.send_message(chat_id, text, parse_mode="Markdown"),
            loop,
        ).result(timeout=15)

    def _set_status(**kwargs):
        _JD_STATUS.setdefault(chat_id, {}).update(kwargs)

    _JD_STATUS[chat_id] = {
        "phase":         "fetching teams",
        "started_at":    datetime.now(),
        "teams_count":   None,
        "members_total": None,
        "tasks_total":   None,
        "tasks_done":    0,
        "errors":        0,
        "completed_at":  None,
        "rows":          None,
        "error_msg":     None,
    }

    token_pairs = [(ct, get_valid_tokens(ct)) for ct in CRED_MAP if get_valid_tokens(ct)]
    token_pairs.sort(key=lambda p: 0 if p[1] is tokens else 1)
    rotator = _TokenRotator(token_pairs)

    sess    = build_session()
    adapter = requests.adapters.HTTPAdapter(
        pool_connections=_JD_WORKERS, pool_maxsize=_JD_WORKERS, max_retries=0
    )
    sess.mount("https://", adapter)
    sess.mount("http://",  adapter)
    headers = _jd_headers(tokens)

    try:
        # ── Step 1: teams + members ────────────────────────────────
        teams = _jd_fetch_teams(sess, headers)
        _set_status(teams_count=len(teams))
        _tg(f"📋 Found *{len(teams)}* teams. Fetching members…")

        members_by_team: Dict[str, List[dict]] = {}
        all_members:     Dict[str, dict]       = {}   # userid → member record (with team_name)
        for team in teams:
            members = _jd_fetch_team_members(sess, headers, team["id"])
            members_by_team[team["id"]] = members
            for m in members:
                m["_team_name"] = team.get("team_name", "")
                all_members[m["userid"]] = m

        total_members = sum(len(v) for v in members_by_team.values())
        _set_status(phase="fetching ongoing tasks", members_total=total_members)
        _tg(f"👥 *{total_members}* team members loaded. Fetching ongoing tasks…")

        # ── Step 2: ongoing tasks list ─────────────────────────────
        first_page  = _jd_fetch_ongoing_page(sess, headers, 1)
        tasks_total = first_page.get("count", 0)
        task_list   = list(first_page.get("results") or [])
        page_size   = len(task_list) if task_list else 10
        if page_size == 0:
            page_size = 10
        total_pages = max(1, -(-tasks_total // page_size))

        _set_status(tasks_total=tasks_total)

        if total_pages > 1:
            with ThreadPoolExecutor(max_workers=_JD_WORKERS) as pool:
                futures = {pool.submit(_jd_fetch_ongoing_page, sess, headers, p): p
                           for p in range(2, total_pages + 1)}
                for fut in _futures_as_completed(futures):
                    task_list.extend(fut.result().get("results") or [])

        _set_status(phase="fetching task details", tasks_total=len(task_list))
        _tg(f"📄 *{len(task_list)}* ongoing tasks. Fetching assignment details…")

        # ── Step 3: detail fetch for each task ─────────────────────
        tasks_by_userid:  Dict[str, List[dict]] = {}   # userid → [task_summary, ...]
        unassigned_tasks: List[dict]            = []
        exhausted_flag = threading.Event()

        with ThreadPoolExecutor(max_workers=_JD_WORKERS) as pool:
            futures = {pool.submit(_jd_fetch_task_detail, sess, rotator, t["id"]): t
                       for t in task_list}
            for fut in _futures_as_completed(futures):
                task_summary = futures[fut]
                try:
                    if exhausted_flag.is_set():
                        fut.cancel()
                        continue
                    detail = fut.result()
                    actors = detail.get("actors") or []
                    vo     = next((a for a in actors if a.get("role") == "VALUATION OFFICER"), None)
                    if vo:
                        uid = (vo.get("user_details") or {}).get("id", "")
                        ext = detail.get("external_process_details") or {}
                        tasks_by_userid.setdefault(uid, []).append({
                            "reference_number":   detail.get("reference_number", ""),
                            "parcel_number":      detail.get("parcel_number", ""),
                            "registry":           detail.get("registry", ""),
                            "date_created":       detail.get("date_created", ""),
                            "consideration_amount": ext.get("consideration_amount", ""),
                        })
                    else:
                        unassigned_tasks.append(task_summary)
                    _set_status(tasks_done=_JD_STATUS[chat_id]["tasks_done"] + 1)
                except _AllTokensExhausted:
                    exhausted_flag.set()
                    unassigned_tasks.append(task_summary)
                    logger.warning("JD: all tokens exhausted at task %s", task_summary.get("id"))
                except Exception as exc:
                    logger.warning("JD detail failed task=%s: %s", task_summary.get("id"), exc)
                    unassigned_tasks.append(task_summary)
                    _set_status(
                        tasks_done=_JD_STATUS[chat_id]["tasks_done"] + 1,
                        errors=_JD_STATUS[chat_id]["errors"] + 1,
                    )

        if exhausted_flag.is_set():
            _tg("⚠️ Tokens exhausted during detail fetch — report will be partial. Refresh tokens and re-run for a complete picture.")

        # ── Step 4: build Excel and send ──────────────────────────
        _set_status(phase="building excel")
        assigned_count   = sum(len(v) for v in tasks_by_userid.values())
        unassigned_count = len(unassigned_tasks)
        _tg(
            f"✅ Analysis complete.\n"
            f"• Assigned: *{assigned_count}* tasks\n"
            f"• Unassigned / no VO: *{unassigned_count}* tasks\n"
            f"Building Excel report…"
        )

        filename   = f"Ardhisasa_Job_Distribution_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        xlsx_bytes = _jd_build_excel(teams, members_by_team, tasks_by_userid, unassigned_tasks)

        asyncio.run_coroutine_threadsafe(
            bot.send_document(
                chat_id,
                document=io.BytesIO(xlsx_bytes),
                filename=filename,
                caption=(
                    f"🏆 Job Distribution Report\n"
                    f"Teams: {len(teams)} | Members: {total_members} | "
                    f"Assigned: {assigned_count} | Unassigned: {unassigned_count}"
                ),
            ),
            loop,
        ).result(timeout=60)

        _set_status(phase="done", completed_at=datetime.now(), rows=total_members)

    except Exception as exc:
        logger.error("JD worker crashed: %s", exc, exc_info=True)
        _set_status(phase="failed", completed_at=datetime.now(), error_msg=str(exc))
        _tg(f"❌ Job distribution failed: `{exc}`")


# ── Job Distribution conversation handlers ────────────────

async def cmd_job_distribution(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    sess           = _get_jd_sess(ctx)
    sess.cred_type = ""

    kbd = _be_cred_keyboard()
    if not kbd:
        await update.message.reply_text(
            "❌ No valid cached tokens. Use *🔑 Refresh Auth* first.",
            parse_mode="Markdown",
            reply_markup=_main_menu(),
        )
        return ConversationHandler.END

    await update.message.reply_text(
        "🏆 *Job Distribution Analysis*\n\n"
        "This report shows how ongoing tasks are distributed across team members.\n\n"
        "👤 *Select the account to run the analysis as:*",
        parse_mode="Markdown",
        reply_markup=kbd,
    )
    return JD.PICK_CRED


async def recv_jd_cred(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    query = update.callback_query
    await query.answer()
    sess           = _get_jd_sess(ctx)
    sess.cred_type = query.data.split(":")[1]
    cred_label     = CRED_LABELS.get(sess.cred_type, sess.cred_type)

    await query.edit_message_text(
        f"✅ Account: *{cred_label}*\n\n"
        "The analysis will:\n"
        "• Fetch all teams and their members\n"
        "• Fetch all Ongoing stamp-duty tasks\n"
        "• Identify the assigned Valuation Officer per task\n"
        "• Export a 3-sheet Excel: Team Summary, Member Distribution, Unassigned Tasks\n\n"
        "Tap *Run Analysis* to start.",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("▶️ Run Analysis", callback_data="jd:yes"),
            InlineKeyboardButton("❌ Cancel",       callback_data="jd:no"),
        ]]),
    )
    return JD.CONFIRM


async def recv_jd_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    query = update.callback_query
    await query.answer()

    if query.data == "jd:no":
        await query.edit_message_text("❌ Analysis cancelled.")
        await ctx.bot.send_message(query.message.chat_id, "Main menu.", reply_markup=_main_menu())
        return ConversationHandler.END

    sess   = _get_jd_sess(ctx)
    tokens = get_valid_tokens(sess.cred_type)
    if not tokens:
        cred_label = CRED_LABELS.get(sess.cred_type, sess.cred_type)
        await query.edit_message_text(
            f"❌ Tokens for *{cred_label}* have expired. Use *🔑 Refresh Auth* first.",
            parse_mode="Markdown",
        )
        await ctx.bot.send_message(query.message.chat_id, "Main menu.", reply_markup=_main_menu())
        return ConversationHandler.END

    await query.edit_message_text("⏳ Analysis running in background — you will be notified when done.")
    await ctx.bot.send_message(query.message.chat_id, "Returning to menu.", reply_markup=_main_menu())

    loop = asyncio.get_event_loop()
    asyncio.ensure_future(
        asyncio.to_thread(_jd_run, tokens, query.message.chat_id, ctx.bot, loop)
    )
    return ConversationHandler.END


# ── Bulk Export conversation handlers ─────────────────────

async def _bulk_export_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    """APScheduler job: run bulk export with saved schedule config."""
    cfg = load_be_schedule()
    if not cfg:
        return
    cred_type = cfg.get("cred_type")
    tokens    = get_valid_tokens(cred_type) if cred_type else _any_valid_tokens()
    if not tokens:
        cred_label = CRED_LABELS.get(cred_type, cred_type or "any")
        logger.warning("Bulk export job: no valid tokens for %s — skipping.", cred_label)
        return
    chat_id    = cfg["chat_id"]
    email      = cfg.get("email", "")
    registries = cfg.get("registries", [])
    county     = cfg.get("county", "")
    loop       = asyncio.get_event_loop()
    asyncio.ensure_future(
        asyncio.to_thread(_bulk_export_run, tokens, chat_id, email, context.bot, loop, registries, county)
    )


async def cmd_bulk_export(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    sess = _get_be_sess(ctx)
    sess.county     = ""
    sess.registries = []
    sess.email      = ""
    sess.cred_type  = ""
    await update.message.reply_text(
        "📤 *Export Valuation Report*\n\nSelect the county to export:",
        parse_mode="Markdown",
        reply_markup=_be_county_keyboard(),
    )
    return BE.COUNTY


async def cmd_export_status(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    chat_id = update.effective_chat.id
    be_st   = _BE_STATUS.get(chat_id)
    jd_st   = _JD_STATUS.get(chat_id)

    if not be_st and not jd_st:
        await update.message.reply_text(
            "ℹ️ No export or analysis has been run in this session.",
            reply_markup=_main_menu(),
        )
        return

    def _elapsed(started_at, completed_at):
        if not started_at:
            return ""
        ref  = completed_at or datetime.now()
        secs = int((ref - started_at).total_seconds())
        return f"{secs // 60}m {secs % 60}s"

    phase_icons = {
        "fetching pages":        "📄",
        "fetching details":      "🔍",
        "fetching teams":        "🔍",
        "fetching ongoing tasks":"📋",
        "fetching task details": "🔍",
        "building excel":        "📊",
        "done":                  "✅",
        "failed":                "❌",
        "paused — tokens exhausted": "⏸",
    }

    all_lines = []

    if be_st:
        phase        = be_st.get("phase", "unknown")
        started_at   = be_st.get("started_at")
        completed_at = be_st.get("completed_at")
        icon         = phase_icons.get(phase, "⏳")
        lines        = [f"*{icon} Export Valuation Report*\n"]
        elapsed      = _elapsed(started_at, completed_at)
        if started_at:
            lines.append(f"Started: `{started_at.strftime('%H:%M:%S')}`")
        if elapsed:
            lines.append(f"Elapsed: `{elapsed}`")
        lines.append(f"Phase: `{phase}`")
        total = be_st.get("total")
        if total is not None:
            lines.append(f"Total records: `{total:,}`")
        total_pages  = be_st.get("total_pages")
        pages_done   = be_st.get("pages_done", 0)
        if total_pages is not None:
            lines.append(f"Pages: `{pages_done}/{total_pages}`")
        details_done  = be_st.get("details_done", 0)
        details_total = be_st.get("details_total")
        if details_total is not None:
            pct = int(details_done / details_total * 100) if details_total else 0
            lines.append(f"Details: `{details_done}/{details_total}` ({pct}%)")
        errors = be_st.get("errors", 0)
        if errors:
            lines.append(f"⚠️ Fetch errors: `{errors}`")
        rows = be_st.get("rows")
        if phase == "done" and rows is not None:
            lines.append(f"Rows exported: `{rows:,}`")
        error_msg = be_st.get("error_msg")
        if phase == "failed" and error_msg:
            lines.append(f"Error: `{error_msg}`")
        all_lines.extend(lines)

    if jd_st:
        if all_lines:
            all_lines.append("")   # blank separator
        phase        = jd_st.get("phase", "unknown")
        started_at   = jd_st.get("started_at")
        completed_at = jd_st.get("completed_at")
        icon         = phase_icons.get(phase, "⏳")
        lines        = [f"*{icon} Job Distribution Analysis*\n"]
        elapsed      = _elapsed(started_at, completed_at)
        if started_at:
            lines.append(f"Started: `{started_at.strftime('%H:%M:%S')}`")
        if elapsed:
            lines.append(f"Elapsed: `{elapsed}`")
        lines.append(f"Phase: `{phase}`")
        teams_count = jd_st.get("teams_count")
        if teams_count is not None:
            lines.append(f"Teams: `{teams_count}`")
        members_total = jd_st.get("members_total")
        if members_total is not None:
            lines.append(f"Members: `{members_total}`")
        tasks_total = jd_st.get("tasks_total")
        tasks_done  = jd_st.get("tasks_done", 0)
        if tasks_total is not None:
            pct = int(tasks_done / tasks_total * 100) if tasks_total else 0
            lines.append(f"Tasks processed: `{tasks_done}/{tasks_total}` ({pct}%)")
        errors = jd_st.get("errors", 0)
        if errors:
            lines.append(f"⚠️ Fetch errors: `{errors}`")
        if phase == "done":
            lines.append(f"Members in report: `{jd_st.get('rows', 0):,}`")
        error_msg = jd_st.get("error_msg")
        if phase == "failed" and error_msg:
            lines.append(f"Error: `{error_msg}`")
        all_lines.extend(lines)

    await update.message.reply_text(
        "\n".join(all_lines),
        parse_mode="Markdown",
        reply_markup=_main_menu(),
    )


async def recv_be_county(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    query = update.callback_query
    await query.answer()
    county = query.data.split(":")[1]
    sess   = _get_be_sess(ctx)
    sess.county     = county
    sess.registries = _BE_COUNTY_REGISTRIES.get(county, [county])
    label           = _BE_COUNTY_LABELS.get(county, county.title())
    reg_list        = ", ".join(sess.registries)
    await query.edit_message_text(
        f"✅ County: *{label}*\nRegistries: `{reg_list}`\n\n"
        "📧 Enter the email address to receive the file, or send `skip` to get it only via Telegram:",
        parse_mode="Markdown",
    )
    return BE.EMAIL


async def recv_be_email(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    text = (update.message.text or "").strip()
    sess = _get_be_sess(ctx)

    if text.lower() == "skip":
        sess.email = ""
    else:
        if "@" not in text or "." not in text.split("@")[-1]:
            await update.message.reply_text(
                "❌ Invalid email. Enter a valid address or send `skip`.",
                parse_mode="Markdown",
            )
            return BE.EMAIL
        sess.email = text

    await update.message.reply_text(
        "🔁 *How often should this report run?*",
        parse_mode="Markdown",
        reply_markup=_be_schedule_keyboard(),
    )
    return BE.SCHEDULE


async def recv_be_schedule(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    query = update.callback_query
    await query.answer()
    secs = int(query.data.split(":")[1])
    sess = _get_be_sess(ctx)
    sess.schedule_seconds = secs

    kbd = _be_cred_keyboard()
    if not kbd:
        await query.edit_message_text(
            "❌ No valid cached tokens. Use *🔑 Refresh Auth* first.",
            parse_mode="Markdown",
        )
        await ctx.bot.send_message(query.message.chat_id, "Main menu.", reply_markup=_main_menu())
        return ConversationHandler.END

    await query.edit_message_text(
        "👤 *Select the account to run the export as:*",
        parse_mode="Markdown",
        reply_markup=kbd,
    )
    return BE.PICK_CRED


async def recv_be_cred(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    query = update.callback_query
    await query.answer()
    sess           = _get_be_sess(ctx)
    sess.cred_type = query.data.split(":")[1]

    cred_label   = CRED_LABELS.get(sess.cred_type, sess.cred_type)
    county_label = _BE_COUNTY_LABELS.get(sess.county, sess.county.title())
    reg_list     = ", ".join(sess.registries)
    secs         = sess.schedule_seconds
    sched_label  = next((l for l, s in _BE_SCHEDULE_OPTIONS if s == secs), "Run Once")
    if secs > 0:
        sched_label += " (repeating)"
    email_label  = sess.email or "Telegram only"

    await query.edit_message_text(
        f"✅ Ready to export.\n\n"
        f"• County: *{county_label}*\n"
        f"• Registries: `{reg_list}`\n"
        f"• Filter: *Completed*\n"
        f"• Schedule: *{sched_label}*\n"
        f"• Account: *{cred_label}*\n"
        f"• Destination: *{email_label}*\n\n"
        "Tap *Run Export* to start.",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("▶️ Run Export", callback_data="be:yes"),
            InlineKeyboardButton("❌ Cancel",     callback_data="be:no"),
        ]]),
    )
    return BE.CONFIRM


async def recv_be_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not allowed(update): return await deny(update)
    query = update.callback_query
    await query.answer()

    if query.data == "be:no":
        await query.edit_message_text("❌ Export cancelled.")
        await ctx.bot.send_message(query.message.chat_id, "Main menu.", reply_markup=_main_menu())
        return ConversationHandler.END

    sess   = _get_be_sess(ctx)
    tokens = get_valid_tokens(sess.cred_type)
    if not tokens:
        cred_label = CRED_LABELS.get(sess.cred_type, sess.cred_type)
        await query.edit_message_text(
            f"❌ Tokens for *{cred_label}* have expired. Use *🔑 Refresh Auth* first.",
            parse_mode="Markdown",
        )
        await ctx.bot.send_message(query.message.chat_id, "Main menu.", reply_markup=_main_menu())
        return ConversationHandler.END

    chat_id = query.message.chat_id
    secs    = sess.schedule_seconds

    if secs > 0:
        # Save schedule and register repeating job
        cfg = {
            "chat_id":          chat_id,
            "county":           sess.county,
            "registries":       sess.registries,
            "email":            sess.email,
            "interval_seconds": secs,
            "cred_type":        sess.cred_type,
        }
        save_be_schedule(cfg)
        # Remove any existing job before adding a new one
        current_jobs = ctx.application.job_queue.get_jobs_by_name("bulk_export_job")
        for job in current_jobs:
            job.schedule_removal()
        ctx.application.job_queue.run_repeating(
            _bulk_export_job,
            interval=secs,
            first=0,   # run immediately then repeat
            name="bulk_export_job",
        )
        label = next((l for l, s in _BE_SCHEDULE_OPTIONS if s == secs), "repeating")
        await query.edit_message_text(
            f"⏳ Export started and scheduled to repeat *{label}*.\n"
            "You will be notified each time it completes.",
            parse_mode="Markdown",
        )
    else:
        await query.edit_message_text("⏳ Export running in background — you will be notified when done.")
        loop = asyncio.get_event_loop()
        asyncio.ensure_future(
            asyncio.to_thread(_bulk_export_run, tokens, chat_id, sess.email, ctx.bot, loop, sess.registries, sess.county)
        )

    await ctx.bot.send_message(chat_id, "Returning to menu.", reply_markup=_main_menu())
    return ConversationHandler.END


# ──────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────
async def _post_init(app) -> None:
    _restore_schedules(app)


def main():
    app = Application.builder().token(BOT_TOKEN).post_init(_post_init).build()

    # Text filter that excludes the cancel button (so it reaches fallbacks)
    not_cancel = filters.TEXT & ~filters.COMMAND & ~_CANCEL_FILTER

    conv = ConversationHandler(
        entry_points=[
            CommandHandler("assign", cmd_assign),
            MessageHandler(filters.Regex(f"^{re.escape(BTN_ASSIGN)}$"), cmd_assign),
        ],
        states={
            S.INPUT_METHOD:       [CallbackQueryHandler(recv_input_method, pattern=r"^input:")],
            S.REF_NUMBERS:        [MessageHandler(not_cancel, recv_refs)],
            S.RECV_PHOTOS:        [
                MessageHandler(filters.PHOTO, recv_photo),
                CallbackQueryHandler(recv_photo_done, pattern=r"^photo:done$"),
            ],
            S.CONFIRM_REFS:       [CallbackQueryHandler(recv_confirm_refs, pattern=r"^refs:")],
            S.REASSIGN_CONFIRM:   [CallbackQueryHandler(recv_reassign_confirm, pattern=r"^reassign:")],
            S.PICK_VALUER_SOURCE: [CallbackQueryHandler(recv_valuer_source, pattern=r"^src:")],
            S.VALUER_NAME:        [MessageHandler(not_cancel, recv_valuer_name)],
            S.CHOOSE_CRED:        [CallbackQueryHandler(recv_cred_choice, pattern=r"^cred:")],
            S.WAIT_OTP:           [MessageHandler(not_cancel, recv_otp)],
            S.SELECT_VALUER:      [CallbackQueryHandler(recv_valuer_select, pattern=r"^valuer:")],
            S.CONFIRM:            [CallbackQueryHandler(recv_confirm, pattern=r"^confirm:")],
        },
        fallbacks=[
            CommandHandler("cancel", cmd_cancel),
            MessageHandler(_CANCEL_FILTER, cmd_cancel),
            MessageHandler(filters.TEXT, fallback),
        ],
        allow_reentry=True,
        per_message=False,
    )

    db_conv = ConversationHandler(
        entry_points=[
            CommandHandler("dlvbatch", cmd_dlv_batch),
            MessageHandler(filters.Regex(f"^{re.escape(BTN_DLV_BATCH)}$"), cmd_dlv_batch),
        ],
        states={
            DB.INPUT_BATCH:   [MessageHandler(not_cancel, recv_db_input)],
            DB.CONFIRM_BATCH: [CallbackQueryHandler(recv_db_confirm, pattern=r"^db:")],
        },
        fallbacks=[
            CommandHandler("cancel", cmd_cancel),
            MessageHandler(_CANCEL_FILTER, cmd_cancel),
            MessageHandler(filters.TEXT, fallback),
        ],
        allow_reentry=True,
        per_message=False,
    )

    auth_conv = ConversationHandler(
        entry_points=[
            CommandHandler("auth", cmd_auth),
            MessageHandler(filters.Regex(f"^{re.escape(BTN_AUTH)}$"), cmd_auth),
        ],
        states={
            AS.CHOOSE_CRED:   [CallbackQueryHandler(recv_auth_cred,  pattern=r"^auth_cred:")],
            AS.FORCE_CONFIRM: [CallbackQueryHandler(recv_auth_force, pattern=r"^auth_force:")],
            AS.WAIT_OTP:      [MessageHandler(not_cancel, recv_auth_otp)],
        },
        fallbacks=[
            CommandHandler("cancel", cmd_cancel),
            MessageHandler(_CANCEL_FILTER, cmd_cancel),
            MessageHandler(filters.TEXT, fallback),
        ],
        allow_reentry=True,
        per_message=False,
    )

    app.add_handler(CommandHandler("start",         cmd_start))
    app.add_handler(CommandHandler("help",          cmd_help))
    app.add_handler(CommandHandler("valuers",       cmd_valuers))
    app.add_handler(CommandHandler("delete_valuer", cmd_delete_valuer))
    app.add_handler(CommandHandler("schedules",     cmd_schedules))
    app.add_handler(CommandHandler("task_batches",  cmd_task_batches))
    app.add_handler(CommandHandler("daemon",        cmd_daemon))
    fetch_conv = ConversationHandler(
        entry_points=[
            CommandHandler("fetch", cmd_fetch_tasks),
            MessageHandler(filters.Regex(f"^{re.escape(BTN_FETCH_TASKS)}$"), cmd_fetch_tasks),
        ],
        states={
            FT.CHOOSE_CRED:   [CallbackQueryHandler(recv_ft_cred,           pattern=r"^cred:")],
            FT.WAIT_OTP:      [MessageHandler(not_cancel, recv_ft_otp)],
            FT.DAYS_BACK:     [
                CallbackQueryHandler(recv_ft_days_callback, pattern=r"^ft_days:"),
                MessageHandler(not_cancel, recv_ft_days_text),
            ],
            FT.COUNTY_FILTER:   [CallbackQueryHandler(recv_ft_county_filter,   pattern=r"^ft_county:")],
            FT.REGISTRY_FILTER: [CallbackQueryHandler(recv_ft_registry_filter, pattern=r"^ft_registry:")],
            FT.AMOUNT_FILTER:    [CallbackQueryHandler(recv_ft_amount_filter,    pattern=r"^ft_amount:")],
            FT.AMOUNT_TEXT:      [MessageHandler(not_cancel, recv_ft_amount_text)],
            FT.SECTIONAL_FILTER: [CallbackQueryHandler(recv_ft_sectional_filter, pattern=r"^ft_sectional:")],
        },
        fallbacks=[
            CommandHandler("cancel", cmd_cancel),
            MessageHandler(_CANCEL_FILTER, cmd_cancel),
            MessageHandler(filters.TEXT, fallback),
        ],
        allow_reentry=True,
        per_message=False,
    )

    af_conv = ConversationHandler(
        entry_points=[
            CommandHandler("autofetch", cmd_auto_fetch),
            MessageHandler(filters.Regex(f"^{re.escape(BTN_AUTO_FETCH)}$"), cmd_auto_fetch),
        ],
        states={
            AF.INTERVAL:    [CallbackQueryHandler(recv_af_interval,     pattern=r"^af:")],
            AF.DAYS_BACK:   [CallbackQueryHandler(recv_af_days,         pattern=r"^af_days:")],
            AF.COUNTY:      [CallbackQueryHandler(recv_af_county,       pattern=r"^ft_county:")],
            AF.REGISTRY:    [CallbackQueryHandler(recv_af_registry,     pattern=r"^ft_registry:")],
            AF.AMOUNT:      [CallbackQueryHandler(recv_af_amount,       pattern=r"^ft_amount:")],
            AF.AMOUNT_TEXT: [MessageHandler(not_cancel, recv_af_amount_text)],
            AF.SECTIONAL:   [CallbackQueryHandler(recv_af_sectional,    pattern=r"^ft_sectional:")],
            AF.EMAIL:       [MessageHandler(not_cancel, recv_af_email)],
        },
        fallbacks=[
            CommandHandler("cancel", cmd_cancel),
            MessageHandler(_CANCEL_FILTER, cmd_cancel),
            MessageHandler(filters.TEXT, fallback),
        ],
        allow_reentry=True,
        per_message=False,
    )

    rs_conv = ConversationHandler(
        entry_points=[
            CommandHandler("receive", cmd_receive),
        ],
        states={
            RS.PICK_STAFF_SOURCE: [CallbackQueryHandler(recv_rt_pick_source,   pattern=r"^rt_src:")],
            RS.STAFF_NAME:        [MessageHandler(not_cancel, recv_rt_staff_name)],
            RS.SELECT_STAFF:      [CallbackQueryHandler(recv_rt_select_staff,  pattern=r"^rt_staff:")],
            RS.CHOOSE_CRED:       [CallbackQueryHandler(recv_rt_cred_choice,   pattern=r"^cred:")],
            RS.WAIT_OTP:          [MessageHandler(not_cancel, recv_rt_otp)],
            RS.TASK_TYPE:         [CallbackQueryHandler(recv_rt_task_type,     pattern=r"^rt_type:")],
            RS.TASK_COUNT:        [MessageHandler(not_cancel, recv_rt_task_count)],
            RS.AMOUNT_RANGE:      [CallbackQueryHandler(recv_rt_amount_choice, pattern=r"^ft_amount:")],
            RS.AMOUNT_TEXT:       [MessageHandler(not_cancel, recv_rt_amount_range)],
            RS.SCHEDULE_CHOICE:   [CallbackQueryHandler(recv_rt_schedule_choice, pattern=r"^rt_sched:")],
            RS.SCHEDULE_INTERVAL: [MessageHandler(not_cancel, recv_rt_schedule_interval)],
            RS.RT_CONFIRM:        [CallbackQueryHandler(recv_rt_confirm,       pattern=r"^rt_confirm:")],
        },
        fallbacks=[
            CommandHandler("cancel", cmd_cancel),
            MessageHandler(_CANCEL_FILTER, cmd_cancel),
            MessageHandler(filters.TEXT, fallback),
        ],
        allow_reentry=True,
        per_message=False,
    )

    be_conv = ConversationHandler(
        entry_points=[
            CommandHandler("bulkexport", cmd_bulk_export),
            MessageHandler(filters.Regex(f"^{re.escape(BTN_BULK_EXPORT)}$"), cmd_bulk_export),
        ],
        states={
            BE.COUNTY:    [CallbackQueryHandler(recv_be_county,    pattern=r"^be_county:")],
            BE.EMAIL:     [MessageHandler(not_cancel, recv_be_email)],
            BE.SCHEDULE:  [CallbackQueryHandler(recv_be_schedule,  pattern=r"^be_sched:")],
            BE.PICK_CRED: [CallbackQueryHandler(recv_be_cred,      pattern=r"^be_cred:")],
            BE.CONFIRM:   [CallbackQueryHandler(recv_be_confirm,   pattern=r"^be:")],
        },
        fallbacks=[
            CommandHandler("cancel", cmd_cancel),
            MessageHandler(_CANCEL_FILTER, cmd_cancel),
            MessageHandler(filters.TEXT, fallback),
        ],
        allow_reentry=True,
        per_message=False,
    )

    jd_conv = ConversationHandler(
        entry_points=[
            CommandHandler("jobdist", cmd_job_distribution),
            MessageHandler(filters.Regex(f"^{re.escape(BTN_JOB_DIST)}$"), cmd_job_distribution),
        ],
        states={
            JD.PICK_CRED: [CallbackQueryHandler(recv_jd_cred,     pattern=r"^be_cred:")],
            JD.CONFIRM:   [CallbackQueryHandler(recv_jd_confirm,  pattern=r"^jd:")],
        },
        fallbacks=[
            CommandHandler("cancel", cmd_cancel),
            MessageHandler(_CANCEL_FILTER, cmd_cancel),
            MessageHandler(filters.TEXT, fallback),
        ],
        allow_reentry=True,
        per_message=False,
    )

    app.add_handler(conv)
    app.add_handler(db_conv)
    app.add_handler(auth_conv)
    app.add_handler(fetch_conv)
    app.add_handler(af_conv)
    app.add_handler(rs_conv)
    app.add_handler(be_conv)
    app.add_handler(jd_conv)
    app.add_handler(MessageHandler(
        filters.Regex(f"^{re.escape(BTN_EXPORT_STATUS)}$"), cmd_export_status
    ))

    # DLV Batch: process queue every 5 minutes
    app.job_queue.run_repeating(_dlv_batch_job, interval=300, first=300, name="dlv_batch_job")

    # Bulk Export: restore saved schedule on startup
    be_cfg = load_be_schedule()
    if be_cfg and be_cfg.get("interval_seconds", 0) > 0:
        app.job_queue.run_repeating(
            _bulk_export_job,
            interval=be_cfg["interval_seconds"],
            first=be_cfg["interval_seconds"],
            name="bulk_export_job",
        )
        logger.info("Bulk export schedule restored: every %ds", be_cfg["interval_seconds"])

    # Auto Fetch: restore saved schedule on startup
    cfg = load_auto_fetch_schedule()
    if cfg:
        interval_secs = cfg.get("interval_minutes", 60) * 60
        app.job_queue.run_repeating(
            _auto_fetch_job,
            interval=interval_secs,
            first=interval_secs,
            name="auto_fetch_job",
        )
        logger.info("Auto Fetch schedule restored: every %d min", cfg.get("interval_minutes"))

    # Button handlers outside an active conversation
    app.add_handler(CallbackQueryHandler(recv_delete_valuer,  pattern=r"^del:"))
    app.add_handler(CallbackQueryHandler(recv_daemon_action,  pattern=r"^daemon:"))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_AUTH)}$"),         cmd_auth))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_TOKEN_STATUS)}$"), cmd_token_status))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_DAEMON)}$"),       cmd_daemon))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_RESTART)}$"),     cmd_restart))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_HELP)}$"),        cmd_help))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_VALUERS)}$"),     cmd_valuers))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_DELETE)}$"),      cmd_delete_valuer))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_DLV_TASKS)}$"),   cmd_dlv_tasks))
    app.add_handler(CallbackQueryHandler(recv_dlv_check,       pattern=r"^dlv_ck:"))
    app.add_handler(CallbackQueryHandler(recv_dlv_queue_action, pattern=r"^dlvq:"))
    app.add_handler(CallbackQueryHandler(recv_ta_valuer,        pattern=r"^ta:"))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_DLV_QUEUE)}$"),    cmd_dlv_queue))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_ASSIGNMENTS)}$"), cmd_assignments))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_AF_RESULTS)}$"), cmd_af_results))
    app.add_handler(CallbackQueryHandler(recv_af_result_detail, pattern=r"^af_result:"))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(BTN_CANCEL)}$"),      cmd_cancel))
    # Lowest-priority: catch valuer name text input during task-assign search
    app.add_handler(
        MessageHandler(filters.TEXT & ~filters.COMMAND & ~_MENU_BUTTON_FILTER, handle_ta_search),
        group=1,
    )

    logger.info("Bot started. Polling for updates…")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
